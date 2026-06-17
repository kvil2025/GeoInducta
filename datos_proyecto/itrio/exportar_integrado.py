"""
EXPORTAR BD INTEGRADA v2 — Actualización 10 junio 2026
======================================================
Fuentes:
  1. BD_Ytrio_LIMPIO.csv (datos historicos, sin modificar)
  2. BD_GEOL_2026 (1).xlsx (campaña mayo 2026, sin modificar)
  3. Datos de muestreo 10.06.xlsx (pXRF ACTUALIZADO, 251 muestras)
     - Hoja: '2026 06 10', 763 filas
     - Columnas: Y=6, Ce=120, La=118, Nd=124, Th=184, Fe=58, Ti=50
  4. BD_GEOL_2026_06_09.xls (coordenadas, sin modificar)

Criterio pXRF: se usa el valor MÁXIMO de cada réplica por muestra
(el usuario pidió "los valores más altos de cada muestra").

NO se inventan datos. Todo se documenta.
"""
import sys, os, shutil
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import numpy as np
import pyproj
import openpyxl
from collections import defaultdict

DATADIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"
OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\itrio"
DRIVE = r"G:\Mi unidad"
PXRF_FILE = r"C:\Users\geolo\Downloads\Datos de muestreo 11.06.xlsx"
COORDS_FILE = r"G:\Mi unidad\BD_GEOL_2026_06_10.xls"

transformer = pyproj.Transformer.from_crs('EPSG:32718', 'EPSG:4326', always_xy=True)

# Mapa de normalización de litologías
lit_map = {
    'Granito': 'Granito', 'Granito ': 'Granito',
    'Granito de Biotita': 'Granito de Biotita',
    'Granito de Biotita?': 'Granito de Biotita',
    'Granodiorita': 'Granodiorita',
    'Diorita': 'Diorita', 'Diotrita?': 'Diorita',
    'Diorita-Metapelita': 'Diorita', 'Diorita - Metapelita': 'Diorita',
    'Aplita': 'Aplita', 'Aplita?': 'Aplita',
    'Kaolinita': 'Kaolinita',
    'Sedimentaria': 'Sedimentaria', 'Sedimentario': 'Sedimentaria',
    'Metapelita': 'Metapelita',
}

# ══════════════════════════════════════════════════════════════════
# 1. BD_Ytrio (sin modificar)
# ══════════════════════════════════════════════════════════════════
df = pd.read_csv(os.path.join(DATADIR, "BD_Ytrio_LIMPIO.csv"))
df['FLAG_OUTLIER'] = df['FLAG_OUTLIER'].fillna('')
df['Litology_STD'] = df['Litology_STD'].fillna('SIN_ASIGNAR')
df['FLAG_DUPLICADO'] = df['FLAG_DUPLICADO'].fillna('')
df['FUENTE'] = 'BD_Ytrio'
df['HORIZONTE'] = ''
df['ROCA_CAJA'] = ''
df['CP'] = ''

cols = ['Sample','UTM_E','UTM_N','COTA_M','Y_ppm','Y_pond',
        'Ce_ppm','La_ppm','Th_ppm','Nd_ppm','Pr_ppm','Fe__','Ti__',
        'Litology_STD','FLAG_OUTLIER','FLAG_DUPLICADO','FUENTE','HORIZONTE','ROCA_CAJA','CP']
data1 = df[cols].dropna(subset=['UTM_E','UTM_N','Y_ppm']).copy()
data1 = data1[(data1['UTM_E'] > 100000) & (data1['UTM_N'] > 1000000)].copy()
data1 = data1.fillna('')

lons, lats = transformer.transform(data1['UTM_E'].values, data1['UTM_N'].values)
data1['lat'] = np.round(lats, 6)
data1['lon'] = np.round(lons, 6)
data1 = data1[(data1['lat'] > -60) & (data1['lat'] < -20) &
              (data1['lon'] > -80) & (data1['lon'] < -60)].copy()
print("1. BD_Ytrio: {} muestras (sin modificar)".format(len(data1)))

# ══════════════════════════════════════════════════════════════════
# 2. BD_GEOL_2026 (sin modificar)
# ══════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(os.path.join(DATADIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']

rows2 = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    cp = row[0]
    xm, ym = row[3], row[4]
    if xm is None or ym is None:
        continue
    horiz = row[7] or ''
    roca = row[8] or ''
    elev = row[2] or ''
    
    def safe(val):
        if val is None: return ''
        if isinstance(val, (int, float)): return val
        if isinstance(val, str) and val.strip().startswith('<'): return 0.5
        return ''
    
    yppm = safe(row[56])
    ypond = safe(row[57])
    fe = safe(row[35])
    ti = safe(row[27])
    
    lit_std = lit_map.get(roca, roca) if roca else 'SIN_ASIGNAR'
    sample = "{}".format(cp) if not horiz else "{}-{}".format(cp, horiz)
    lon, lat = transformer.transform(xm, ym)
    
    if not isinstance(yppm, (int, float)):
        continue
    
    rows2.append({
        'Sample': sample, 'UTM_E': xm, 'UTM_N': ym, 'COTA_M': elev,
        'Y_ppm': yppm, 'Y_pond': ypond, 'Ce_ppm': '', 'La_ppm': '', 'Th_ppm': '',
        'Nd_ppm': '', 'Pr_ppm': '', 'Fe__': fe, 'Ti__': ti,
        'Litology_STD': lit_std, 'FLAG_OUTLIER': '', 'FLAG_DUPLICADO': '',
        'FUENTE': 'BD_GEOL_2026', 'HORIZONTE': horiz, 'ROCA_CAJA': roca, 'CP': cp or '',
        'lat': round(lat, 6), 'lon': round(lon, 6),
    })

data2 = pd.DataFrame(rows2)
print("2. BD_GEOL_2026: {} muestras (sin modificar)".format(len(data2)))

# ══════════════════════════════════════════════════════════════════
# 3. pXRF ACTUALIZADO (Datos de muestreo 10.06.xlsx)
#    Criterio: valor MÁXIMO por muestra (no promedio)
# ══════════════════════════════════════════════════════════════════

# 3a. Cargar coordenadas + litología + Ypond
df_geo = pd.read_excel(COORDS_FILE, engine='xlrd', sheet_name='BD_29May26')
geo_lookup = {}
for _, gr in df_geo.iterrows():
    sid = str(int(gr['IDSAMPLE'])) if pd.notna(gr['IDSAMPLE']) else None
    if sid and pd.notna(gr['Xm']) and pd.notna(gr['Ym']):
        roca_raw = str(gr.get('ROCA CAJA', '')).strip() if pd.notna(gr.get('ROCA CAJA')) else ''
        roca_std = lit_map.get(roca_raw, roca_raw) if roca_raw else 'SIN_ASIGNAR'
        ypond_val = gr.get('Ypond', '')
        ypond_val = round(float(ypond_val), 1) if pd.notna(ypond_val) and ypond_val > 0 else ''
        
        geo_lookup[sid] = {
            'Xm': gr['Xm'], 'Ym': gr['Ym'],
            'CP': str(gr.get('CP', '')) if pd.notna(gr.get('CP')) else '',
            'HORIZONTE': str(gr.get('HORIZONTE', '')) if pd.notna(gr.get('HORIZONTE')) else '',
            'ROCA_CAJA': roca_raw, 'Litology_STD': roca_std,
            'Elevation': gr.get('Elevation', ''),
            'Ypond': ypond_val,
        }

# 3b. Leer pXRF — COLUMNAS del archivo 10.06:
#     Y=6, Ce=120, La=118, Nd=124, Th=184, Fe=58, Ti=50
wb2 = openpyxl.load_workbook(PXRF_FILE, data_only=True)
ws2 = wb2['2026 06 11']

elem_cols = {'Y_ppm': 6, 'Ce_ppm': 120, 'La_ppm': 118, 'Nd_ppm': 124, 'Th_ppm': 184, 'Fe__': 58, 'Ti__': 50}

# Acumular todas las réplicas por Sample ID base
pxrf_data = defaultdict(lambda: {e: [] for e in elem_cols})
total_rows = 0

for r in range(2, ws2.max_row + 1):
    sid_raw = ws2.cell(row=r, column=1).value
    if not sid_raw:
        continue
    sid = str(sid_raw).strip()
    base = sid.split('_')[0]
    total_rows += 1
    
    for elem, col in elem_cols.items():
        v = ws2.cell(row=r, column=col).value
        if v is not None and v != 'ND' and not str(v).startswith('<'):
            try:
                pxrf_data[base][elem].append(float(v))
            except:
                pass

print("\n3. pXRF ACTUALIZADO (Datos de muestreo 10.06.xlsx)")
print("   Archivo: {}".format(PXRF_FILE))
print("   Hoja: 2026 06 10")
print("   Filas leidas: {}".format(total_rows))
print("   Sample IDs unicos: {}".format(len(pxrf_data)))
print("   Criterio: VALOR MÁXIMO por muestra")

# 3c. Crear registros usando MÁXIMO
rows3 = []
pxrf_no_coords = 0

for base_id, elems in pxrf_data.items():
    if base_id in ('Ejrmplo', 'Prueba', 'ejemplo'):
        continue
    
    geo = geo_lookup.get(base_id)
    if geo is None:
        pxrf_no_coords += 1
        continue
    
    xm, ym = geo['Xm'], geo['Ym']
    if xm == -999 or ym == -999:
        pxrf_no_coords += 1
        continue
    
    lon, lat = transformer.transform(xm, ym)
    if not (-60 < lat < -20 and -80 < lon < -60):
        continue
    
    # Usar MÁXIMO (no promedio)
    rec = {
        'Sample': base_id, 'UTM_E': xm, 'UTM_N': ym,
        'COTA_M': geo.get('Elevation', ''),
    }
    if isinstance(rec['COTA_M'], (int, float)) and rec['COTA_M'] == -999:
        rec['COTA_M'] = ''
    
    has_any = False
    n_replicas = 0
    for elem, vals_list in elems.items():
        if vals_list:
            rec[elem] = round(max(vals_list), 1)  # ← MÁXIMO
            has_any = True
            n_replicas = max(n_replicas, len(vals_list))
        else:
            rec[elem] = ''
    
    if not has_any:
        continue
    
    # Limpiar NaN strings
    cp = geo.get('CP', '')
    horiz = geo.get('HORIZONTE', '')
    roca = geo.get('ROCA_CAJA', '')
    cp = cp if str(cp) != 'nan' else ''
    horiz = horiz if str(horiz) != 'nan' else ''
    roca = roca if str(roca) != 'nan' else ''
    
    rec.update({
        'Y_pond': geo.get('Ypond', ''),
        'Pr_ppm': '',
        'Litology_STD': geo.get('Litology_STD', 'SIN_ASIGNAR'),
        'FLAG_OUTLIER': '', 'FLAG_DUPLICADO': '',
        'FUENTE': 'pXRF_2026',
        'HORIZONTE': horiz, 'ROCA_CAJA': roca, 'CP': cp,
        'lat': round(lat, 6), 'lon': round(lon, 6),
        'N_REPLICAS': n_replicas,
    })
    rows3.append(rec)

data3 = pd.DataFrame(rows3)
print("   Con coordenadas: {}".format(len(data3)))
print("   Sin coordenadas (descartadas): {}".format(pxrf_no_coords))

# ══════════════════════════════════════════════════════════════════
# 4. COMBINAR TODO
# ══════════════════════════════════════════════════════════════════
all_cols = set()
for d in [data1, data2, data3]:
    all_cols.update(d.columns)
all_cols = sorted(all_cols)

for d in [data1, data2, data3]:
    for c in all_cols:
        if c not in d.columns:
            d[c] = ''

data_all = pd.concat([data1[all_cols], data2[all_cols], data3[all_cols]], ignore_index=True)
data_all = data_all.fillna('')

# Ordenar columnas
priority_cols = ['Sample', 'CP', 'FUENTE', 'UTM_E', 'UTM_N', 'lat', 'lon', 'COTA_M',
                 'Y_ppm', 'Y_pond', 'Ce_ppm', 'La_ppm', 'Th_ppm', 'Nd_ppm', 'Pr_ppm',
                 'Fe__', 'Ti__', 'Litology_STD', 'HORIZONTE', 'ROCA_CAJA',
                 'FLAG_OUTLIER', 'FLAG_DUPLICADO', 'N_REPLICAS']
ordered_cols = [c for c in priority_cols if c in data_all.columns]
remaining = [c for c in data_all.columns if c not in ordered_cols]
data_all = data_all[ordered_cols + remaining]

print("\n" + "=" * 60)
print("TOTAL INTEGRADO: {} muestras".format(len(data_all)))
print("  BD_Ytrio:     {:,}".format(len(data1)))
print("  BD_GEOL_2026: {:,}".format(len(data2)))
print("  pXRF_2026:    {:,} (MÁXIMO por muestra)".format(len(data3)))
print("  Columnas:     {}".format(len(data_all.columns)))

# ══════════════════════════════════════════════════════════════════
# 5. GUARDAR
# ══════════════════════════════════════════════════════════════════
csv_path = os.path.join(OUTDIR, "BD_INTEGRADA_2026.csv")
data_all.to_csv(csv_path, index=False, encoding='utf-8-sig')
print("\n✅ CSV: {} ({:.0f} KB)".format(csv_path, os.path.getsize(csv_path)/1024))

xlsx_path = os.path.join(OUTDIR, "BD_INTEGRADA_2026.xlsx")
with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
    data_all.to_excel(writer, sheet_name='BD_Integrada', index=False)
    
    # Hoja resumen
    dfs = {'BD_Ytrio': data1, 'BD_GEOL_2026': data2, 'pXRF_2026': data3, 'TOTAL': data_all}
    rows_sum = []
    for name, d in dfs.items():
        yp = d['Y_ppm'].apply(pd.to_numeric, errors='coerce')
        yw = d['Y_pond'].apply(pd.to_numeric, errors='coerce')
        rows_sum.append({
            'Fuente': name, 'N_Muestras': len(d),
            'Y_media': round(yp.mean(), 1) if yp.notna().any() else '',
            'Y_max': round(yp.max(), 1) if yp.notna().any() else '',
            'Y_pond_media': round(yw.mean(), 1) if yw.notna().any() else '',
            'N_con_Ypond': int(yw.notna().sum()),
            'Criterio_pXRF': 'MAX por muestra' if 'pXRF' in name else 'Original',
        })
    pd.DataFrame(rows_sum).to_excel(writer, sheet_name='Resumen', index=False)
    
    # Hoja litologías
    lit_counts = data_all.groupby(['FUENTE', 'Litology_STD']).size().reset_index(name='N')
    lit_counts.to_excel(writer, sheet_name='Litologias', index=False)

print("✅ XLSX: {} ({:.0f} KB)".format(xlsx_path, os.path.getsize(xlsx_path)/1024))

# Copiar a Drive
for src, dst in [(xlsx_path, os.path.join(DRIVE, "BD_INTEGRADA_2026.xlsx")),
                  (csv_path, os.path.join(DRIVE, "BD_INTEGRADA_2026.csv"))]:
    try:
        shutil.copy2(src, dst)
        print("✅ Drive: {}".format(dst))
    except Exception as e:
        print("⚠️ Error Drive: {}".format(e))

# Copiar pXRF a itrio para respaldo
dst_pxrf = os.path.join(OUTDIR, "Datos de muestreo 10.06.xlsx")
shutil.copy2(PXRF_FILE, dst_pxrf)
print("✅ Copia pXRF: {}".format(dst_pxrf))

print("\n🎉 Base de datos integrada ACTUALIZADA (criterio: MÁXIMO por muestra)")
