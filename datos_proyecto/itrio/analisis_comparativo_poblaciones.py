#!/usr/bin/env python3
"""
===========================================================================
ANÁLISIS COMPARATIVO DE DOS POBLACIONES DE DATOS GEOQUÍMICOS
---------------------------------------------------------------------------
Población 1: BD_Ytrio_LIMPIO.csv  (datos históricos, ~2907 muestras)
Población 2: BD_GEOL_2026 (1).xlsx (datos nuevos, ~30-40 muestras)

Compara ambas poblaciones bajo múltiples métodos de correlación y
análisis estadístico para datos geoquímicos.
===========================================================================
"""
import pandas as pd
import numpy as np
import openpyxl
import pyproj
import os
import sys
import warnings
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

from scipy import stats
from scipy.spatial.distance import mahalanobis
from scipy.cluster.hierarchy import linkage, fcluster
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
import matplotlib.ticker as mticker

# ─── Configuración ────────────────────────────────────────────────
OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"
plt.rcParams.update({
    'figure.dpi': 150,
    'savefig.dpi': 150,
    'font.size': 9,
    'axes.titlesize': 11,
    'axes.labelsize': 9,
    'figure.facecolor': 'white'
})

# ─── 1. CARGAR DATOS ──────────────────────────────────────────────
print("=" * 70)
print("  CARGANDO DATOS")
print("=" * 70)

# Población 1: BD_Ytrio_LIMPIO
df1 = pd.read_csv(os.path.join(OUTDIR, "BD_Ytrio_LIMPIO.csv"))
print(f"Población 1 (BD_Ytrio_LIMPIO): {len(df1)} muestras, {len(df1.columns)} columnas")

# Población 2: BD_GEOL_2026
wb = openpyxl.load_workbook(os.path.join(OUTDIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']

# Leer headers
headers = [cell.value for cell in ws[1]]
print(f"  Headers BD_GEOL: {len(headers)} columnas")

# Leer filas
data2 = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    data2.append(row)
df2_raw = pd.DataFrame(data2, columns=headers)
print(f"Población 2 (BD_GEOL_2026): {len(df2_raw)} filas totales")

# ─── 2. MAPEAR COLUMNAS COMUNES ──────────────────────────────────
# Identificar las columnas geoquímicas comunes
print("\n--- Mapeando columnas ---")
print(f"Columnas Pob1: {list(df1.columns)}")
print(f"\nColumnas Pob2: {list(df2_raw.columns)}")

# Mapeo de columnas entre ambas fuentes
# Pob1 tiene: Y_ppm, Ce_ppm, La_ppm, Pr_ppm, Nd_ppm, Th_ppm, Fe__, Ti__, K__, Ca_ppm, V_ppm, Cr_ppm, Mn_, Cl_ppm
# Pob2 tiene: K(%), Ca(%), Ti(%), V(ppm), Cr(ppm), Mn(%), Fe(%), Y(ppm), etc.

# Extraer datos numéricos de Pob2
def safe_float(val):
    """Convierte valor a float, tratando < como 0.5"""
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip()
        if val.startswith('<'):
            return 0.5
        try:
            return float(val.replace(',', '.'))
        except:
            return np.nan
    return np.nan

# Buscar las columnas por nombre parcial en df2_raw
def find_col(df, pattern):
    """Buscar columna por patrón"""
    for c in df.columns:
        if c is not None and pattern.lower() in str(c).lower():
            return c
    return None

# Identificar columnas de Pob2
print("\n--- Buscando columnas en Pob2 ---")
for c in df2_raw.columns:
    if c is not None:
        non_null = df2_raw[c].dropna()
        if len(non_null) > 0:
            print(f"  {str(c):30s} | {len(non_null):4d} vals | ejemplo: {non_null.iloc[0]}")

# Crear df2 con las columnas mapeadas
# Coordenadas
col_x = None
col_y = None
col_elev = None
for c in df2_raw.columns:
    if c is not None:
        cs = str(c).lower()
        if 'xm' in cs or cs == 'x' or 'este' in cs or 'utm_e' in cs:
            col_x = c
        if 'ym' in cs or cs == 'y' or 'norte' in cs or 'utm_n' in cs:
            col_y = c
        if 'elev' in cs or 'cota' in cs or 'alt' in cs:
            col_elev = c

print(f"\nCoordenadas: X={col_x}, Y={col_y}, Elev={col_elev}")

# Extraer datos geoquímicos de Pob2 usando índices de columna
# Basado en procesar_bd_geol.py: K=col23, Ca=col25, Ti=col27, V=col29, Cr=col31, Mn=col33, Fe=col35
# Y_ppm=col56

df2 = pd.DataFrame()
# Coordenadas
if col_x:
    df2['UTM_E'] = df2_raw[col_x].apply(safe_float)
if col_y:
    df2['UTM_N'] = df2_raw[col_y].apply(safe_float)
if col_elev:
    df2['COTA_M'] = df2_raw[col_elev].apply(safe_float)

# Geoquimica por posición (basado en procesar_bd_geol.py)
if len(headers) > 56:
    col_indices = {
        'K__': 23, 'Ca_ppm': 25, 'Ti__': 27, 'V_ppm': 29,
        'Cr_ppm': 31, 'Mn_': 33, 'Fe__': 35, 'Y_ppm': 56, 'Y_pond': 57
    }
    for name, idx in col_indices.items():
        if idx < len(headers):
            df2[name] = df2_raw.iloc[:, idx].apply(safe_float)
            non_null = df2[name].dropna().count()
            print(f"  Extraída {name} (col {idx}): {non_null} valores")
            # Convert % to ppm for Ca, V, Cr if present
            for col in ['Ca_ppm', 'V_ppm', 'Cr_ppm']:
                if col in df2.columns:
                    df2[col] = df2[col] * 10000

# Roca/Litología
col_roca = None
for c in df2_raw.columns:
    if c is not None and ('roca' in str(c).lower() or 'litol' in str(c).lower()):
        col_roca = c
        break
# Usar índice 8 (basado en procesar_bd_geol.py)
if col_roca is None and len(headers) > 8:
    col_roca_idx = 8
    df2['Litology'] = df2_raw.iloc[:, col_roca_idx]
else:
    df2['Litology'] = df2_raw[col_roca] if col_roca else 'Desconocida'

# CP / Sample
if headers[0] is not None:
    df2['Sample'] = df2_raw.iloc[:, 0].astype(str) + '-' + df2_raw.iloc[:, 1].astype(str)

# Filtrar filas con al menos algún dato geoquímico
geochem_cols = ['K__', 'Ca_ppm', 'Ti__', 'V_ppm', 'Cr_ppm', 'Mn_', 'Fe__', 'Y_ppm']
geochem_available = [c for c in geochem_cols if c in df2.columns]
df2 = df2.dropna(subset=['UTM_E', 'UTM_N'], how='any')
df2_geochem = df2.dropna(subset=geochem_available, how='all').copy()

print(f"\nPob2 con coordenadas válidas: {len(df2)} filas")
print(f"Pob2 con algún dato geoquímico: {len(df2_geochem)} filas")

# ─── 3. VARIABLES COMUNES ────────────────────────────────────────
# Variables que existen en ambas poblaciones
common_vars = []
for var in geochem_available:
    if var in df1.columns:
        n1 = df1[var].dropna().count()
        n2 = df2_geochem[var].dropna().count()
        if n1 > 0 and n2 > 0:
            common_vars.append(var)
            print(f"  Variable común: {var} -> Pob1: {n1} vals, Pob2: {n2} vals")

print(f"\nVariables comunes para análisis: {common_vars}")

# ─── 4. ESTADÍSTICAS DESCRIPTIVAS COMPARATIVAS ───────────────────
print("\n" + "=" * 70)
print("  ANÁLISIS ESTADÍSTICO COMPARATIVO")
print("=" * 70)

results = []
for var in common_vars:
    v1 = df1[var].dropna().values
    v2 = df2_geochem[var].dropna().values
    if len(v2) < 2:
        continue
    
    r = {
        'Variable': var,
        'N_Pob1': len(v1), 'N_Pob2': len(v2),
        'Mean_Pob1': np.mean(v1), 'Mean_Pob2': np.mean(v2),
        'Median_Pob1': np.median(v1), 'Median_Pob2': np.median(v2),
        'Std_Pob1': np.std(v1, ddof=1), 'Std_Pob2': np.std(v2, ddof=1),
        'CV_Pob1': np.std(v1, ddof=1) / np.mean(v1) * 100 if np.mean(v1) != 0 else np.nan,
        'CV_Pob2': np.std(v2, ddof=1) / np.mean(v2) * 100 if np.mean(v2) != 0 else np.nan,
        'Min_Pob1': np.min(v1), 'Min_Pob2': np.min(v2),
        'Max_Pob1': np.max(v1), 'Max_Pob2': np.max(v2),
        'P25_Pob1': np.percentile(v1, 25), 'P25_Pob2': np.percentile(v2, 25),
        'P75_Pob1': np.percentile(v1, 75), 'P75_Pob2': np.percentile(v2, 75),
        'Skew_Pob1': stats.skew(v1), 'Skew_Pob2': stats.skew(v2),
        'Kurt_Pob1': stats.kurtosis(v1), 'Kurt_Pob2': stats.kurtosis(v2),
    }
    results.append(r)
    
    print(f"\n{'─' * 50}")
    print(f"  {var}")
    print(f"{'─' * 50}")
    print(f"  {'':20s} {'Pob1 (Ytrio)':>15s} {'Pob2 (GEOL)':>15s}")
    print(f"  {'N':20s} {r['N_Pob1']:>15d} {r['N_Pob2']:>15d}")
    print(f"  {'Media':20s} {r['Mean_Pob1']:>15.2f} {r['Mean_Pob2']:>15.2f}")
    print(f"  {'Mediana':20s} {r['Median_Pob1']:>15.2f} {r['Median_Pob2']:>15.2f}")
    print(f"  {'Desv. Std':20s} {r['Std_Pob1']:>15.2f} {r['Std_Pob2']:>15.2f}")
    print(f"  {'CV (%)':20s} {r['CV_Pob1']:>15.1f} {r['CV_Pob2']:>15.1f}")
    print(f"  {'Mín':20s} {r['Min_Pob1']:>15.2f} {r['Min_Pob2']:>15.2f}")
    print(f"  {'Máx':20s} {r['Max_Pob1']:>15.2f} {r['Max_Pob2']:>15.2f}")
    print(f"  {'Skewness':20s} {r['Skew_Pob1']:>15.3f} {r['Skew_Pob2']:>15.3f}")
    print(f"  {'Kurtosis':20s} {r['Kurt_Pob1']:>15.3f} {r['Kurt_Pob2']:>15.3f}")

# ─── 5. TESTS ESTADÍSTICOS ───────────────────────────────────────
print("\n" + "=" * 70)
print("  TESTS DE COMPARACIÓN DE POBLACIONES")
print("=" * 70)

test_results = []
for var in common_vars:
    v1 = df1[var].dropna().values
    v2 = df2_geochem[var].dropna().values
    if len(v2) < 3:
        continue
    
    tr = {'Variable': var}
    
    # 5a. Mann-Whitney U (no paramétrico, no asume normalidad)
    try:
        u_stat, u_pval = stats.mannwhitneyu(v1, v2, alternative='two-sided')
        tr['MannWhitney_U'] = u_stat
        tr['MannWhitney_p'] = u_pval
        tr['MannWhitney_sig'] = '***' if u_pval < 0.001 else '**' if u_pval < 0.01 else '*' if u_pval < 0.05 else 'ns'
    except:
        tr['MannWhitney_U'] = np.nan
        tr['MannWhitney_p'] = np.nan
        tr['MannWhitney_sig'] = 'error'
    
    # 5b. Kolmogorov-Smirnov (compara distribuciones)
    try:
        ks_stat, ks_pval = stats.ks_2samp(v1, v2)
        tr['KS_stat'] = ks_stat
        tr['KS_p'] = ks_pval
        tr['KS_sig'] = '***' if ks_pval < 0.001 else '**' if ks_pval < 0.01 else '*' if ks_pval < 0.05 else 'ns'
    except:
        tr['KS_stat'] = np.nan
        tr['KS_p'] = np.nan
        tr['KS_sig'] = 'error'
    
    # 5c. Welch's t-test (no asume varianzas iguales)
    try:
        t_stat, t_pval = stats.ttest_ind(v1, v2, equal_var=False)
        tr['Welch_t'] = t_stat
        tr['Welch_p'] = t_pval
        tr['Welch_sig'] = '***' if t_pval < 0.001 else '**' if t_pval < 0.01 else '*' if t_pval < 0.05 else 'ns'
    except:
        tr['Welch_t'] = np.nan
        tr['Welch_p'] = np.nan
        tr['Welch_sig'] = 'error'
    
    # 5d. Effect size (Cohen's d)
    try:
        pooled_std = np.sqrt(((len(v1)-1)*np.std(v1, ddof=1)**2 + (len(v2)-1)*np.std(v2, ddof=1)**2) / (len(v1)+len(v2)-2))
        if pooled_std > 0:
            tr['Cohen_d'] = (np.mean(v1) - np.mean(v2)) / pooled_std
        else:
            tr['Cohen_d'] = 0
        # Interpretar
        d_abs = abs(tr['Cohen_d'])
        tr['Effect_size'] = 'Insignificante' if d_abs < 0.2 else 'Pequeño' if d_abs < 0.5 else 'Mediano' if d_abs < 0.8 else 'Grande'
    except:
        tr['Cohen_d'] = np.nan
        tr['Effect_size'] = 'error'
    
    # 5e. Anderson-Darling test de normalidad
    try:
        ad1 = stats.anderson(v1, dist='norm')
        ad2 = stats.anderson(v2, dist='norm')
        tr['AD_Pob1_stat'] = ad1.statistic
        tr['AD_Pob2_stat'] = ad2.statistic
        tr['Normal_Pob1'] = 'Sí' if ad1.statistic < ad1.critical_values[2] else 'No'
        tr['Normal_Pob2'] = 'Sí' if ad2.statistic < ad2.critical_values[2] else 'No'
    except:
        tr['Normal_Pob1'] = 'error'
        tr['Normal_Pob2'] = 'error'
    
    # 5f. Levene test (igualdad de varianzas)
    try:
        lev_stat, lev_pval = stats.levene(v1, v2)
        tr['Levene_stat'] = lev_stat
        tr['Levene_p'] = lev_pval
        tr['Varianzas_iguales'] = 'Sí' if lev_pval > 0.05 else 'No'
    except:
        tr['Levene_stat'] = np.nan
        tr['Levene_p'] = np.nan
        tr['Varianzas_iguales'] = 'error'
    
    test_results.append(tr)
    
    print(f"\n{var}:")
    print(f"  Mann-Whitney U: U={tr.get('MannWhitney_U', 'N/A'):.0f}, p={tr.get('MannWhitney_p', 'N/A'):.4e} [{tr.get('MannWhitney_sig', '')}]")
    print(f"  Kolmogorov-Smirnov: D={tr.get('KS_stat', 'N/A'):.4f}, p={tr.get('KS_p', 'N/A'):.4e} [{tr.get('KS_sig', '')}]")
    print(f"  Welch's t-test: t={tr.get('Welch_t', 'N/A'):.3f}, p={tr.get('Welch_p', 'N/A'):.4e} [{tr.get('Welch_sig', '')}]")
    print(f"  Cohen's d: {tr.get('Cohen_d', 'N/A'):.3f} ({tr.get('Effect_size', '')})")
    print(f"  Normalidad: Pob1={tr.get('Normal_Pob1', '')}, Pob2={tr.get('Normal_Pob2', '')}")
    print(f"  Varianzas iguales (Levene): {tr.get('Varianzas_iguales', '')} (p={tr.get('Levene_p', 'N/A'):.4e})")

# ─── 6. CORRELACIONES INTERNAS ───────────────────────────────────
print("\n" + "=" * 70)
print("  MATRICES DE CORRELACIÓN INTERNAS")
print("=" * 70)

# Variables para correlación
corr_vars = [v for v in common_vars if v in df1.columns and v in df2_geochem.columns]
if len(corr_vars) >= 2:
    # 6a. Pearson
    print("\n--- Correlación de Pearson (Pob1) ---")
    corr1_pearson = df1[corr_vars].corr(method='pearson')
    print(corr1_pearson.round(3).to_string())
    
    print("\n--- Correlación de Pearson (Pob2) ---")
    corr2_pearson = df2_geochem[corr_vars].corr(method='pearson')
    print(corr2_pearson.round(3).to_string())
    
    # 6b. Spearman (rangos, más robusto para datos geoquímicos)
    print("\n--- Correlación de Spearman (Pob1) ---")
    corr1_spearman = df1[corr_vars].corr(method='spearman')
    print(corr1_spearman.round(3).to_string())
    
    print("\n--- Correlación de Spearman (Pob2) ---")
    corr2_spearman = df2_geochem[corr_vars].corr(method='spearman')
    print(corr2_spearman.round(3).to_string())
    
    # 6c. Kendall tau (más robusto con pocos datos)
    print("\n--- Correlación de Kendall Tau (Pob1) ---")
    corr1_kendall = df1[corr_vars].corr(method='kendall')
    print(corr1_kendall.round(3).to_string())
    
    print("\n--- Correlación de Kendall Tau (Pob2) ---")
    corr2_kendall = df2_geochem[corr_vars].corr(method='kendall')
    print(corr2_kendall.round(3).to_string())
    
    # 6d. Diferencia de correlaciones
    print("\n--- Diferencia de Correlaciones (Spearman Pob2 - Pob1) ---")
    diff_corr = corr2_spearman - corr1_spearman
    print(diff_corr.round(3).to_string())

# ─── 7. LOG-RATIOS Y CORRELACIONES COMPOSICIONALES ───────────────
print("\n" + "=" * 70)
print("  ANÁLISIS DE LOG-RATIOS (COMPOSICIONAL)")
print("=" * 70)

# Para datos geoquímicos composicionales, usar CLR (centered log-ratio)
def clr_transform(df, cols):
    """Centred log-ratio transform"""
    data = df[cols].copy()
    data = data.replace(0, np.nan)
    data = data.dropna()
    log_data = np.log(data)
    geo_mean = log_data.mean(axis=1)
    clr = log_data.subtract(geo_mean, axis=0)
    return clr

if len(corr_vars) >= 3:
    try:
        clr1 = clr_transform(df1, corr_vars)
        clr2 = clr_transform(df2_geochem, corr_vars)
        
        print(f"\nCLR Pob1: {len(clr1)} muestras válidas")
        print(f"CLR Pob2: {len(clr2)} muestras válidas")
        
        if len(clr2) >= 3:
            print("\n--- Correlación Spearman en espacio CLR (Pob1) ---")
            print(clr1.corr(method='spearman').round(3).to_string())
            
            print("\n--- Correlación Spearman en espacio CLR (Pob2) ---")
            print(clr2.corr(method='spearman').round(3).to_string())
    except Exception as e:
        print(f"  Error en CLR: {e}")

# ─── 8. ANÁLISIS DE RATIOS ELEMENTALES ───────────────────────────
print("\n" + "=" * 70)
print("  RATIOS ELEMENTALES")
print("=" * 70)

ratios_to_compute = []
if 'Y_ppm' in common_vars and 'Fe__' in common_vars:
    ratios_to_compute.append(('Y/Fe', 'Y_ppm', 'Fe__'))
if 'Y_ppm' in common_vars and 'Ti__' in common_vars:
    ratios_to_compute.append(('Y/Ti', 'Y_ppm', 'Ti__'))
if 'Y_ppm' in common_vars and 'K__' in common_vars:
    ratios_to_compute.append(('Y/K', 'Y_ppm', 'K__'))
if 'Fe__' in common_vars and 'Ti__' in common_vars:
    ratios_to_compute.append(('Fe/Ti', 'Fe__', 'Ti__'))
if 'Fe__' in common_vars and 'Mn_' in common_vars:
    ratios_to_compute.append(('Fe/Mn', 'Fe__', 'Mn_'))
if 'K__' in common_vars and 'Ti__' in common_vars:
    ratios_to_compute.append(('K/Ti', 'K__', 'Ti__'))
if 'V_ppm' in common_vars and 'Cr_ppm' in common_vars:
    ratios_to_compute.append(('V/Cr', 'V_ppm', 'Cr_ppm'))

ratio_results = []
for rname, num, den in ratios_to_compute:
    # Pob1
    mask1 = (df1[num] > 0) & (df1[den] > 0)
    r1 = (df1.loc[mask1, num] / df1.loc[mask1, den]).dropna()
    # Pob2
    mask2 = (df2_geochem[num] > 0) & (df2_geochem[den] > 0)
    r2 = (df2_geochem.loc[mask2, num] / df2_geochem.loc[mask2, den]).dropna()
    
    if len(r2) >= 2:
        rr = {
            'Ratio': rname,
            'Mean_Pob1': np.mean(r1), 'Mean_Pob2': np.mean(r2),
            'Median_Pob1': np.median(r1), 'Median_Pob2': np.median(r2),
            'Std_Pob1': np.std(r1, ddof=1), 'Std_Pob2': np.std(r2, ddof=1),
        }
        
        try:
            u_s, u_p = stats.mannwhitneyu(r1, r2, alternative='two-sided')
            rr['MW_p'] = u_p
            rr['MW_sig'] = '***' if u_p < 0.001 else '**' if u_p < 0.01 else '*' if u_p < 0.05 else 'ns'
        except:
            rr['MW_p'] = np.nan
            rr['MW_sig'] = 'error'
        
        ratio_results.append(rr)
        print(f"\n{rname}: Pob1 med={rr['Median_Pob1']:.3f}, Pob2 med={rr['Median_Pob2']:.3f}, MW p={rr.get('MW_p', 'N/A'):.4e} [{rr.get('MW_sig', '')}]")

# ─── 9. REGRESIONES Y FACTORES ───────────────────────────────────
print("\n" + "=" * 70)
print("  REGRESIONES ENTRE VARIABLES")
print("=" * 70)

# Para cada par de variables, comparar las pendientes de regresión
if len(corr_vars) >= 2:
    regression_comparisons = []
    for i, v1_name in enumerate(corr_vars):
        for v2_name in corr_vars[i+1:]:
            # Pob1
            mask1 = df1[[v1_name, v2_name]].dropna()
            if len(mask1) > 2:
                slope1, intercept1, r1, p1, se1 = stats.linregress(mask1[v1_name], mask1[v2_name])
            else:
                continue
            # Pob2
            mask2 = df2_geochem[[v1_name, v2_name]].dropna()
            if len(mask2) > 2:
                slope2, intercept2, r2, p2, se2 = stats.linregress(mask2[v1_name], mask2[v2_name])
            else:
                continue
            
            rc = {
                'X': v1_name, 'Y_var': v2_name,
                'Slope_Pob1': slope1, 'R2_Pob1': r1**2,
                'Slope_Pob2': slope2, 'R2_Pob2': r2**2,
                'Slope_diff': slope2 - slope1,
                'R2_diff': r2**2 - r1**2,
            }
            regression_comparisons.append(rc)
            print(f"  {v1_name} vs {v2_name}: slope1={slope1:.4f} (R²={r1**2:.3f}), slope2={slope2:.4f} (R²={r2**2:.3f})")
    
    if regression_comparisons:
        df_reg = pd.DataFrame(regression_comparisons)
        print(f"\nMayor divergencia de pendiente: {df_reg.loc[df_reg['Slope_diff'].abs().idxmax()]}")

# ─── 10. PERCENTIL COMPARATIVO ────────────────────────────────────
print("\n" + "=" * 70)
print("  POSICIÓN PERCENTIL DE POB2 EN POB1")
print("=" * 70)

for var in common_vars:
    v1 = df1[var].dropna().values
    v2 = df2_geochem[var].dropna().values
    if len(v2) < 1:
        continue
    
    # ¿En qué percentil de Pob1 cae la mediana de Pob2?
    med2 = np.median(v2)
    pct = stats.percentileofscore(v1, med2)
    print(f"  {var}: Mediana Pob2 ({med2:.2f}) → percentil {pct:.1f}% en Pob1")

# ═══════════════════════════════════════════════════════════════════
# FIGURAS
# ═══════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("  GENERANDO FIGURAS")
print("=" * 70)

# ─── FIGURA 1: Boxplots comparativos ─────────────────────────────
n_vars = len(common_vars)
if n_vars >= 1:
    fig, axes = plt.subplots(2, (n_vars + 1) // 2, figsize=(14, 8))
    axes = axes.flatten() if n_vars > 1 else [axes]
    fig.suptitle('Boxplots Comparativos: Población 1 (Ytrio) vs Población 2 (GEOL)', fontsize=13, fontweight='bold')
    
    for idx, var in enumerate(common_vars):
        if idx >= len(axes):
            break
        ax = axes[idx]
        v1 = df1[var].dropna().values
        v2 = df2_geochem[var].dropna().values
        
        bp = ax.boxplot([v1, v2], labels=['Pob1\n(Ytrio)', 'Pob2\n(GEOL)'], 
                       patch_artist=True, widths=0.6,
                       medianprops=dict(color='red', linewidth=2))
        bp['boxes'][0].set_facecolor('#4A90D9')
        bp['boxes'][0].set_alpha(0.6)
        bp['boxes'][1].set_facecolor('#E67E22')
        bp['boxes'][1].set_alpha(0.6)
        ax.set_title(var, fontweight='bold')
        ax.grid(True, alpha=0.3)
        
        # Añadir n
        ax.text(1, ax.get_ylim()[1]*0.95, f'n={len(v1)}', ha='center', fontsize=7, color='#4A90D9')
        ax.text(2, ax.get_ylim()[1]*0.95, f'n={len(v2)}', ha='center', fontsize=7, color='#E67E22')
    
    for idx in range(len(common_vars), len(axes)):
        axes[idx].set_visible(False)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig01_boxplots.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig01_boxplots.png")

# ─── FIGURA 2: Histogramas superpuestos ──────────────────────────
if n_vars >= 1:
    fig, axes = plt.subplots(2, (n_vars + 1) // 2, figsize=(14, 8))
    axes = axes.flatten() if n_vars > 1 else [axes]
    fig.suptitle('Histogramas Superpuestos (Densidad)', fontsize=13, fontweight='bold')
    
    for idx, var in enumerate(common_vars):
        if idx >= len(axes):
            break
        ax = axes[idx]
        v1 = df1[var].dropna().values
        v2 = df2_geochem[var].dropna().values
        
        # Usar bins comunes
        all_vals = np.concatenate([v1, v2])
        bins = np.histogram_bin_edges(all_vals, bins=30)
        
        ax.hist(v1, bins=bins, density=True, alpha=0.5, color='#4A90D9', label=f'Pob1 (n={len(v1)})')
        ax.hist(v2, bins=bins, density=True, alpha=0.6, color='#E67E22', label=f'Pob2 (n={len(v2)})')
        ax.set_title(var, fontweight='bold')
        ax.legend(fontsize=7)
        ax.grid(True, alpha=0.3)
    
    for idx in range(len(common_vars), len(axes)):
        axes[idx].set_visible(False)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig02_histogramas.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig02_histogramas.png")

# ─── FIGURA 3: Matrices de correlación lado a lado ───────────────
if len(corr_vars) >= 2:
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    fig.suptitle('Matrices de Correlación Spearman: Pob1 vs Pob2 y Diferencia', fontsize=13, fontweight='bold')
    
    for ax, mat, title in zip(axes, [corr1_spearman, corr2_spearman, diff_corr],
                                ['Pob1 (Ytrio)', 'Pob2 (GEOL)', 'Diferencia (Pob2 - Pob1)']):
        im = ax.imshow(mat.values, cmap='RdBu_r', vmin=-1, vmax=1, aspect='auto')
        ax.set_xticks(range(len(corr_vars)))
        ax.set_yticks(range(len(corr_vars)))
        ax.set_xticklabels(corr_vars, rotation=45, ha='right', fontsize=7)
        ax.set_yticklabels(corr_vars, fontsize=7)
        ax.set_title(title, fontsize=10, fontweight='bold')
        
        # Anotar valores
        for i in range(len(corr_vars)):
            for j in range(len(corr_vars)):
                val = mat.iloc[i, j]
                color = 'white' if abs(val) > 0.6 else 'black'
                ax.text(j, i, f'{val:.2f}', ha='center', va='center', fontsize=6, color=color)
        
        plt.colorbar(im, ax=ax, shrink=0.8)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig03_correlaciones.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig03_correlaciones.png")

# ─── FIGURA 4: QQ-plots comparando distribuciones ────────────────
if n_vars >= 1:
    fig, axes = plt.subplots(2, (n_vars + 1) // 2, figsize=(14, 8))
    axes = axes.flatten() if n_vars > 1 else [axes]
    fig.suptitle('QQ-Plots: Pob2 vs Distribución Teórica de Pob1', fontsize=13, fontweight='bold')
    
    for idx, var in enumerate(common_vars):
        if idx >= len(axes):
            break
        ax = axes[idx]
        v1 = df1[var].dropna().values
        v2 = df2_geochem[var].dropna().values
        
        if len(v2) < 2:
            ax.set_visible(False)
            continue
        
        # QQ-plot: cuantiles de Pob2 vs cuantiles teóricos de Pob1
        quantiles = np.linspace(0.01, 0.99, min(len(v2), 50))
        q1 = np.quantile(v1, quantiles)
        q2 = np.quantile(v2, quantiles)
        
        ax.scatter(q1, q2, s=20, alpha=0.7, color='#E67E22', edgecolors='k', linewidths=0.5)
        lims = [min(q1.min(), q2.min()), max(q1.max(), q2.max())]
        ax.plot(lims, lims, 'k--', alpha=0.5, label='1:1')
        ax.set_xlabel('Cuantiles Pob1', fontsize=7)
        ax.set_ylabel('Cuantiles Pob2', fontsize=7)
        ax.set_title(var, fontweight='bold')
        ax.legend(fontsize=7)
        ax.grid(True, alpha=0.3)
    
    for idx in range(len(common_vars), len(axes)):
        axes[idx].set_visible(False)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig04_QQ.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig04_QQ.png")

# ─── FIGURA 5: CDF comparativas ─────────────────────────────────
if n_vars >= 1:
    fig, axes = plt.subplots(2, (n_vars + 1) // 2, figsize=(14, 8))
    axes = axes.flatten() if n_vars > 1 else [axes]
    fig.suptitle('Funciones de Distribución Acumulada (CDF)', fontsize=13, fontweight='bold')
    
    for idx, var in enumerate(common_vars):
        if idx >= len(axes):
            break
        ax = axes[idx]
        v1 = np.sort(df1[var].dropna().values)
        v2 = np.sort(df2_geochem[var].dropna().values)
        
        cdf1 = np.arange(1, len(v1)+1) / len(v1)
        cdf2 = np.arange(1, len(v2)+1) / len(v2)
        
        ax.step(v1, cdf1, color='#4A90D9', label=f'Pob1 (n={len(v1)})', linewidth=1.5)
        ax.step(v2, cdf2, color='#E67E22', label=f'Pob2 (n={len(v2)})', linewidth=2)
        ax.set_title(var, fontweight='bold')
        ax.set_ylabel('CDF')
        ax.legend(fontsize=7)
        ax.grid(True, alpha=0.3)
    
    for idx in range(len(common_vars), len(axes)):
        axes[idx].set_visible(False)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig05_CDF.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig05_CDF.png")

# ─── FIGURA 6: Scatter plots bivariados con ambas poblaciones ────
if len(corr_vars) >= 2:
    pairs = [(corr_vars[i], corr_vars[j]) for i in range(len(corr_vars)) for j in range(i+1, len(corr_vars))]
    n_pairs = min(len(pairs), 12)
    cols_fig = 4
    rows_fig = (n_pairs + cols_fig - 1) // cols_fig
    
    fig, axes = plt.subplots(rows_fig, cols_fig, figsize=(16, 4*rows_fig))
    if rows_fig > 1:
        axes = axes.flatten()
    elif cols_fig > 1:
        axes = axes.flatten()
    else:
        axes = [axes]
    fig.suptitle('Scatter Plots Bivariados: Pob1 (azul) vs Pob2 (naranja)', fontsize=13, fontweight='bold')
    
    for idx, (vx, vy) in enumerate(pairs[:n_pairs]):
        ax = axes[idx]
        # Pob1
        mask1 = df1[[vx, vy]].dropna()
        ax.scatter(mask1[vx], mask1[vy], s=3, alpha=0.15, color='#4A90D9', label='Pob1')
        # Pob2
        mask2 = df2_geochem[[vx, vy]].dropna()
        ax.scatter(mask2[vx], mask2[vy], s=30, alpha=0.8, color='#E67E22', edgecolors='black', linewidths=0.5, label='Pob2', zorder=5)
        
        ax.set_xlabel(vx, fontsize=7)
        ax.set_ylabel(vy, fontsize=7)
        ax.set_title(f'{vx} vs {vy}', fontsize=8, fontweight='bold')
        ax.legend(fontsize=6, markerscale=2)
        ax.grid(True, alpha=0.3)
    
    for idx in range(n_pairs, len(axes)):
        axes[idx].set_visible(False)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig06_scatters.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig06_scatters.png")

# ─── FIGURA 7: Mapa de ubicación de ambas poblaciones ────────────
fig, ax = plt.subplots(1, 1, figsize=(10, 8))
ax.set_title('Ubicación Espacial: Pob1 vs Pob2', fontsize=13, fontweight='bold')

# Pob1
ax.scatter(df1['UTM_E'], df1['UTM_N'], s=3, alpha=0.15, color='#4A90D9', label=f'Pob1 (n={len(df1)})')
# Pob2
if 'UTM_E' in df2_geochem.columns and 'UTM_N' in df2_geochem.columns:
    ax.scatter(df2_geochem['UTM_E'], df2_geochem['UTM_N'], s=40, alpha=0.8, color='#E67E22',
              edgecolors='black', linewidths=0.5, label=f'Pob2 (n={len(df2_geochem)})', zorder=5)
ax.set_xlabel('UTM_E')
ax.set_ylabel('UTM_N')
ax.legend()
ax.grid(True, alpha=0.3)
ax.set_aspect('equal')
plt.tight_layout()
fig.savefig(os.path.join(OUTDIR, 'comp_fig07_mapa.png'), bbox_inches='tight')
plt.close(fig)
print("  ✅ comp_fig07_mapa.png")

# ─── FIGURA 8: Resumen visual de tests ──────────────────────────
if test_results:
    df_tests = pd.DataFrame(test_results)
    
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle('Resumen de Tests Estadísticos', fontsize=13, fontweight='bold')
    
    # Panel izq: p-values
    ax = axes[0]
    vars_t = [t['Variable'] for t in test_results]
    mw_p = [t.get('MannWhitney_p', 1) for t in test_results]
    ks_p = [t.get('KS_p', 1) for t in test_results]
    welch_p = [t.get('Welch_p', 1) for t in test_results]
    
    x = np.arange(len(vars_t))
    w = 0.25
    ax.barh(x - w, [-np.log10(max(p, 1e-300)) for p in mw_p], w, label='Mann-Whitney', color='#4A90D9', alpha=0.7)
    ax.barh(x, [-np.log10(max(p, 1e-300)) for p in ks_p], w, label='KS', color='#E67E22', alpha=0.7)
    ax.barh(x + w, [-np.log10(max(p, 1e-300)) for p in welch_p], w, label='Welch t', color='#2ECC71', alpha=0.7)
    ax.axvline(-np.log10(0.05), color='red', linestyle='--', alpha=0.5, label='p=0.05')
    ax.axvline(-np.log10(0.01), color='red', linestyle=':', alpha=0.3, label='p=0.01')
    ax.set_yticks(x)
    ax.set_yticklabels(vars_t, fontsize=8)
    ax.set_xlabel('-log10(p-value)')
    ax.set_title('Significancia de Tests', fontweight='bold')
    ax.legend(fontsize=7)
    ax.grid(True, alpha=0.3, axis='x')
    
    # Panel der: Effect sizes
    ax = axes[1]
    cohens = [t.get('Cohen_d', 0) for t in test_results]
    colors = ['#E74C3C' if abs(c) > 0.8 else '#E67E22' if abs(c) > 0.5 else '#F1C40F' if abs(c) > 0.2 else '#2ECC71' for c in cohens]
    ax.barh(x, cohens, color=colors, alpha=0.7, edgecolor='k', linewidth=0.5)
    ax.axvline(0, color='k', linewidth=0.5)
    ax.axvline(-0.8, color='red', linestyle=':', alpha=0.3)
    ax.axvline(0.8, color='red', linestyle=':', alpha=0.3)
    ax.axvline(-0.5, color='orange', linestyle=':', alpha=0.3)
    ax.axvline(0.5, color='orange', linestyle=':', alpha=0.3)
    ax.set_yticks(x)
    ax.set_yticklabels(vars_t, fontsize=8)
    ax.set_xlabel("Cohen's d")
    ax.set_title('Tamaño del Efecto (Cohen d)', fontweight='bold')
    ax.grid(True, alpha=0.3, axis='x')
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig08_tests.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig08_tests.png")

# ─── FIGURA 9: Violin plots comparativos ─────────────────────────
if n_vars >= 1:
    fig, axes = plt.subplots(2, (n_vars + 1) // 2, figsize=(14, 8))
    axes = axes.flatten() if n_vars > 1 else [axes]
    fig.suptitle('Violin Plots Comparativos', fontsize=13, fontweight='bold')
    
    for idx, var in enumerate(common_vars):
        if idx >= len(axes):
            break
        ax = axes[idx]
        v1 = df1[var].dropna().values
        v2 = df2_geochem[var].dropna().values
        
        parts1 = ax.violinplot([v1], positions=[1], showmeans=True, showmedians=True)
        parts2 = ax.violinplot([v2], positions=[2], showmeans=True, showmedians=True)
        
        for pc in parts1['bodies']:
            pc.set_facecolor('#4A90D9')
            pc.set_alpha(0.5)
        for pc in parts2['bodies']:
            pc.set_facecolor('#E67E22')
            pc.set_alpha(0.5)
        
        ax.set_xticks([1, 2])
        ax.set_xticklabels(['Pob1\n(Ytrio)', 'Pob2\n(GEOL)'])
        ax.set_title(var, fontweight='bold')
        ax.grid(True, alpha=0.3)
    
    for idx in range(len(common_vars), len(axes)):
        axes[idx].set_visible(False)
    
    plt.tight_layout()
    fig.savefig(os.path.join(OUTDIR, 'comp_fig09_violin.png'), bbox_inches='tight')
    plt.close(fig)
    print("  ✅ comp_fig09_violin.png")

# ─── 11. GUARDAR RESULTADOS ──────────────────────────────────────
print("\n" + "=" * 70)
print("  GUARDANDO RESULTADOS")
print("=" * 70)

# Estadísticas
if results:
    df_stats = pd.DataFrame(results)
    df_stats.to_csv(os.path.join(OUTDIR, 'comp_estadisticas_descriptivas.csv'), index=False)
    print("  ✅ comp_estadisticas_descriptivas.csv")

# Tests
if test_results:
    df_tests = pd.DataFrame(test_results)
    df_tests.to_csv(os.path.join(OUTDIR, 'comp_tests_estadisticos.csv'), index=False)
    print("  ✅ comp_tests_estadisticos.csv")

# Ratios
if ratio_results:
    df_ratios = pd.DataFrame(ratio_results)
    df_ratios.to_csv(os.path.join(OUTDIR, 'comp_ratios_elementales.csv'), index=False)
    print("  ✅ comp_ratios_elementales.csv")

# ─── 12. RESUMEN EJECUTIVO ───────────────────────────────────────
print("\n" + "=" * 70)
print("  RESUMEN EJECUTIVO")
print("=" * 70)

print(f"""
COMPARACIÓN DE POBLACIONES GEOQUÍMICAS
=======================================

Población 1 (BD_Ytrio_LIMPIO): {len(df1)} muestras
Población 2 (BD_GEOL_2026):    {len(df2_geochem)} muestras con geoquímica
Variables comunes analizadas:   {len(common_vars)} ({', '.join(common_vars)})

MÉTODOS APLICADOS:
1. Estadísticas descriptivas (media, mediana, CV, skewness, kurtosis)
2. Test de Mann-Whitney U (no paramétrico, medianas)
3. Test de Kolmogorov-Smirnov (forma de distribución)
4. Welch's t-test (medias con varianzas desiguales)
5. Cohen's d (tamaño del efecto)
6. Test de normalidad Anderson-Darling
7. Test de Levene (homogeneidad de varianzas)
8. Correlaciones Pearson, Spearman, Kendall (internas)
9. Transformación CLR composicional
10. Ratios elementales y su comparación
11. Regresión lineal comparativa (pendientes)
12. Posición percentil cruzada
13. QQ-plots inter-poblacionales
14. CDFs comparativas
""")

# Variables con diferencias significativas
if test_results:
    sig_vars = [t['Variable'] for t in test_results if t.get('MannWhitney_p', 1) < 0.05]
    ns_vars = [t['Variable'] for t in test_results if t.get('MannWhitney_p', 1) >= 0.05]
    print(f"Variables con diferencias SIGNIFICATIVAS (Mann-Whitney p<0.05):")
    for v in sig_vars:
        t = [t for t in test_results if t['Variable'] == v][0]
        print(f"  ⚠️  {v}: p={t['MannWhitney_p']:.2e}, Cohen d={t.get('Cohen_d', 0):.2f} ({t.get('Effect_size', '')})")
    
    print(f"\nVariables SIN diferencia significativa:")
    for v in ns_vars:
        t = [t for t in test_results if t['Variable'] == v][0]
        print(f"  ✅ {v}: p={t['MannWhitney_p']:.2e}")

print("\n✅ Análisis comparativo completado exitosamente.")
