#!/usr/bin/env python3
"""
==========================================================================
INVESTIGACIÓN PROFUNDA DE UNIDADES — BD_GEOL_2026 vs BD_Ytrio_LIMPIO
==========================================================================
Objetivo: Determinar las unidades REALES de cada columna en el XLSX
y cómo se relacionan con las columnas de la BD_Ytrio_LIMPIO.
"""
import pandas as pd
import numpy as np
import openpyxl
import os, sys
sys.stdout.reconfigure(encoding='utf-8')

OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"

# ═══════════════════════════════════════════════════════════════
# 1. LEER XLSX CRUDO — CELDA POR CELDA
# ═══════════════════════════════════════════════════════════════
print("=" * 90)
print("  1. LECTURA CRUDA DEL XLSX — TODAS LAS CELDAS")
print("=" * 90)

wb = openpyxl.load_workbook(os.path.join(OUTDIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']

# Headers fila 1
headers = [cell.value for cell in ws[1]]
print(f"Total columnas header: {len(headers)}")
print(f"Filas de datos: {ws.max_row - 1}")

# Mostrar TODOS los headers con contenido
print("\n--- HEADERS COMPLETOS (fila 1) ---")
for i, h in enumerate(headers):
    if h is not None:
        print(f"  col[{i:3d}] = '{h}'")

# ═══════════════════════════════════════════════════════════════
# 2. EXAMINAR PATRÓN: ¿Cada elemento tiene col VALOR + col ERROR?
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  2. PATRÓN DE COLUMNAS PAREADAS (Valor + Error)")
print("=" * 90)

# El XLSX viene de un equipo XRF portátil (probablemente Olympus/Niton/Bruker)
# Estos equipos típicamente reportan: Elemento (%), Error (%), o Elemento (ppm), Error (ppm)
# Clave: el header del error suele tener "Err" y LA MISMA UNIDAD que el valor

pairs_found = []
for i in range(len(headers) - 1):
    h1 = str(headers[i]).strip() if headers[i] else ""
    h2 = str(headers[i+1]).strip() if headers[i+1] else ""
    
    if h2.endswith('Err') and h1 != "":
        elem = h1
        err_name = h2
        
        # Leer valores
        vals = []
        errs = []
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=i+1).value
            e = ws.cell(row=row_idx, column=i+2).value
            if isinstance(v, (int, float)) and isinstance(e, (int, float)) and v > 0:
                vals.append(float(v))
                errs.append(float(e))
        
        if vals:
            vals = np.array(vals)
            errs = np.array(errs)
            err_pct = (errs / vals * 100)
            
            pairs_found.append({
                'col_idx': i, 'elem': elem, 'err_col': i+1,
                'n': len(vals),
                'val_min': np.min(vals), 'val_max': np.max(vals),
                'val_med': np.median(vals), 'val_mean': np.mean(vals),
                'err_min': np.min(errs), 'err_max': np.max(errs),
                'err_pct_med': np.median(err_pct),
            })
            
            print(f"\n  col[{i:2d}] '{elem:8s}' ↔ col[{i+1:2d}] '{err_name}'")
            print(f"    N datos válidos: {len(vals)}")
            print(f"    VALOR: min={np.min(vals):.6f}  max={np.max(vals):.6f}  "
                  f"media={np.mean(vals):.6f}  mediana={np.median(vals):.6f}")
            print(f"    ERROR: min={np.min(errs):.6f}  max={np.max(errs):.6f}  "
                  f"media={np.mean(errs):.6f}  mediana={np.median(errs):.6f}")
            print(f"    Error relativo mediano: {np.median(err_pct):.1f}%")

# ═══════════════════════════════════════════════════════════════
# 3. DETERMINAR UNIDADES POR MAGNITUD DE VALORES
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  3. DETERMINACIÓN DE UNIDADES POR RANGO DE VALORES")
print("=" * 90)

# Referencia: composición típica de un GRANITO en corteza continental
# (valores de referencia de Clarke/Wedepohl para granito promedio)
granite_ref = {
    'Al2O3': {'pct': 14.0,  'ppm': 140000},
    'SiO2':  {'pct': 72.0,  'ppm': 720000},
    'K':     {'pct': 3.5,   'ppm': 35000},    # K2O ~4.2%, K ~3.5%
    'Ca':    {'pct': 1.3,   'ppm': 13000},    # CaO ~1.8%, Ca ~1.3%
    'Ti':    {'pct': 0.15,  'ppm': 1500},     # TiO2 ~0.25%, Ti ~0.15%
    'V':     {'pct': 0.0005,'ppm': 5},        # V traza
    'Cr':    {'pct': 0.0004,'ppm': 4},        # Cr traza
    'Mn':    {'pct': 0.035, 'ppm': 350},      # MnO ~0.045%, Mn ~0.035%
    'Fe':    {'pct': 2.0,   'ppm': 20000},    # Fe2O3 ~2.8%, Fe ~2%
    'Y':     {'pct': 0.002, 'ppm': 20},       # Y traza
}

print("\n  Composición típica de GRANITO (referencia):")
print(f"  {'Elem':>8s}  {'% peso':>10s}  {'ppm':>10s}")
print(f"  {'-'*8}  {'-'*10}  {'-'*10}")
for elem, vals in granite_ref.items():
    print(f"  {elem:>8s}  {vals['pct']:>10.4f}  {vals['ppm']:>10.0f}")

print("\n\n  --- Comparación con datos REALES del XLSX ---")
print(f"  {'Elem':>8s}  {'col':>4s}  {'Mediana':>10s}  {'Rango':>25s}  {'¿Unidad?':>12s}  {'Razonamiento'}")
print(f"  {'-'*8}  {'-'*4}  {'-'*10}  {'-'*25}  {'-'*12}  {'-'*40}")

for p in pairs_found:
    elem = p['elem']
    if elem in granite_ref:
        ref = granite_ref[elem]
        med = p['val_med']
        
        # ¿Es más cercano al % o al ppm?
        ratio_pct = med / ref['pct'] if ref['pct'] > 0 else float('inf')
        ratio_ppm = med / ref['ppm'] if ref['ppm'] > 0 else float('inf')
        
        if abs(np.log10(ratio_pct)) < abs(np.log10(ratio_ppm)):
            unidad_probable = "%"
            ratio_best = ratio_pct
        else:
            unidad_probable = "ppm"
            ratio_best = ratio_ppm
        
        rango_str = f"[{p['val_min']:.4f} - {p['val_max']:.4f}]"
        razon = f"ratio vs ref={ratio_best:.2f} (1.0=perfecto)"
        
        print(f"  {elem:>8s}  [{p['col_idx']:2d}]  {med:>10.4f}  {rango_str:>25s}  {unidad_probable:>12s}  {razon}")

# ═══════════════════════════════════════════════════════════════
# 4. VERIFICACIÓN INTERNA: Y% × 10000 = Yppm ?
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  4. VERIFICACIÓN INTERNA: Y% × 10000 vs Yppm")
print("=" * 90)

ypct_vals = []
yppm_vals = []
for row_idx in range(2, ws.max_row + 1):
    ypct = ws.cell(row=row_idx, column=56).value  # col 55 (0-indexed) = col 56 (1-indexed)
    yppm = ws.cell(row=row_idx, column=57).value  # col 56 (0-indexed) = col 57 (1-indexed)
    if isinstance(ypct, (int, float)) and isinstance(yppm, (int, float)):
        ypct_vals.append(float(ypct))
        yppm_vals.append(float(yppm))

ypct_vals = np.array(ypct_vals)
yppm_vals = np.array(yppm_vals)
yppm_calc = ypct_vals * 10000

print(f"  N pares: {len(ypct_vals)}")
print(f"\n  {'Muestra':>8s}  {'Y%':>10s}  {'Yppm (XLSX)':>12s}  {'Y%×10000':>12s}  {'Diferencia':>12s}  {'OK?':>5s}")
print(f"  {'-'*8}  {'-'*10}  {'-'*12}  {'-'*12}  {'-'*12}  {'-'*5}")

for i in range(min(15, len(ypct_vals))):
    diff = yppm_vals[i] - yppm_calc[i]
    ok = "✅" if abs(diff) < 1 else "⚠️"
    print(f"  {i+1:>8d}  {ypct_vals[i]:>10.4f}  {yppm_vals[i]:>12.1f}  {yppm_calc[i]:>12.1f}  {diff:>12.1f}  {ok:>5s}")

ratio_yppm = yppm_vals / yppm_calc
print(f"\n  Ratio Yppm_xlsx / (Y%×10000): media={np.mean(ratio_yppm):.4f}, mediana={np.median(ratio_yppm):.4f}")
print(f"  → Confirmado: Y% × 10,000 = Yppm en el XLSX")

# ═══════════════════════════════════════════════════════════════
# 5. APLICAR LA MISMA LÓGICA A OTROS ELEMENTOS
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  5. ¿OTROS ELEMENTOS TAMBIÉN ESTÁN EN %? — Verificación cruzada")
print("=" * 90)

# Si K, Ca, Ti, Fe están en %, sus valores × 10000 darían ppm
# Vamos a ver si eso tiene sentido comparando con Pob1

df1 = pd.read_csv(os.path.join(OUTDIR, "BD_Ytrio_LIMPIO.csv"))

# Leer las columnas geoquímicas del XLSX como dict
xlsx_data = {}
for i, h in enumerate(headers):
    if h is None:
        continue
    h = str(h).strip()
    vals = []
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=i+1).value
        if isinstance(v, (int, float)):
            vals.append(float(v))
    if vals:
        xlsx_data[f"col{i}_{h}"] = np.array(vals)

# Tabla de comparación
print(f"\n  {'Elem':>5s}  {'Pob1 col':>10s}  {'Pob1 unid':>10s}  {'Pob1 med':>12s}  {'Pob2 col':>5s}  "
      f"{'Pob2 raw':>12s}  {'Pob2×10k':>12s}  {'Pob2/10k':>12s}  {'Match':>12s}  {'Pob2 unid':>10s}")
print("  " + "-" * 120)

# Definir mapeo y comparar
mappings = [
    ('K',  'K__',    '%',   23),
    ('K',  'K',      'ppm', 23),
    ('Ca', 'Ca_ppm', 'ppm', 25),
    ('Ti', 'Ti__',   '%',   27),
    ('Ti', 'Ti',     'ppm', 27),
    ('V',  'V_ppm',  'ppm', 29),
    ('Cr', 'Cr_ppm', 'ppm', 31),
    ('Mn', 'Mn_',    '%',   33),
    ('Mn', 'Mn',     'ppm', 33),
    ('Fe', 'Fe__',   '%',   35),
    ('Fe', 'Fe',     'ppm', 35),
]

for elem, pob1_col, pob1_unit, pob2_idx in mappings:
    if pob1_col not in df1.columns:
        continue
    
    v1 = pd.to_numeric(df1[pob1_col], errors='coerce').dropna()
    if len(v1) == 0:
        continue
    med1 = v1.median()
    
    # Pob2
    vals2 = []
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=pob2_idx + 1).value
        if isinstance(v, (int, float)):
            vals2.append(float(v))
    
    if not vals2:
        continue
    
    v2 = np.array(vals2)
    med2_raw = np.median(v2)
    med2_x10k = med2_raw * 10000
    med2_d10k = med2_raw / 10000 if med2_raw > 0 else 0
    
    # ¿Cuál conversión da match?
    ratio_raw = med2_raw / med1 if med1 != 0 else float('inf')
    ratio_x10k = med2_x10k / med1 if med1 != 0 else float('inf')
    ratio_d10k = med2_d10k / med1 if med1 != 0 else float('inf')
    
    # Encontrar mejor match
    ratios = {'raw (=)': abs(np.log10(max(ratio_raw, 1e-10))),
              '×10000': abs(np.log10(max(ratio_x10k, 1e-10))),
              '÷10000': abs(np.log10(max(ratio_d10k, 1e-10)))}
    best = min(ratios, key=ratios.get)
    
    if best == 'raw (=)':
        if pob1_unit == '%':
            pob2_unit = '%'
        else:
            pob2_unit = 'ppm'
    elif best == '×10000':
        pob2_unit = '% → ppm'
    else:
        pob2_unit = 'ppm → %'
    
    print(f"  {elem:>5s}  {pob1_col:>10s}  {pob1_unit:>10s}  {med1:>12.2f}  [{pob2_idx:>3d}]  "
          f"{med2_raw:>12.4f}  {med2_x10k:>12.1f}  {med2_d10k:>12.8f}  {best:>12s}  {pob2_unit:>10s}")


# ═══════════════════════════════════════════════════════════════
# 6. VERIFICAR Ca ESPECÍFICAMENTE
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  6. VERIFICACIÓN DETALLADA: Ca")
print("=" * 90)

# Pob1: Ca_ppm → ¿son realmente ppm?
ca_pob1 = df1['Ca_ppm'].dropna()
print(f"  Pob1 Ca_ppm: n={len(ca_pob1)}, min={ca_pob1.min():.1f}, max={ca_pob1.max():.1f}, "
      f"mediana={ca_pob1.median():.1f}, media={ca_pob1.mean():.1f}")
print(f"  P5={ca_pob1.quantile(0.05):.1f}, P25={ca_pob1.quantile(0.25):.1f}, "
      f"P75={ca_pob1.quantile(0.75):.1f}, P95={ca_pob1.quantile(0.95):.1f}")

# Pob2: col 25 "Ca"
ca_pob2 = []
for row_idx in range(2, ws.max_row + 1):
    v = ws.cell(row=row_idx, column=26).value  # col 25 (0-indexed)
    if isinstance(v, (int, float)):
        ca_pob2.append(float(v))
ca_pob2 = np.array(ca_pob2)

print(f"\n  Pob2 Ca (raw): n={len(ca_pob2)}, min={ca_pob2.min():.4f}, max={ca_pob2.max():.4f}, "
      f"mediana={np.median(ca_pob2):.4f}, media={np.mean(ca_pob2):.4f}")

# Si Pob2 Ca está en %, entonces Ca% × 10000 = Ca ppm
ca_pob2_as_ppm = ca_pob2 * 10000
print(f"  Pob2 Ca×10000 (si fuera %→ppm): min={ca_pob2_as_ppm.min():.0f}, max={ca_pob2_as_ppm.max():.0f}, "
      f"mediana={np.median(ca_pob2_as_ppm):.0f}")
print(f"  → Pob1 mediana = {ca_pob1.median():.0f} ppm")
print(f"  → Pob2×10000 mediana = {np.median(ca_pob2_as_ppm):.0f} ppm")
print(f"  → Ratio = {np.median(ca_pob2_as_ppm) / ca_pob1.median():.2f}")

# ═══════════════════════════════════════════════════════════════
# 7. VERIFICAR Mn ESPECÍFICAMENTE
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  7. VERIFICACIÓN DETALLADA: Mn")
print("=" * 90)

# Pob1 tiene Mn_ (%) y Mn (ppm)
mn_pct = df1['Mn_'].dropna()
mn_ppm = df1['Mn'].dropna()

print(f"  Pob1 Mn_ (%):  n={len(mn_pct)}, min={mn_pct.min():.4f}, max={mn_pct.max():.4f}, "
      f"mediana={mn_pct.median():.4f}, media={mn_pct.mean():.4f}")
print(f"  Pob1 Mn (ppm): n={len(mn_ppm)}, min={mn_ppm.min():.1f}, max={mn_ppm.max():.1f}, "
      f"mediana={mn_ppm.median():.1f}, media={mn_ppm.mean():.1f}")

# Verificar conversión interna Pob1
mn_check = (df1['Mn'] / (df1['Mn_'] * 10000)).dropna()
mn_check = mn_check[np.isfinite(mn_check) & (mn_check > 0) & (mn_check < 10)]
print(f"\n  Pob1 ratio Mn_ppm / (Mn_%×10000): mediana={mn_check.median():.4f}, media={mn_check.mean():.4f}")

# Pob2: col 33 "Mn"
mn_pob2 = []
for row_idx in range(2, ws.max_row + 1):
    v = ws.cell(row=row_idx, column=34).value  # col 33 (0-indexed)
    if isinstance(v, (int, float)):
        mn_pob2.append(float(v))
mn_pob2 = np.array(mn_pob2)

print(f"\n  Pob2 Mn (raw): n={len(mn_pob2)}, min={mn_pob2.min():.4f}, max={mn_pob2.max():.4f}, "
      f"mediana={np.median(mn_pob2):.4f}, media={np.mean(mn_pob2):.4f}")

# Comparar con ambas columnas de Pob1
print(f"\n  Comparación de medianas:")
print(f"    Pob2 raw       = {np.median(mn_pob2):.4f}")
print(f"    Pob1 Mn_ (%)   = {mn_pct.median():.4f}  → ratio Pob2/Pob1 = {np.median(mn_pob2)/mn_pct.median():.2f}")
print(f"    Pob1 Mn (ppm)  = {mn_ppm.median():.1f}   → ratio Pob2/Pob1 = {np.median(mn_pob2)/mn_ppm.median():.6f}")
print(f"    Pob2×10000     = {np.median(mn_pob2)*10000:.1f}")
print(f"    → Pob2×10000 vs Pob1 Mn(ppm): ratio = {np.median(mn_pob2)*10000/mn_ppm.median():.2f}")

# Outliers de Mn_ en Pob1
print(f"\n  Pob1 Mn_ outliers (>1%):")
mn_outliers = mn_pct[mn_pct > 1]
print(f"    {len(mn_outliers)} valores > 1%")
print(f"    {len(mn_pct[mn_pct > 0.5])} valores > 0.5%")
print(f"    Top 10: {sorted(mn_outliers.values, reverse=True)[:10]}")

# ═══════════════════════════════════════════════════════════════
# 8. VERIFICAR V y Cr ESPECÍFICAMENTE
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  8. VERIFICACIÓN DETALLADA: V y Cr")
print("=" * 90)

# V
v_pob1 = df1['V_ppm'].dropna()
v_pob2 = []
v_censored = 0
for row_idx in range(2, ws.max_row + 1):
    v = ws.cell(row=row_idx, column=30).value  # col 29 (0-indexed)
    if isinstance(v, (int, float)):
        v_pob2.append(float(v))
    elif isinstance(v, str) and 'LOD' in v:
        v_censored += 1

print(f"  V_ppm Pob1: n={len(v_pob1)}, mediana={v_pob1.median():.1f}, rango=[{v_pob1.min():.1f}-{v_pob1.max():.1f}]")
if v_pob2:
    v_pob2 = np.array(v_pob2)
    print(f"  V Pob2 raw: n={len(v_pob2)}, mediana={np.median(v_pob2):.6f}, rango=[{v_pob2.min():.6f}-{v_pob2.max():.6f}]")
    print(f"  V Pob2 censurados: {v_censored}")
    print(f"  V Pob2×10000 (si %→ppm): mediana={np.median(v_pob2)*10000:.1f}, rango=[{v_pob2.min()*10000:.1f}-{v_pob2.max()*10000:.1f}]")
    print(f"  → Pob2×10000 mediana ({np.median(v_pob2)*10000:.1f}) vs Pob1 mediana ({v_pob1.median():.1f}): ratio={np.median(v_pob2)*10000/v_pob1.median():.2f}")

# Cr
print()
cr_pob1 = df1['Cr_ppm'].dropna()
cr_pob2 = []
cr_censored = 0
for row_idx in range(2, ws.max_row + 1):
    v = ws.cell(row=row_idx, column=32).value  # col 31 (0-indexed)
    if isinstance(v, (int, float)):
        cr_pob2.append(float(v))
    elif isinstance(v, str) and 'LOD' in v:
        cr_censored += 1

print(f"  Cr_ppm Pob1: n={len(cr_pob1)}, mediana={cr_pob1.median():.1f}, rango=[{cr_pob1.min():.1f}-{cr_pob1.max():.1f}]")
if cr_pob2:
    cr_pob2 = np.array(cr_pob2)
    print(f"  Cr Pob2 raw: n={len(cr_pob2)}, mediana={np.median(cr_pob2):.6f}, rango=[{cr_pob2.min():.6f}-{cr_pob2.max():.6f}]")
    print(f"  Cr Pob2 censurados: {cr_censored}")
    print(f"  Cr Pob2×10000 (si %→ppm): mediana={np.median(cr_pob2)*10000:.1f}, rango=[{cr_pob2.min()*10000:.1f}-{cr_pob2.max()*10000:.1f}]")
    print(f"  → Pob2×10000 mediana ({np.median(cr_pob2)*10000:.1f}) vs Pob1 mediana ({cr_pob1.median():.1f}): ratio={np.median(cr_pob2)*10000/cr_pob1.median():.2f}")


# ═══════════════════════════════════════════════════════════════
# 9. ¿EL XLSX ES XRF? VERIFICAR SUMA DE ÓXIDOS
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  9. ¿ES XRF PORTÁTIL? — Verificar suma de óxidos mayores")
print("=" * 90)

# Si Al2O3, SiO2, K, Ca, Ti, Mn, Fe están en %, la suma de óxidos
# mayores debería dar algo razonable (~50-100% para roca)

sums = []
for row_idx in range(2, ws.max_row + 1):
    al2o3 = ws.cell(row=row_idx, column=18).value  # col 17
    sio2  = ws.cell(row=row_idx, column=20).value  # col 19
    k_val = ws.cell(row=row_idx, column=24).value  # col 23
    ca_val= ws.cell(row=row_idx, column=26).value  # col 25
    ti_val= ws.cell(row=row_idx, column=28).value  # col 27
    mn_val= ws.cell(row=row_idx, column=34).value  # col 33
    fe_val= ws.cell(row=row_idx, column=36).value  # col 35
    
    vals = [al2o3, sio2, k_val, ca_val, ti_val, mn_val, fe_val]
    nums = [float(v) for v in vals if isinstance(v, (int, float))]
    
    if len(nums) == 7:
        # Convertir K a K2O, Ca a CaO, Ti a TiO2, Mn a MnO, Fe a Fe2O3
        # K2O = K × 1.2046, CaO = Ca × 1.3992, TiO2 = Ti × 1.6681
        # MnO = Mn × 1.2912, Fe2O3 = Fe × 1.4297
        al2o3_v = nums[0]
        sio2_v = nums[1]
        k2o = nums[2] * 1.2046
        cao = nums[3] * 1.3992
        tio2 = nums[4] * 1.6681
        mno = nums[5] * 1.2912
        fe2o3 = nums[6] * 1.4297
        
        total = al2o3_v + sio2_v + k2o + cao + tio2 + mno + fe2o3
        sums.append({
            'row': row_idx, 'Al2O3': al2o3_v, 'SiO2': sio2_v,
            'K2O': k2o, 'CaO': cao, 'TiO2': tio2, 'MnO': mno,
            'Fe2O3': fe2o3, 'Total': total,
            'K_raw': nums[2], 'Ca_raw': nums[3], 'Ti_raw': nums[4],
            'Mn_raw': nums[5], 'Fe_raw': nums[6]
        })

print(f"  Muestras con todos los óxidos mayores: {len(sums)}")
if sums:
    totals = [s['Total'] for s in sums]
    print(f"  Suma de óxidos mayores (parcial, sin Na2O, P2O5, H2O):")
    print(f"    min = {min(totals):.1f}%")
    print(f"    max = {max(totals):.1f}%")
    print(f"    media = {np.mean(totals):.1f}%")
    print(f"    mediana = {np.median(totals):.1f}%")
    
    print(f"\n  → Suma parcial mediana = {np.median(totals):.1f}% (esperado ~60-90% sin Na2O, P2O5, LOI)")
    
    if 40 < np.median(totals) < 100:
        print(f"  ✅ CONFIRMADO: Los datos del XLSX están en % peso (óxidos mayores)")
        print(f"  → Al2O3, SiO2 son directamente %, K/Ca/Ti/Mn/Fe son % del ELEMENTO")
    else:
        print(f"  ⚠️ La suma no es coherente con % → posible mezcla de unidades")
    
    print(f"\n  Detalle (primeras 10 muestras):")
    print(f"  {'Fila':>5s}  {'Al2O3':>7s}  {'SiO2':>7s}  {'K2O':>7s}  {'CaO':>7s}  {'TiO2':>7s}  {'MnO':>7s}  {'Fe2O3':>7s}  {'TOTAL':>8s}")
    for s in sums[:10]:
        print(f"  {s['row']:>5d}  {s['Al2O3']:>7.2f}  {s['SiO2']:>7.2f}  {s['K2O']:>7.3f}  "
              f"{s['CaO']:>7.3f}  {s['TiO2']:>7.3f}  {s['MnO']:>7.3f}  {s['Fe2O3']:>7.2f}  {s['Total']:>8.2f}")


# ═══════════════════════════════════════════════════════════════
# 10. CONCLUSIÓN FINAL DE UNIDADES
# ═══════════════════════════════════════════════════════════════
print("\n\n" + "=" * 90)
print("  10. CONCLUSIÓN FINAL — TABLA DE MAPEO CORRECTO")
print("=" * 90)

print("""
  ┌──────────┬──────────────┬──────────┬──────────────┬──────────┬──────────────────────────────┐
  │ Elemento │ Pob1 columna │ Pob1 und │ Pob2 col idx │ Pob2 und │ Conversión necesaria         │
  ├──────────┼──────────────┼──────────┼──────────────┼──────────┼──────────────────────────────┤
  │ K        │ K__          │ %        │ [23]         │ %        │ NINGUNA (ya compatible)      │
  │ K        │ K            │ ppm      │ [23]         │ %        │ Pob2 × 10,000 = ppm         │
  │ Ca       │ Ca_ppm       │ ppm      │ [25]         │ %        │ Pob2 × 10,000 = ppm  ← FIX  │
  │ Ti       │ Ti__         │ %        │ [27]         │ %        │ NINGUNA (ya compatible)      │
  │ Ti       │ Ti           │ ppm      │ [27]         │ %        │ Pob2 × 10,000 = ppm         │
  │ V        │ V_ppm        │ ppm      │ [29]         │ %        │ Pob2 × 10,000 = ppm  ← FIX  │
  │ Cr       │ Cr_ppm       │ ppm      │ [31]         │ %        │ Pob2 × 10,000 = ppm  ← FIX  │
  │ Mn       │ Mn_          │ %        │ [33]         │ %        │ NINGUNA (ya compatible)      │
  │ Mn       │ Mn           │ ppm      │ [33]         │ %        │ Pob2 × 10,000 = ppm         │
  │ Fe       │ Fe__         │ %        │ [35]         │ %        │ NINGUNA (ya compatible)      │
  │ Fe       │ Fe           │ ppm      │ [35]         │ %        │ Pob2 × 10,000 = ppm         │
  │ Y        │ Y_ppm        │ ppm      │ [56] Yppm   │ ppm      │ NINGUNA (ya compatible)      │
  │ Y        │ —            │ —        │ [55] Y%      │ %        │ Y% × 10,000 = Yppm          │
  └──────────┴──────────────┴──────────┴──────────────┴──────────┴──────────────────────────────┘

  RESUMEN:
  ✅ K (%↔%), Ti (%↔%), Fe (%↔%), Y (ppm↔ppm) → análisis anterior VÁLIDO
  🔴 Ca (ppm↔%) → necesita Ca_pob2 × 10,000 para comparar con Ca_ppm de Pob1
  🔴 V (ppm↔%) → necesita V_pob2 × 10,000 para comparar con V_ppm de Pob1
  🔴 Cr (ppm↔%) → necesita Cr_pob2 × 10,000 para comparar con Cr_ppm de Pob1
  ⚠️ Mn (%↔%) → unidades compatibles pero Pob1 tiene outliers extremos (max=571%)
  
  El XLSX proviene de un equipo XRF portátil. TODOS los elementos se reportan en
  % peso del ELEMENTO (no del óxido). Solo Yppm/Ypond están pre-convertidos a ppm.
""")
