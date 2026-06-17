#!/usr/bin/env python3
"""
Genera mapa de isoconcentraciones de Y (ppm) con contornos cada 25 ppm.
Usa string.Template para evitar conflictos con llaves de JS.
"""
import pandas as pd
import numpy as np
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
import pyproj
import openpyxl
from scipy.interpolate import griddata
from scipy.ndimage import gaussian_filter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from string import Template

OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"

# ══════════════════════════════════════════════════════════════════
# 1. CARGAR TODOS LOS DATOS
# ══════════════════════════════════════════════════════════════════
transformer = pyproj.Transformer.from_crs('EPSG:32718', 'EPSG:4326', always_xy=True)

# --- Datos existentes ---
df = pd.read_csv(os.path.join(OUTDIR, "BD_Ytrio_LIMPIO.csv"))
df = df.dropna(subset=['UTM_E','UTM_N','Y_ppm'])
df = df[(df['UTM_E'] > 100000) & (df['UTM_N'] > 1000000)].copy()
lons_old, lats_old = transformer.transform(df['UTM_E'].values, df['UTM_N'].values)
df['lat'] = lats_old
df['lon'] = lons_old
df = df[(df['lat'] > -60) & (df['lat'] < -20) & (df['lon'] > -80) & (df['lon'] < -60)]

points_old = list(zip(df['lon'].values, df['lat'].values, df['Y_ppm'].values,
                      df['Sample'].values, df['UTM_E'].values, df['UTM_N'].values))
print(f"Datos existentes: {len(points_old)} puntos")

# --- Datos nuevos ---
wb = openpyxl.load_workbook(os.path.join(OUTDIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']
points_new = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    cp, idsample, elev, xm, ym = row[0], row[1], row[2], row[3], row[4]
    yppm = row[56]
    if xm is None or ym is None or yppm is None or not isinstance(yppm, (int, float)):
        continue
    lon, lat = transformer.transform(xm, ym)
    points_new.append((lon, lat, float(yppm), f"{cp}", xm, ym))

print(f"Datos nuevos: {len(points_new)} puntos")

# --- Combinar ---
all_points = points_old + points_new
lons = np.array([p[0] for p in all_points])
lats = np.array([p[1] for p in all_points])
vals = np.array([p[2] for p in all_points])
utm_es = np.array([p[4] for p in all_points])
utm_ns = np.array([p[5] for p in all_points])

print(f"Total: {len(all_points)} puntos")
print(f"Y ppm: min={vals.min():.1f}, max={vals.max():.1f}, mean={vals.mean():.1f}")

# ══════════════════════════════════════════════════════════════════
# 2. INTERPOLACIÓN EN UTM
# ══════════════════════════════════════════════════════════════════
margin_m = 2000
x_min, x_max = utm_es.min() - margin_m, utm_es.max() + margin_m
y_min, y_max = utm_ns.min() - margin_m, utm_ns.max() + margin_m

nx = min(int((x_max - x_min) / 500), 300)
ny = min(int((y_max - y_min) / 500), 300)

print(f"Grid: {nx} x {ny} = {nx*ny:,} celdas")

xi = np.linspace(x_min, x_max, nx)
yi = np.linspace(y_min, y_max, ny)
XI, YI = np.meshgrid(xi, yi)

ZI = griddata((utm_es, utm_ns), vals, (XI, YI), method='linear')
ZI_nearest = griddata((utm_es, utm_ns), vals, (XI, YI), method='nearest')

from scipy.spatial import cKDTree
tree = cKDTree(np.column_stack([utm_es, utm_ns]))
grid_points = np.column_stack([XI.ravel(), YI.ravel()])
distances, _ = tree.query(grid_points)
dist_grid = distances.reshape(XI.shape)

max_dist = 5000
mask_fill = np.isnan(ZI) & (dist_grid < max_dist)
ZI[mask_fill] = ZI_nearest[mask_fill]

ZI_smooth = gaussian_filter(ZI, sigma=1.8)
ZI_smooth[dist_grid > max_dist] = np.nan

print(f"Interpolación completada. NaN: {np.isnan(ZI_smooth).sum():,} / {ZI_smooth.size:,}")

# ══════════════════════════════════════════════════════════════════
# 3. GENERAR CONTORNOS
# ══════════════════════════════════════════════════════════════════
LONS_grid, LATS_grid = transformer.transform(XI, YI)

max_level = min(int(np.nanmax(ZI_smooth)), 600)
levels = list(range(25, max_level + 25, 25))
if levels[-1] < np.nanmax(vals):
    levels.append(levels[-1] + 25)
print(f"Niveles: {levels}")

colors_hex = [
    '#313695','#4575b4','#74add1','#abd9e9',
    '#e0f3f8','#ffffbf','#fee090',
    '#fdae61','#f46d43','#d73027',
    '#a50026','#7a0019','#540010',
]

def get_fill_color(level_min, level_max):
    mid = (level_min + level_max) / 2
    t = min(1.0, mid / 300)
    idx = min(int(t * (len(colors_hex) - 1)), len(colors_hex) - 2)
    return colors_hex[idx]

fig, ax = plt.subplots(figsize=(10, 10))
cs = ax.contourf(LONS_grid, LATS_grid, ZI_smooth, levels=[0] + levels, extend='max')
cs_lines = ax.contour(LONS_grid, LATS_grid, ZI_smooth, levels=levels)
plt.close(fig)

filled_features = []
line_features = []

fill_levels_all = list(cs.levels)
for i, segs in enumerate(cs.allsegs):
    level_min = fill_levels_all[i] if i < len(fill_levels_all) else levels[-1]
    level_max = fill_levels_all[i+1] if i+1 < len(fill_levels_all) else level_min + 25
    color = get_fill_color(level_min, level_max)
    for seg in segs:
        if len(seg) >= 3:
            step = max(1, len(seg) // 200)
            decimated = seg[::step].tolist()
            ring = [[round(c[0], 4), round(c[1], 4)] for c in decimated]
            if ring[0] != ring[-1]:
                ring.append(ring[0])
            if len(ring) < 4:
                continue
            filled_features.append({
                "type": "Feature",
                "properties": {"level_min": float(level_min), "level_max": float(level_max),
                               "color": color, "label": f"{int(level_min)}-{int(level_max)} ppm"},
                "geometry": {"type": "Polygon", "coordinates": [ring]}
            })

for i, segs in enumerate(cs_lines.allsegs):
    level = float(cs_lines.levels[i])
    for seg in segs:
        step = max(1, len(seg) // 150)
        decimated = seg[::step].tolist()
        vertices = [[round(c[0], 4), round(c[1], 4)] for c in decimated]
        if len(vertices) >= 2:
            line_features.append({
                "type": "Feature",
                "properties": {"level": level, "label": f"{int(level)} ppm"},
                "geometry": {"type": "LineString", "coordinates": vertices}
            })

print(f"Polígonos: {len(filled_features)}, Líneas: {len(line_features)}")

# ══════════════════════════════════════════════════════════════════
# 4. DATOS DE PUNTOS
# ══════════════════════════════════════════════════════════════════
sample_records = []
for p in all_points:
    sample_records.append({
        'lon': round(float(p[0]), 6), 'lat': round(float(p[1]), 6),
        'Y_ppm': round(float(p[2]), 1), 'Sample': str(p[3]),
        'UTM_E': round(float(p[4]), 0), 'UTM_N': round(float(p[5]), 0),
    })

center_lat = float(np.median(lats))
center_lon = float(np.median(lons))

# ══════════════════════════════════════════════════════════════════
# 5. GENERAR HTML — sin f-strings, usando replace
# ══════════════════════════════════════════════════════════════════
json_filled = json.dumps({"type":"FeatureCollection","features":filled_features})
json_lines = json.dumps({"type":"FeatureCollection","features":line_features})
json_samples = json.dumps(sample_records, ensure_ascii=False)
levels_json = json.dumps(levels)

n_total = len(all_points)
y_min_val = float(vals.min())
y_med_val = float(np.median(vals))
y_mean_val = float(vals.mean())
y_max_val = float(vals.max())
n_anom = int((vals >= 50).sum())

html = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Mapa de Isoconcentraciones — Y (ppm) | Chile</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Inter','Segoe UI',sans-serif;background:#0a0e14;color:#e0e0e0;display:flex;flex-direction:column;height:100vh;overflow:hidden}
#header{background:linear-gradient(135deg,#0d1b2a,#1a2a3a,#162035);padding:12px 24px;display:flex;align-items:center;justify-content:space-between;border-bottom:2px solid #1e3a5a;flex-shrink:0;z-index:1000}
#header h1{font-size:18px;font-weight:700;color:#4fc3f7;letter-spacing:.5px}
#header .sub{font-size:11px;color:#78909c;margin-top:3px}
.pill{background:rgba(30,42,58,.8);border:1px solid #2a4a6a;border-radius:20px;padding:5px 16px;font-size:12px;backdrop-filter:blur(4px)}
.pill b{color:#4fc3f7}
#main{display:flex;flex:1;overflow:hidden}
#sidebar{width:300px;background:linear-gradient(180deg,#111820,#0d1218);border-right:1px solid #1e3040;display:flex;flex-direction:column;flex-shrink:0;z-index:500;overflow-y:auto;overflow-x:hidden}
#sidebar::-webkit-scrollbar{width:6px}
#sidebar::-webkit-scrollbar-thumb{background:#2a3a4a;border-radius:3px}
.sec{padding:10px 14px;border-bottom:1px solid #1a2535}
.sec h3{font-size:10px;color:#546e7a;text-transform:uppercase;letter-spacing:1.2px;margin-bottom:6px;font-weight:600}
.lbl{font-size:10px;color:#78909c;display:block;margin-bottom:3px}
select{width:100%;background:#1a2535;border:1px solid #2a3a4a;color:#e0e0e0;padding:5px 8px;border-radius:6px;font-size:11px}
input[type=range]{width:100%;accent-color:#4fc3f7;margin:3px 0}
.rval{display:flex;justify-content:space-between;font-size:9px;color:#546e7a}
.lbtns{display:flex;gap:4px;margin-top:5px}
.lbtn{flex:1;padding:5px;background:#1a2535;border:1px solid #2a3a4a;color:#78909c;border-radius:6px;cursor:pointer;font-size:11px;text-align:center;transition:all .2s}
.lbtn.on{background:#0077b6;border-color:#4fc3f7;color:#fff}
.lbtn:hover:not(.on){background:#223040;color:#b0bec5}
.ovl{display:flex;align-items:center;gap:8px;padding:4px 0;font-size:10px}
.ovl label{cursor:pointer;color:#b0bec5}
.tgl{position:relative;display:inline-block;width:36px;height:20px;flex-shrink:0}
.tgl input{position:absolute;width:36px;height:20px;opacity:0;margin:0;cursor:pointer;z-index:2}
.tgl .sl{position:absolute;inset:0;background:#2a3a4a;border-radius:10px;transition:.3s;pointer-events:none}
.tgl .sl::before{content:'';position:absolute;height:16px;width:16px;left:2px;bottom:2px;background:#546e7a;border-radius:50%;transition:.3s}
.tgl input:checked+.sl{background:#0077b6}
.tgl input:checked+.sl::before{transform:translateX(16px);background:#4fc3f7}
#mapwrap{flex:1;position:relative}
#map{width:100%;height:100%;z-index:1}
.leaflet-container{background:#0a0e14!important}
#statbar{position:absolute;top:10px;right:12px;display:flex;gap:8px;z-index:800}
.stat{background:rgba(13,27,42,.92);border:1px solid #2a4a6a;border-radius:16px;padding:5px 14px;font-size:12px;backdrop-filter:blur(6px)}
.stat b{color:#4fc3f7}
.ctt{background:rgba(10,20,35,.97)!important;border:1px solid #2a4a6a!important;border-radius:10px!important;color:#e0e0e0!important;font-size:12px!important;padding:10px 14px!important;box-shadow:0 4px 24px rgba(0,0,0,.7)!important}
.lscale{display:flex;flex-direction:column;gap:2px;margin-top:4px}
.lgi{display:flex;align-items:center;gap:6px;font-size:10px}
.lgs{width:20px;height:12px;border-radius:3px;border:1px solid rgba(255,255,255,.15);flex-shrink:0}
</style>
</head>
<body>
<div id="header">
  <div>
    <h1>&#x1F5FA;&#xFE0F; MAPA DE ISOCONCENTRACIONES — Y (ppm)</h1>
    <div class="sub">Interpolaci&oacute;n de %%N_TOTAL%% muestras | Contornos cada 25 ppm | Regi&oacute;n del Biob&iacute;o, Chile</div>
  </div>
  <div style="display:flex;gap:8px;flex-wrap:wrap">
    <div class="pill">Muestras: <b>%%N_TOTAL%%</b></div>
    <div class="pill">Y media: <b>%%Y_MEAN%%</b> ppm</div>
    <div class="pill">Y m&aacute;x: <b>%%Y_MAX%%</b> ppm</div>
  </div>
</div>
<div id="main">
  <div id="sidebar">
    <div class="sec">
      <h3>&#x1F5FA;&#xFE0F; Capa Base</h3>
      <div class="lbtns">
        <div class="lbtn" id="btn-osm" onclick="setLyr('osm')">Calles</div>
        <div class="lbtn" id="btn-sat" onclick="setLyr('sat')">Sat&eacute;lite</div>
        <div class="lbtn on" id="btn-topo" onclick="setLyr('topo')">Topo</div>
        <div class="lbtn" id="btn-dark" onclick="setLyr('dark')">Oscuro</div>
      </div>
    </div>
    <div class="sec">
      <h3>&#x1F4CA; Capas</h3>
      <div class="ovl">
        <label class="tgl"><input type="checkbox" id="chkFill" checked onchange="tglFill(this.checked)"><span class="sl"></span></label>
        <label>&#x1F3A8; Relleno de contornos</label>
      </div>
      <div class="ovl">
        <label class="tgl"><input type="checkbox" id="chkLines" checked onchange="tglLines(this.checked)"><span class="sl"></span></label>
        <label>&#x3030;&#xFE0F; L&iacute;neas de contorno</label>
      </div>
      <div class="ovl">
        <label class="tgl"><input type="checkbox" id="chkLabels" checked onchange="tglLabels(this.checked)"><span class="sl"></span></label>
        <label>&#x1F3F7;&#xFE0F; Etiquetas (ppm)</label>
      </div>
      <div class="ovl">
        <label class="tgl"><input type="checkbox" id="chkPts" checked onchange="tglPts(this.checked)"><span class="sl"></span></label>
        <label>&#x1F4CD; Puntos de muestreo</label>
      </div>
      <div class="ovl">
        <label class="tgl"><input type="checkbox" id="chkGeo" onchange="tglGeo(this.checked)"><span class="sl"></span></label>
        <label>&#x1FAA8; Geolog&iacute;a (Macrostrat)</label>
      </div>
    </div>
    <div class="sec">
      <h3>&#x1F39B;&#xFE0F; Opacidad contornos</h3>
      <input type="range" id="fillA" min="0.1" max="0.9" step="0.05" value="0.55" oninput="setFillA(this.value)">
      <div class="rval"><span>10%</span><span id="faV">55%</span><span>90%</span></div>
    </div>
    <div class="sec">
      <h3>&#x1F4CD; Tama&ntilde;o puntos</h3>
      <input type="range" id="ptSz" min="2" max="12" value="5" oninput="setPtSz(+this.value)">
      <div class="rval"><span>2</span><span id="psV">5</span><span>12</span></div>
    </div>
    <div class="sec">
      <h3>&#x1F3A8; Escala Y (ppm)</h3>
      <div class="lscale" id="legItems"></div>
    </div>
    <div class="sec">
      <h3>&#x1F4C8; Estad&iacute;sticas</h3>
      <div style="font-size:11px;line-height:1.8">
        <div>Total muestras: <b style="color:#4fc3f7">%%N_TOTAL%%</b></div>
        <div>Y m&iacute;n: <b style="color:#66bb6a">%%Y_MIN%% ppm</b></div>
        <div>Y mediana: <b style="color:#ffa726">%%Y_MED%% ppm</b></div>
        <div>Y media: <b style="color:#ffa726">%%Y_MEAN%% ppm</b></div>
        <div>Y m&aacute;x: <b style="color:#ff5722">%%Y_MAX%% ppm</b></div>
        <div>Anomal&iacute;as (&ge;50): <b style="color:#ff5722">%%N_ANOM%%</b></div>
        <div>Interpolaci&oacute;n: <b style="color:#78909c">Linear + Gaussian</b></div>
        <div>Resoluci&oacute;n: <b style="color:#78909c">~500 m/celda</b></div>
      </div>
    </div>
  </div>
  <div id="mapwrap">
    <div id="map"></div>
    <div id="statbar">
      <div class="stat">Zoom: <b id="zoomLv">-</b></div>
    </div>
  </div>
</div>
<script>
var FILLED = %%FILLED_JSON%%;
var LINES = %%LINES_JSON%%;
var SAMPLES = %%SAMPLES_JSON%%;
var LEVELS = %%LEVELS_JSON%%;
var CENTER = [%%CENTER_LAT%%, %%CENTER_LON%%];

var COLORS = ['#313695','#4575b4','#74add1','#abd9e9','#e0f3f8','#ffffbf','#fee090','#fdae61','#f46d43','#d73027','#a50026','#7a0019','#540010'];

function lvColor(val) {
  var t = Math.min(1, val / 300);
  var idx = Math.min(Math.floor(t * (COLORS.length - 1)), COLORS.length - 2);
  return COLORS[idx];
}

// Leaflet setup
var LYRS = {
  osm: L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {maxZoom:19,attribution:'OSM'}),
  sat: L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', {maxZoom:19,attribution:'Esri'}),
  topo: L.tileLayer('https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png', {maxZoom:17,attribution:'OpenTopoMap'}),
  dark: L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png', {maxZoom:19,attribution:'CartoDB'})
};

var map = L.map('map', {center: CENTER, zoom: 9, zoomControl: true});
var curLyr = LYRS.topo.addTo(map);

function setLyr(k) {
  map.removeLayer(curLyr);
  curLyr = LYRS[k].addTo(map);
  document.querySelectorAll('.lbtn').forEach(function(b){b.classList.remove('on')});
  document.getElementById('btn-'+k).classList.add('on');
  if (geoLyr && map.hasLayer(geoLyr)) geoLyr.bringToFront();
  fillGrp.bringToFront();
  lineGrp.bringToFront();
  labelGrp.bringToFront();
  ptGrp.bringToFront();
}

map.on('zoomend', function(){ document.getElementById('zoomLv').textContent = map.getZoom(); });

// Geology
var geoLyr = null;
map.createPane('geo'); map.getPane('geo').style.zIndex = 350;
function tglGeo(on) {
  if (on) {
    if (!geoLyr) geoLyr = L.tileLayer('https://tiles.macrostrat.org/carto/{z}/{x}/{y}.png', {maxZoom:19,opacity:0.4,pane:'geo'});
    geoLyr.addTo(map);
  } else if (geoLyr) map.removeLayer(geoLyr);
}

// Panes
map.createPane('cFill'); map.getPane('cFill').style.zIndex = 400;
map.createPane('cLine'); map.getPane('cLine').style.zIndex = 410;
map.createPane('cLabel'); map.getPane('cLabel').style.zIndex = 420;
map.createPane('pts'); map.getPane('pts').style.zIndex = 450;

// Filled contours
var fillAlpha = 0.55;
var fillGrp = L.layerGroup().addTo(map);

function renderFill() {
  fillGrp.clearLayers();
  L.geoJSON(FILLED, {
    pane: 'cFill',
    style: function(f) {
      return {fillColor: f.properties.color, fillOpacity: fillAlpha, color: 'rgba(255,255,255,0.08)', weight: 0.3};
    },
    onEachFeature: function(f, layer) {
      layer.bindTooltip(f.properties.label, {className:'ctt', sticky:true});
    }
  }).addTo(fillGrp);
}
renderFill();

function tglFill(on) { on ? fillGrp.addTo(map) : map.removeLayer(fillGrp); }
function setFillA(v) {
  fillAlpha = parseFloat(v);
  document.getElementById('faV').textContent = Math.round(v*100)+'%';
  renderFill();
}

// Contour lines
var lineGrp = L.layerGroup().addTo(map);
function renderLines() {
  lineGrp.clearLayers();
  L.geoJSON(LINES, {
    pane: 'cLine',
    style: function(f) {
      var lv = f.properties.level;
      var main = (lv % 50 === 0);
      return {
        color: main ? 'rgba(255,255,255,0.7)' : 'rgba(200,200,200,0.35)',
        weight: main ? 1.8 : 0.8,
        dashArray: main ? null : '4 3'
      };
    }
  }).addTo(lineGrp);
}
renderLines();
function tglLines(on) { on ? lineGrp.addTo(map) : map.removeLayer(lineGrp); }

// Labels
var labelGrp = L.layerGroup().addTo(map);
function renderLabels() {
  labelGrp.clearLayers();
  LINES.features.forEach(function(f) {
    var lv = f.properties.level;
    if (lv % 25 !== 0) return;
    var coords = f.geometry.coordinates;
    if (coords.length < 4) return;
    var positions = [Math.floor(coords.length * 0.33), Math.floor(coords.length * 0.66)];
    positions.forEach(function(idx) {
      if (idx >= coords.length) return;
      var pt = coords[idx];
      var clr = lv >= 100 ? '#ff7043' : lv >= 50 ? '#ffa726' : '#b0bec5';
      var mk = L.marker([pt[1], pt[0]], {
        pane: 'cLabel',
        icon: L.divIcon({
          className: '',
          html: '<div style="background:rgba(10,14,20,0.85);color:'+clr+';font-size:10px;font-weight:600;padding:1px 5px;border-radius:4px;border:1px solid rgba(255,255,255,0.2);white-space:nowrap">'+lv+'</div>',
          iconSize: [0,0], iconAnchor: [15,8]
        })
      });
      labelGrp.addLayer(mk);
    });
  });
}
renderLabels();
function tglLabels(on) { on ? labelGrp.addTo(map) : map.removeLayer(labelGrp); }

// Sample points
var ptGrp = L.layerGroup().addTo(map);
var ptSz = 5;
function renderPts() {
  ptGrp.clearLayers();
  SAMPLES.forEach(function(s) {
    var col = lvColor(s.Y_ppm);
    var bCol = s.Y_ppm >= 100 ? '#fff' : s.Y_ppm >= 50 ? '#ffd54f' : 'rgba(255,255,255,0.3)';
    var bW = s.Y_ppm >= 100 ? 2 : s.Y_ppm >= 50 ? 1.5 : 0.5;
    var m = L.circleMarker([s.lat, s.lon], {
      pane:'pts', radius:ptSz, fillColor:col, color:bCol, weight:bW, fillOpacity:0.85, opacity:1
    });
    var yClr = s.Y_ppm > 100 ? '#ff5722' : s.Y_ppm > 50 ? '#ffa726' : '#66bb6a';
    m.bindTooltip(
      '<div style="font-weight:700;color:#4fc3f7;font-size:13px;margin-bottom:4px">' + s.Sample + '</div>' +
      '<div><b>Y:</b> <span style="color:'+yClr+';font-weight:700">' + s.Y_ppm.toFixed(1) + ' ppm</span></div>' +
      '<div style="font-size:10px;color:#78909c;margin-top:2px">UTM: ' + s.UTM_E.toLocaleString() + ' E / ' + s.UTM_N.toLocaleString() + ' N</div>',
      {className:'ctt', sticky:true}
    );
    ptGrp.addLayer(m);
  });
}
renderPts();
function tglPts(on) { on ? ptGrp.addTo(map) : map.removeLayer(ptGrp); }
function setPtSz(v) { ptSz = v; document.getElementById('psV').textContent = v; renderPts(); }

// Legend
var legDiv = document.getElementById('legItems');
var allLvs = [0].concat(LEVELS);
for (var i = 0; i < allLvs.length; i++) {
  var lmin = allLvs[i];
  var lmax = (i + 1 < allLvs.length) ? allLvs[i + 1] : lmin + 25;
  var col = lvColor((lmin + lmax) / 2);
  var d = document.createElement('div');
  d.className = 'lgi';
  var txt = (i + 1 < allLvs.length) ? (lmin + ' - ' + lmax + ' ppm') : ('>= ' + lmin + ' ppm');
  d.innerHTML = '<div class="lgs" style="background:'+col+'"></div><span>'+txt+'</span>';
  legDiv.appendChild(d);
}

// Map legend control
var legCtrl = L.control({position:'bottomright'});
legCtrl.onAdd = function() {
  var d = L.DomUtil.create('div','');
  d.style.cssText = 'background:rgba(13,27,42,0.94);border:1px solid #2a4a6a;border-radius:10px;padding:12px 16px;min-width:160px;backdrop-filter:blur(6px)';
  var items = '';
  var keyLvs = [0, 25, 50, 75, 100, 150, 200, 300];
  keyLvs.forEach(function(lv) {
    var c = lvColor(lv + 12);
    var lb = lv === 0 ? '0 - 25' : lv === 300 ? '>= 300' : lv + ' - ' + (lv + 25);
    items += '<div style="display:flex;align-items:center;gap:6px;margin:2px 0"><div style="width:18px;height:10px;border-radius:3px;background:'+c+';border:1px solid rgba(255,255,255,0.1)"></div><span style="font-size:10px;color:#b0bec5">'+lb+' ppm</span></div>';
  });
  d.innerHTML = '<div style="font-size:11px;color:#78909c;text-transform:uppercase;margin-bottom:6px;font-weight:600;letter-spacing:1px">Y ppm - Contornos</div>' + items +
    '<div style="margin-top:8px;font-size:9px;color:#546e7a"><div>&#9473;&#9473; Contorno principal (c/50)</div><div>&#9476;&#9476; Contorno secundario (c/25)</div><div>&#9679; Puntos de muestreo</div></div>';
  return d;
};
legCtrl.addTo(map);

// Fit bounds
var bounds = SAMPLES.map(function(s) { return [s.lat, s.lon]; });
if (bounds.length) map.fitBounds(bounds, {padding:[40,40]});

setTimeout(function() {
  map.invalidateSize();
  console.log('Map OK: samples=' + SAMPLES.length + ' filled=' + FILLED.features.length + ' lines=' + LINES.features.length);
}, 300);
</script>
</body>
</html>"""

# Replace placeholders (no braces involved!)
html = html.replace('%%FILLED_JSON%%', json_filled)
html = html.replace('%%LINES_JSON%%', json_lines)
html = html.replace('%%SAMPLES_JSON%%', json_samples)
html = html.replace('%%LEVELS_JSON%%', levels_json)
html = html.replace('%%CENTER_LAT%%', str(center_lat))
html = html.replace('%%CENTER_LON%%', str(center_lon))
html = html.replace('%%N_TOTAL%%', f"{n_total:,}")
html = html.replace('%%Y_MIN%%', f"{y_min_val:.1f}")
html = html.replace('%%Y_MED%%', f"{y_med_val:.1f}")
html = html.replace('%%Y_MEAN%%', f"{y_mean_val:.1f}")
html = html.replace('%%Y_MAX%%', f"{y_max_val:.1f}")
html = html.replace('%%N_ANOM%%', str(n_anom))

outpath = os.path.join(OUTDIR, "mapa_isoconcentraciones_Y.html")
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✅ Mapa generado: {outpath}")
print(f"📊 Tamaño: {os.path.getsize(outpath)/1024:.0f} KB")
print(f"🗺️  Contornos: {len(levels)} niveles")
print(f"📍 Puntos: {len(all_points)}")
