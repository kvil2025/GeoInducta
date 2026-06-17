import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 1. Estructura de la tabla principal
columns = [
    "Estacion_CP", 
    "Muestra_ID", 
    "Intervalo", 
    "Litologia", 
    "Mineralogia_Siglas", 
    "Caolinizacion_Caol", 
    "Oxidos_Fe_OxFe", 
    "Horizonte", 
    "Fecha_Hora", 
    "Coordenada_GPS", 
    "Observaciones", 
    "Geologo"
]

# Datos iniciales mapeados a la escala numérica 1-5
mock_data = [
    {
        "Estacion_CP": "Qui-CA-107", "Muestra_ID": "1140115", "Intervalo": "(0 - 1) m",
        "Litologia": "Granito", "Mineralogia_Siglas": "Qz ± bt", "Caolinizacion_Caol": "4 (+)",
        "Oxidos_Fe_OxFe": "2 (-)", "Horizonte": "LP-US", "Fecha_Hora": "2026-05-27 09:30:00",
        "Coordenada_GPS": "-37.5255, -72.6345", "Observaciones": "Sup (0-1): LP-US. Caolinización fuerte.", "Geologo": "Gro"
    },
    {
        "Estacion_CP": "Qui-CA-107", "Muestra_ID": "1140116", "Intervalo": "(1 - 2) m",
        "Litologia": "Granito", "Mineralogia_Siglas": "Qz ± bt", "Caolinizacion_Caol": "2 (-)",
        "Oxidos_Fe_OxFe": "1 (--)", "Horizonte": "UP", "Fecha_Hora": "2026-05-27 09:45:00",
        "Coordenada_GPS": "-37.5255, -72.6345", "Observaciones": "Inf (1-2): UP. Sin presencia de óxidos.", "Geologo": "Gro"
    }
]

df_main = pd.DataFrame(mock_data, columns=columns)

# 2. Listas de validación con escala numérica 1-5
litologias = ["Granito", "Pegmatita", "Granito con stockwork", "Otro"]
# Escala de intensidad geológica 1-5
intensidades = ["5 (++)", "4 (+)", "3 (±)", "2 (-)", "1 (--)"]
horizontes = ["UP", "LP", "US", "LS", "LP-US", "US/LP", "N/A"]

max_len = max(len(litologias), len(intensidades), len(horizontes))
df_val = pd.DataFrame({
    "Litologias_Ref": litologias + [""] * (max_len - len(litologias)),
    "Caolinizacion_Ref": intensidades + [""] * (max_len - len(intensidades)),
    "OxidosFe_Ref": intensidades + [""] * (max_len - len(intensidades)),
    "Horizontes_Ref": horizontes + [""] * (max_len - len(horizontes))
})

file_path = "Base_Datos_AppSheet.xlsx"
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df_main.to_excel(writer, sheet_name="Registro_Terreno", index=False)
    df_val.to_excel(writer, sheet_name="Listas_Validacion", index=False)

# Aplicar estilos
wb = openpyxl.load_workbook(file_path)
ws_main = wb["Registro_Terreno"]
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

for col in range(1, len(columns) + 1):
    cell = ws_main.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
for col in ws_main.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws_main.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(file_path)
print("¡Plantilla Base_Datos_AppSheet.xlsx actualizada con la escala geológica de 1 a 5!")
