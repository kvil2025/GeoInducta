import sys, os; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import pandas as pd

wb = openpyxl.load_workbook(r'c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\itrio\Datos de muestreo 11.06.xlsx', data_only=True)
ws = wb['2026 06 11']

print('=== Buscar Dy (Disprosio) y Tb (Terbio) en pXRF ===')
for c in range(1, ws.max_column+1):
    h = ws.cell(row=1, column=c).value
    if h:
        hs = str(h).strip()
        if hs in ('Dy', 'Tb', 'Gd', 'Er', 'Ho', 'Yb', 'Lu', 'Eu', 'Sm', 'Pr'):
            all_vals = []
            for r in range(2, ws.max_row+1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() not in ('ND','','None') and not str(v).startswith('<'):
                    try:
                        all_vals.append(float(v))
                    except: pass
            mx = max(all_vals) if all_vals else 0
            mn = sum(all_vals)/len(all_vals) if all_vals else 0
            print('  col {:3d}: {:4s} -> {} valores de {} filas | max={:.0f} media={:.0f}'.format(
                c, hs, len(all_vals), ws.max_row-1, mx, mn))

print()
df = pd.read_csv(r'c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\BD_Ytrio_LIMPIO.csv')
print('=== BD_Ytrio columnas REE ===')
for col in df.columns:
    if any(x in col.lower() for x in ['dy','tb','gd','er','ho','yb','lu','eu','sm','pr']):
        vals = pd.to_numeric(df[col], errors='coerce')
        print('  {}: {} valores, max={:.1f}, media={:.1f}'.format(col, vals.notna().sum(), vals.max(), vals.mean()))
