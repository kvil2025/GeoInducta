#!/usr/bin/env python3
"""
Script completo para generar Base_Datos_AppSheet_FINAL.xlsx
Proyecto: Inducta - Exploración Quinhue
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Columnas de la tabla principal ──────────────────────────
COLUMNS = [
    "Estacion_CP", "Muestra_ID", "Intervalo", "Sector",
    "Litologia", "Mineralogia_Siglas", "Horizonte",
    "Caolinizacion_Caol", "Oxidos_Fe_OxFe", "Estructuras",
    "Fecha", "Hora", "Coordenada_GPS",
    "UTM_Este", "UTM_Norte", "Elevacion_m",
    "Foto_URL", "Observaciones", "Geologo", "Estado_Sync"
]

# ─── Datos de ejemplo (3 filas reales de Quinhue) ────────────
SAMPLE_DATA = [
    {
        "Estacion_CP": "Qui-CA-107", "Muestra_ID": "1140115",
        "Intervalo": "(0-1) m", "Sector": "Este",
        "Litologia": "Granito", "Mineralogia_Siglas": "Qz \u00b1 bt",
        "Horizonte": "LP-US", "Caolinizacion_Caol": "4 (+)",
        "Oxidos_Fe_OxFe": "2 (-)", "Estructuras": "Sin estructura singular",
        "Fecha": "2026-05-27", "Hora": "09:30:00",
        "Coordenada_GPS": "-37.5255, -72.6345",
        "UTM_Este": 720345.0, "UTM_Norte": 5845210.0, "Elevacion_m": 133.6,
        "Foto_URL": "",
        "Observaciones": "Sup (0-1): LP-US. Alteracion fuerte de feldespatos. Muestra recolectada cerca acceso principal.",
        "Geologo": "Gro", "Estado_Sync": "Sincronizado"
    },
    {
        "Estacion_CP": "Qui-CA-112", "Muestra_ID": "1140126",
        "Intervalo": "Pared Oeste (0-1)", "Sector": "Oeste",
        "Litologia": "Pegmatita", "Mineralogia_Siglas": "Qz + Msc",
        "Horizonte": "LP", "Caolinizacion_Caol": "3 (\u00b1)",
        "Oxidos_Fe_OxFe": "3 (\u00b1)", "Estructuras": "Cuerpo Pegmatitico",
        "Fecha": "2026-05-27", "Hora": "11:15:00",
        "Coordenada_GPS": "-37.5260, -72.6350",
        "UTM_Este": 720290.0, "UTM_Norte": 5845155.0, "Elevacion_m": 141.2,
        "Foto_URL": "",
        "Observaciones": "Pared Oeste: Cambia lateralmente a cuerpo pegmatitico. Stockwork de Qz + Msc bien desarrollado.",
        "Geologo": "Gro", "Estado_Sync": "Sincronizado"
    },
    {
        "Estacion_CP": "Qui-CA-117", "Muestra_ID": "1140133",
        "Intervalo": "(0-1) m", "Sector": "Este",
        "Litologia": "Granito", "Mineralogia_Siglas": "Qz + bt + Msc",
        "Horizonte": "LP", "Caolinizacion_Caol": "3 (\u00b1)",
        "Oxidos_Fe_OxFe": "3 (\u00b1)", "Estructuras": "Pegmatita intruida",
        "Fecha": "2026-05-27", "Hora": "13:00:00",
        "Coordenada_GPS": "-37.5270, -72.6360",
        "UTM_Este": 720200.0, "UTM_Norte": 5845060.0, "Elevacion_m": 147.8,
        "Foto_URL": "",
        "Observaciones": "Zona superior intruida por pegmatita de Qz + Msc. Contacto neto con granito encajante.",
        "Geologo": "Gro", "Estado_Sync": "Sincronizado"
    }
]

# ─── Listas de validación ─────────────────────────────────────
LISTS = {
    "Litologias_Ref":    ["Granito", "Pegmatita", "Granito con stockwork", "Brecha", "Otro"],
    "Minerales_Ref":     ["Qz", "bt", "Msc", "Kfs", "Pl", "Qz+bt", "Qz+Msc", "Qz+bt+Msc", "Qz-bt", "Qz-Msc"],
    "Intensidad_Ref":    ["5 (++)", "4 (+)", "3 (\u00b1)", "2 (-)", "1 (--)"],
    "Horizonte_Ref":     ["UP", "LP", "US", "LS", "LP-US", "US/LP", "N/A"],
    "Sector_Ref":        ["Este", "Oeste"],
    "Estructuras_Ref":   ["Stockwork Qz+Msc", "Pegmatita intruida", "Cuerpo Pegmatitico", "Sin estructura singular", "Otra"]
}

# ─── Leyenda de abreviaturas ──────────────────────────────────
LEGEND = [
    ("Qz", "Cuarzo"),
    ("bt", "Biotita"),
    ("Msc", "Moscovita"),
    ("Kfs", "Feldespato potasico (K-Feldespato)"),
    ("Pl", "Plagioclasa"),
    ("Caol", "Caolinizacion - Alteracion arcillosa supergena de feldespatos"),
    ("OxFe", "Oxidos de Hierro (limonitas, hematita, goethita)"),
    ("UP", "Upper Profile - Perfil superior de la calicata"),
    ("LP", "Lower Profile - Perfil inferior de la calicata"),
    ("US", "Upper Surface - Superficie superior del afloramiento"),
    ("LS", "Lower Surface - Superficie inferior del afloramiento"),
    ("5 (++)", "Muy fuerte / Intensidad maxima de alteracion (valor numerico 5)"),
    ("4 (+)",  "Fuerte (valor numerico 4)"),
    ("3 (+-)", "Moderado / Presente con variacion (valor numerico 3)"),
    ("2 (-)",  "Debil (valor numerico 2)"),
    ("1 (--)", "Nulo / Sin alteracion observada (valor numerico 1)"),
    ("Gro",    "Geologo Responsable (firma de la estacion)"),
    ("GPS",    "Punto de control georreferenciado sin muestra asociada"),
]

def apply_header_style(ws, n_cols, fill_color="2C3E50"):
    hf = PatternFill("solid", fgColor=fill_color)
    hfont = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        vals = [str(c.value or "") for c in col]
        w = min(max(len(s) for s in vals) + 2, max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(w, min_w)

def build_excel():
    fp = "Base_Datos_AppSheet_FINAL.xlsx"

    # ── Sheet 1: Registro_Terreno ──────────────────────────────
    df_main = pd.DataFrame(SAMPLE_DATA, columns=COLUMNS)

    # ── Sheet 2: Listas_Validacion ────────────────────────────
    max_len = max(len(v) for v in LISTS.values())
    padded = {k: v + [""] * (max_len - len(v)) for k, v in LISTS.items()}
    df_lists = pd.DataFrame(padded)

    # ── Sheet 3: Leyenda ──────────────────────────────────────
    df_legend = pd.DataFrame(LEGEND, columns=["Sigla / Codigo", "Descripcion"])

    with pd.ExcelWriter(fp, engine="openpyxl") as writer:
        df_main.to_excel(writer, sheet_name="Registro_Terreno", index=False)
        df_lists.to_excel(writer, sheet_name="Listas_Validacion", index=False)
        df_legend.to_excel(writer, sheet_name="Leyenda", index=False)

    wb = openpyxl.load_workbook(fp)

    # ── Format Sheet 1 ────────────────────────────────────────
    ws1 = wb["Registro_Terreno"]
    apply_header_style(ws1, len(COLUMNS))
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    gray_fill = PatternFill("solid", fgColor="F8F9FA")
    bold_font = Font(name="Calibri", size=10, bold=True)
    normal_font = Font(name="Calibri", size=10)

    for row_idx in range(2, ws1.max_row + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = bold_font if col_idx == 1 else normal_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = gray_fill

    auto_width(ws1)
    ws1.column_dimensions["R"].width = 50  # Observaciones wide
    ws1.column_dimensions["H"].width = 14  # Caol
    ws1.column_dimensions["I"].width = 14  # OxFe

    # ── Format Sheet 2 ────────────────────────────────────────
    ws2 = wb["Listas_Validacion"]
    apply_header_style(ws2, len(LISTS), fill_color="1A5276")
    ws2.row_dimensions[1].height = 25
    auto_width(ws2)
    for row_idx in range(2, ws2.max_row + 1):
        for col_idx in range(1, len(LISTS) + 1):
            ws2.cell(row=row_idx, column=col_idx).font = Font(name="Calibri", size=10)
            ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="center")

    # ── Format Sheet 3 ────────────────────────────────────────
    ws3 = wb["Leyenda"]
    apply_header_style(ws3, 2, fill_color="154360")
    ws3.row_dimensions[1].height = 25
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 60
    for row_idx in range(2, ws3.max_row + 1):
        ws3.cell(row=row_idx, column=1).font = Font(name="Calibri", size=10, bold=True, color="154360")
        ws3.cell(row=row_idx, column=2).font = Font(name="Calibri", size=10)
        ws3.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)

    wb.save(fp)
    print(f"OK - Base_Datos_AppSheet_FINAL.xlsx generado con exito!")
    print(f"   Hoja 1: Registro_Terreno   - {len(SAMPLE_DATA)} filas de ejemplo, {len(COLUMNS)} columnas")
    print(f"   Hoja 2: Listas_Validacion  - {len(LISTS)} columnas de referencia para AppSheet")
    print(f"   Hoja 3: Leyenda            - {len(LEGEND)} abreviaturas documentadas")

build_excel()
