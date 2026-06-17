import openpyxl, sys, collections
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\itrio\Datos de muestreo 09.06.xlsx', data_only=True)
ws = wb['2026 06 09']

# Headers completos col 79+
print('=== Headers cols 79-110 ===')
for col in range(79, min(ws.max_column+1, 120)):
    v = ws.cell(row=1, column=col).value
    if v:
        print(f'  Col {col:3d}: {str(v).strip()[:60]}')

# Analizar Sample IDs unicos
print('\n=== Sample IDs ===')
samples = {}
for r in range(2, ws.max_row+1):
    sid = ws.cell(row=r, column=1).value
    y_val = ws.cell(row=r, column=7).value
    if sid:
        sid = str(sid).strip()
        if sid not in samples:
            samples[sid] = {'count':0, 'y_vals':[]}
        samples[sid]['count'] += 1
        if y_val and y_val != 'ND':
            try:
                samples[sid]['y_vals'].append(float(y_val))
            except:
                pass

print(f'Total filas: {ws.max_row-1}')
print(f'Samples unicos: {len(samples)}')
print(f'\nPrimeros 20 Sample IDs:')
for i, (sid, info) in enumerate(samples.items()):
    if i >= 20: break
    y_avg = sum(info['y_vals'])/len(info['y_vals']) if info['y_vals'] else 0
    print(f'  {sid:20s} -> {info["count"]:2d} mediciones, Y_avg={y_avg:.1f} ppm')

# Estadísticas de Y
print('\n=== Estadisticas Y (ppm) ===')
all_y = []
nd_count = 0
for r in range(2, ws.max_row+1):
    y = ws.cell(row=r, column=7).value
    if y == 'ND' or y is None:
        nd_count += 1
    else:
        try:
            all_y.append(float(y))
        except:
            nd_count += 1

print(f'  Valores numericos: {len(all_y)}')
print(f'  ND (no detectado): {nd_count}')
if all_y:
    all_y.sort()
    print(f'  Min: {min(all_y):.1f}')
    print(f'  Max: {max(all_y):.1f}')
    print(f'  Media: {sum(all_y)/len(all_y):.1f}')
    print(f'  Mediana: {all_y[len(all_y)//2]:.1f}')
    print(f'  P95: {all_y[int(len(all_y)*0.95)]:.1f}')
    print(f'  Y >= 50: {sum(1 for y in all_y if y >= 50)}')
    print(f'  Y >= 100: {sum(1 for y in all_y if y >= 100)}')

# REE disponibles
print('\n=== Elementos REE disponibles ===')
ree = ['Y','Ce','La','Nd','Pr','Dy','Er','Eu','Gd','Ho','Lu','Sm','Tb','Tm','Yb']
for elem in ree:
    found = False
    for col in range(1, ws.max_column+1):
        h = ws.cell(row=1, column=col).value
        if h and str(h).strip() == elem:
            # Count non-ND values
            vals = 0
            for r in range(2, min(ws.max_row+1, 100)):
                v = ws.cell(row=r, column=col).value
                if v and v != 'ND':
                    try:
                        float(v)
                        vals += 1
                    except:
                        pass
            print(f'  {elem:3s} -> Col {col}, ~{vals} valores en primeras 100 filas')
            found = True
            break
    if not found:
        print(f'  {elem:3s} -> NO encontrado')
