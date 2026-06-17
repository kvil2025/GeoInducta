#!/usr/bin/env python3
"""
Exporta el mapa de isoconcentraciones de Y (ppm) en:
  1. GeoTIFF (raster interpolado)
  2. Shapefile de contornos (líneas)
  3. Shapefile de contornos rellenos (polígonos)

CRS de salida: EPSG:32718 (UTM Zona 18S) — compatible con software GIS.
"""
import pandas as pd
import numpy as np
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import pyproj
import openpyxl
from scipy.interpolate import griddata
from scipy.ndimage import gaussian_filter
from scipy.spatial import cKDTree
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import rasterio
from rasterio.transform import from_bounds
from rasterio.crs import CRS
import geopandas as gpd
from shapely.geometry import Polygon, LineString, Point

OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"
EXPORT_DIR = os.path.join(OUTDIR, "export_gis")
os.makedirs(EXPORT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════
# 1. CARGAR TODOS LOS DATOS
# ══════════════════════════════════════════════════════════════════
transformer = pyproj.Transformer.from_crs('EPSG:32718', 'EPSG:4326', always_xy=True)

# --- Datos existentes ---
df = pd.read_csv(os.path.join(OUTDIR, "BD_Ytrio_LIMPIO.csv"))
df = df.dropna(subset=['UTM_E','UTM_N','Y_ppm'])
df = df[(df['UTM_E'] > 100000) & (df['UTM_N'] > 1000000)].copy()
lons_old, lats_old = transformer.transform(df['UTM_E'].values, df['UTM_N'].values)
df['lat'] = lats_old
df['lon'] = lons_old
df = df[(df['lat'] > -60) & (df['lat'] < -20) & (df['lon'] > -80) & (df['lon'] < -60)]

utm_e_old = df['UTM_E'].values
utm_n_old = df['UTM_N'].values
y_old = df['Y_ppm'].values
samples_old = df['Sample'].values
print(f"Datos existentes: {len(utm_e_old)} puntos")

# --- Datos nuevos ---
wb = openpyxl.load_workbook(os.path.join(OUTDIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']
utm_e_new, utm_n_new, y_new, samples_new_list = [], [], [], []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    cp, xm, ym, yppm = row[0], row[3], row[4], row[56]
    if xm is None or ym is None or yppm is None or not isinstance(yppm, (int, float)):
        continue
    utm_e_new.append(float(xm))
    utm_n_new.append(float(ym))
    y_new.append(float(yppm))
    samples_new_list.append(str(cp))

print(f"Datos nuevos: {len(utm_e_new)} puntos")

# --- Combinar ---
utm_es = np.concatenate([utm_e_old, utm_e_new])
utm_ns = np.concatenate([utm_n_old, utm_n_new])
vals = np.concatenate([y_old, y_new])
all_samples = list(samples_old) + samples_new_list

print(f"Total: {len(vals)} puntos")

# ══════════════════════════════════════════════════════════════════
# 2. INTERPOLACIÓN EN UTM (resolución 200m para GeoTIFF)
# ══════════════════════════════════════════════════════════════════
margin_m = 2000
x_min = float(utm_es.min() - margin_m)
x_max = float(utm_es.max() + margin_m)
y_min = float(utm_ns.min() - margin_m)
y_max = float(utm_ns.max() + margin_m)

res = 200  # 200m resolution
nx = int((x_max - x_min) / res)
ny = int((y_max - y_min) / res)
nx = min(nx, 500)
ny = min(ny, 500)

print(f"Grid: {nx} x {ny} = {nx*ny:,} celdas ({res}m)")

xi = np.linspace(x_min, x_max, nx)
yi = np.linspace(y_min, y_max, ny)
XI, YI = np.meshgrid(xi, yi)

ZI = griddata((utm_es, utm_ns), vals, (XI, YI), method='linear')
ZI_nearest = griddata((utm_es, utm_ns), vals, (XI, YI), method='nearest')

tree = cKDTree(np.column_stack([utm_es, utm_ns]))
grid_points = np.column_stack([XI.ravel(), YI.ravel()])
distances, _ = tree.query(grid_points)
dist_grid = distances.reshape(XI.shape)

max_dist = 5000
mask_fill = np.isnan(ZI) & (dist_grid < max_dist)
ZI[mask_fill] = ZI_nearest[mask_fill]

ZI_smooth = gaussian_filter(ZI, sigma=1.8)
ZI_smooth[dist_grid > max_dist] = np.nan

print(f"Interpolación OK. NaN: {np.isnan(ZI_smooth).sum():,} / {ZI_smooth.size:,}")

# ══════════════════════════════════════════════════════════════════
# 3. EXPORTAR GeoTIFF
# ══════════════════════════════════════════════════════════════════
print("\n── Exportando GeoTIFF ──")

# Flip Y axis (GeoTIFF: top-to-bottom)
raster_data = np.flipud(ZI_smooth).astype(np.float32)

transform = from_bounds(x_min, y_min, x_max, y_max, nx, ny)
crs = CRS.from_epsg(32718)

geotiff_path = os.path.join(EXPORT_DIR, "Y_ppm_interpolado.tif")
with rasterio.open(
    geotiff_path, 'w',
    driver='GTiff',
    height=ny, width=nx,
    count=1,
    dtype='float32',
    crs=crs,
    transform=transform,
    nodata=-9999.0,
    compress='lzw',
) as dst:
    # Replace NaN with nodata
    data_out = raster_data.copy()
    data_out[np.isnan(data_out)] = -9999.0
    dst.write(data_out, 1)
    dst.update_tags(
        DESCRIPTION='Ytrio (Y) interpolado en ppm',
        SOURCE='BD_Ytrio_LIMPIO.csv + BD_GEOL_2026',
        METHOD='scipy.griddata linear + gaussian_filter',
        RESOLUTION_M=str(res),
        MAX_INTERP_DIST_M=str(max_dist),
        N_SAMPLES=str(len(vals)),
    )

print(f"✅ GeoTIFF: {geotiff_path}")
print(f"   Tamaño: {os.path.getsize(geotiff_path)/1024:.0f} KB")
print(f"   CRS: EPSG:32718 (UTM 18S)")
print(f"   Resolución: {res}m")
print(f"   Dimensiones: {nx} x {ny}")

# ══════════════════════════════════════════════════════════════════
# 4. GENERAR CONTORNOS Y EXPORTAR SHAPEFILES
# ══════════════════════════════════════════════════════════════════
print("\n── Generando contornos ──")

max_level = min(int(np.nanmax(ZI_smooth)), 600)
levels = list(range(25, max_level + 25, 25))
if levels[-1] < np.nanmax(vals):
    levels.append(levels[-1] + 25)
print(f"Niveles: {levels}")

fig, ax = plt.subplots()
# Contour lines in UTM
cs_lines = ax.contour(XI, YI, ZI_smooth, levels=levels)
# Contour fill in UTM
cs_fill = ax.contourf(XI, YI, ZI_smooth, levels=[0] + levels, extend='max')
plt.close(fig)

# ── 4a. Shapefile de líneas de contorno ──
print("\n── Exportando Shapefile de líneas ──")
line_records = []
for i, segs in enumerate(cs_lines.allsegs):
    level = float(cs_lines.levels[i])
    for seg in segs:
        if len(seg) >= 2:
            coords = [(float(c[0]), float(c[1])) for c in seg]
            line_records.append({
                'geometry': LineString(coords),
                'Y_ppm': level,
                'tipo': 'contorno',
            })

if line_records:
    gdf_lines = gpd.GeoDataFrame(line_records, crs='EPSG:32718')
    shp_lines_path = os.path.join(EXPORT_DIR, "contornos_Y_lineas.shp")
    gdf_lines.to_file(shp_lines_path, driver='ESRI Shapefile', encoding='utf-8')
    print(f"✅ Shapefile líneas: {shp_lines_path}")
    print(f"   Features: {len(gdf_lines)}")

# ── 4b. Shapefile de polígonos rellenos ──
print("\n── Exportando Shapefile de polígonos ──")
poly_records = []
fill_levels_all = list(cs_fill.levels)
for i, segs in enumerate(cs_fill.allsegs):
    level_min = fill_levels_all[i] if i < len(fill_levels_all) else levels[-1]
    level_max = fill_levels_all[i+1] if i+1 < len(fill_levels_all) else level_min + 25
    for seg in segs:
        if len(seg) >= 3:
            coords = [(float(c[0]), float(c[1])) for c in seg]
            try:
                poly = Polygon(coords)
                if poly.is_valid and poly.area > 0:
                    poly_records.append({
                        'geometry': poly,
                        'Y_min': float(level_min),
                        'Y_max': float(level_max),
                        'label': f"{int(level_min)}-{int(level_max)}",
                    })
            except:
                pass

if poly_records:
    gdf_polys = gpd.GeoDataFrame(poly_records, crs='EPSG:32718')
    shp_polys_path = os.path.join(EXPORT_DIR, "contornos_Y_poligonos.shp")
    gdf_polys.to_file(shp_polys_path, driver='ESRI Shapefile', encoding='utf-8')
    print(f"✅ Shapefile polígonos: {shp_polys_path}")
    print(f"   Features: {len(gdf_polys)}")

# ── 4c. Shapefile de puntos de muestreo ──
print("\n── Exportando Shapefile de puntos ──")
lons_all, lats_all = transformer.transform(utm_es, utm_ns)
point_records = []
for i in range(len(vals)):
    point_records.append({
        'geometry': Point(float(utm_es[i]), float(utm_ns[i])),
        'Sample': str(all_samples[i]),
        'Y_ppm': float(vals[i]),
        'UTM_E': float(utm_es[i]),
        'UTM_N': float(utm_ns[i]),
        'Lat': float(lats_all[i]),
        'Lon': float(lons_all[i]),
    })

gdf_pts = gpd.GeoDataFrame(point_records, crs='EPSG:32718')
shp_pts_path = os.path.join(EXPORT_DIR, "muestras_Y_puntos.shp")
gdf_pts.to_file(shp_pts_path, driver='ESRI Shapefile', encoding='utf-8')
print(f"✅ Shapefile puntos: {shp_pts_path}")
print(f"   Features: {len(gdf_pts)}")

# ══════════════════════════════════════════════════════════════════
# 5. RESUMEN
# ══════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("EXPORTACIÓN COMPLETA")
print("="*60)
print(f"\n📁 Directorio: {EXPORT_DIR}")
print(f"\nArchivos generados:")
for f in sorted(os.listdir(EXPORT_DIR)):
    fpath = os.path.join(EXPORT_DIR, f)
    sz = os.path.getsize(fpath)
    print(f"  📄 {f:40s}  {sz/1024:8.1f} KB")

print(f"\n🗺️  CRS: EPSG:32718 (UTM Zona 18S, WGS84)")
print(f"📊 GeoTIFF resolución: {res}m")
print(f"📍 Puntos: {len(vals)}")
print(f"〰️  Contornos: {len(levels)} niveles ({', '.join(str(l) for l in levels)} ppm)")
