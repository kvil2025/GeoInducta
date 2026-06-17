import sys, os; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\geolo\Downloads\Datos de muestreo 11.06.xlsx', data_only=True)
print('Hojas:', wb.sheetnames)

for sn in wb.sheetnames[:3]:
    ws = wb[sn]
    print('\n=== Hoja: {} ==='.format(sn))
    print('Filas: {}, Columnas: {}'.format(ws.max_row, ws.max_column))
    
    # Headers relevantes
    targets = ['Y', 'Ce', 'La', 'Nd', 'Th', 'Fe', 'Ti', 'Sample']
    print('Columnas clave:')
    for c in range(1, min(ws.max_column+1, 250)):
        h = ws.cell(row=1, column=c).value
        if h:
            hs = str(h).strip()
            if hs in targets or c <= 6:
                vals = [ws.cell(row=r, column=c).value for r in range(2, min(6, ws.max_row+1))]
                print('  col {:3d}: {:10s} -> {}'.format(c, hs, vals))
    
    # Sample IDs
    sids = set()
    for r in range(2, ws.max_row+1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip() not in ('', 'Ejrmplo', 'Prueba'):
            sids.add(str(v).strip().split('_')[0])
    print('Sample IDs unicos: {}'.format(len(sids)))
    if sids:
        print('Ejemplo: {}'.format(sorted(list(sids))[:10]))

# Comparar con datos existentes
import pandas as pd
existing = pd.read_csv(os.path.join(r'c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\itrio', 'BD_INTEGRADA_2026.csv'))
existing_pxrf = existing[existing['FUENTE']=='pXRF_2026']['Sample'].str.replace('pXRF-','',regex=False).unique()
print('\n=== COMPARACION ===')
print('pXRF existentes en BD: {}'.format(len(existing_pxrf)))

ws = wb[wb.sheetnames[0]]
new_ids = set()
for r in range(2, ws.max_row+1):
    v = ws.cell(row=r, column=1).value
    if v and str(v).strip() not in ('', 'Ejrmplo', 'Prueba'):
        new_ids.add(str(v).strip().split('_')[0])

overlap = new_ids & set(existing_pxrf)
only_new = new_ids - set(existing_pxrf)
print('IDs en archivo nuevo: {}'.format(len(new_ids)))
print('Overlap con BD: {}'.format(len(overlap)))
print('Solo nuevos: {}'.format(len(only_new)))
if only_new:
    print('Nuevos IDs: {}'.format(sorted(list(only_new))[:20]))
