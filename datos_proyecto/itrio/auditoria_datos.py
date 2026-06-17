#!/usr/bin/env python3
"""
==========================================================================
AUDITORÍA DE CONSISTENCIA DE DATOS
- Verificar separadores decimales (coma vs punto)
- Verificar unidades (ppm vs %, absolutos vs relativos)
- Verificar rangos coherentes
- Detectar problemas de mapeo de columnas
==========================================================================
"""
import pandas as pd
import numpy as np
import openpyxl
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"

# ═══════════════════════════════════════════════════════════════════
# 1. CARGAR POB1 (CSV)
# ═══════════════════════════════════════════════════════════════════
print("=" * 80)
print("  AUDITORÍA POBLACIÓN 1: BD_Ytrio_LIMPIO.csv")
print("=" * 80)

df1 = pd.read_csv(os.path.join(OUTDIR, "BD_Ytrio_LIMPIO.csv"))
print(f"Filas: {len(df1)}, Columnas: {len(df1.columns)}")

# Columnas de interés con sus unidades esperadas
pob1_cols = {
    'K__':    {'unidad_esperada': '%',   'rango_tipico': (0, 10)},
    'K':      {'unidad_esperada': 'ppm', 'rango_tipico': (0, 100000)},
    'Ca_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 100000)},
    'Ti__':   {'unidad_esperada': '%',   'rango_tipico': (0, 5)},
    'Ti':     {'unidad_esperada': 'ppm', 'rango_tipico': (0, 50000)},
    'V_ppm':  {'unidad_esperada': 'ppm', 'rango_tipico': (0, 2000)},
    'Cr_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 2000)},
    'Mn_':    {'unidad_esperada': '%',   'rango_tipico': (0, 1)},
    'Mn':     {'unidad_esperada': 'ppm', 'rango_tipico': (0, 10000)},
    'Fe__':   {'unidad_esperada': '%',   'rango_tipico': (0, 20)},
    'Fe':     {'unidad_esperada': 'ppm', 'rango_tipico': (0, 200000)},
    'Y_ppm':  {'unidad_esperada': 'ppm', 'rango_tipico': (0, 500)},
    'Ce_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 2000)},
    'Th_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 500)},
    'La_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 1000)},
    'Pr_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 2000)},
    'Nd_ppm': {'unidad_esperada': 'ppm', 'rango_tipico': (0, 3000)},
}

print(f"\n{'Variable':12s} {'Unidad':6s} {'N':>6s} {'NaN':>6s} {'Min':>10s} {'Max':>10s} {'Media':>10s} {'Mediana':>10s} {'P5':>10s} {'P95':>10s} {'OK?':>5s}")
print("-" * 105)
for col, info in pob1_cols.items():
    if col not in df1.columns:
        print(f"{col:12s} — NO EXISTE EN POB1")
        continue
    vals = pd.to_numeric(df1[col], errors='coerce')
    n = vals.notna().sum()
    na = vals.isna().sum()
    if n == 0:
        print(f"{col:12s} {info['unidad_esperada']:6s} {n:>6d} {na:>6d} — TODO NaN")
        continue
    mn, mx = vals.min(), vals.max()
    mean = vals.mean()
    med = vals.median()
    p5, p95 = vals.quantile(0.05), vals.quantile(0.95)
    rmin, rmax = info['rango_tipico']
    ok = "✅" if rmin <= mn and mx <= rmax * 2 else "⚠️"
    print(f"{col:12s} {info['unidad_esperada']:6s} {n:>6d} {na:>6d} {mn:>10.2f} {mx:>10.2f} {mean:>10.2f} {med:>10.2f} {p5:>10.2f} {p95:>10.2f} {ok:>5s}")

# Verificar relación K__ (%) * 10000 ≈ K (ppm)
print("\n--- Verificación de conversión K__ (%) vs K (ppm) ---")
k_pct = df1['K__'].dropna()
k_ppm = df1['K'].dropna()
ratio_k = (df1['K'] / (df1['K__'] * 10000)).dropna()
print(f"  K(ppm) / (K(%)*10000) → media={ratio_k.mean():.4f}, mediana={ratio_k.median():.4f}")
print(f"  Esperado ≈ 1.0 si K__ es % y K es ppm")

# Verificar Ti
print("\n--- Verificación de conversión Ti__ (%) vs Ti (ppm) ---")
ratio_ti = (df1['Ti'] / (df1['Ti__'] * 10000)).dropna()
print(f"  Ti(ppm) / (Ti(%)*10000) → media={ratio_ti.mean():.4f}, mediana={ratio_ti.median():.4f}")

# Verificar Fe
print("\n--- Verificación de conversión Fe__ (%) vs Fe (ppm) ---")
ratio_fe = (df1['Fe'] / (df1['Fe__'] * 10000)).dropna()
print(f"  Fe(ppm) / (Fe(%)*10000) → media={ratio_fe.mean():.4f}, mediana={ratio_fe.median():.4f}")

# Verificar Mn
print("\n--- Verificación de conversión Mn_ (%) vs Mn (ppm) ---")
ratio_mn = (df1['Mn'] / (df1['Mn_'] * 10000)).dropna()
print(f"  Mn(ppm) / (Mn(%)*10000) → media={ratio_mn.mean():.4f}, mediana={ratio_mn.median():.4f}")


# ═══════════════════════════════════════════════════════════════════
# 2. CARGAR POB2 (XLSX) - EXAMEN EXHAUSTIVO COLUMNA POR COLUMNA
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  AUDITORÍA POBLACIÓN 2: BD_GEOL_2026 (1).xlsx")
print("=" * 80)

wb = openpyxl.load_workbook(os.path.join(OUTDIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']

# Leer TODOS los headers (fila 1)
headers_raw = []
for cell in ws[1]:
    headers_raw.append(cell.value)

# Imprimir todos los headers con índice
print(f"\nTotal columnas en XLSX: {len(headers_raw)}")
print(f"\n{'Idx':>4s} {'Header':40s} {'Tipo datos':15s} {'N valores':>10s} {'Ejemplo 1':>15s} {'Ejemplo 2':>15s} {'Min':>12s} {'Max':>12s}")
print("-" * 130)

col_data = {}
for col_idx in range(len(headers_raw)):
    header = headers_raw[col_idx]
    vals = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx + 1)
        vals.append(cell.value)
    
    non_null = [v for v in vals if v is not None]
    
    if len(non_null) == 0:
        tipo = "VACÍA"
        ejemplo1 = ejemplo2 = "—"
        mn = mx = "—"
    else:
        # Detectar tipo
        tipos = set()
        for v in non_null[:20]:
            if isinstance(v, (int, float)):
                tipos.add('num')
            elif isinstance(v, str):
                if v.strip().startswith('<'):
                    tipos.add('censored')
                else:
                    tipos.add('str')
            else:
                tipos.add(type(v).__name__)
        
        tipo = '+'.join(sorted(tipos))
        ejemplo1 = str(non_null[0])[:15] if len(non_null) > 0 else "—"
        ejemplo2 = str(non_null[1])[:15] if len(non_null) > 1 else "—"
        
        # Si numérico, calcular min/max
        nums = []
        for v in non_null:
            if isinstance(v, (int, float)):
                nums.append(float(v))
            elif isinstance(v, str) and v.strip().startswith('<'):
                try:
                    nums.append(float(v.strip().lstrip('<')))
                except:
                    pass
        
        if nums:
            mn = f"{min(nums):.4f}"
            mx = f"{max(nums):.4f}"
        else:
            mn = mx = "—"
    
    col_data[col_idx] = {'header': header, 'vals': vals, 'non_null': non_null}
    print(f"{col_idx:>4d} {str(header):40s} {tipo:15s} {len(non_null):>10d} {ejemplo1:>15s} {ejemplo2:>15s} {mn:>12s} {mx:>12s}")


# ═══════════════════════════════════════════════════════════════════
# 3. COMPARAR COLUMNAS EQUIVALENTES EN DETALLE
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  COMPARACIÓN DETALLADA DE COLUMNAS EQUIVALENTES")
print("=" * 80)

# Mapeo manual basado en el header del XLSX
# Necesitamos saber qué columna del XLSX corresponde a qué columna del CSV
# En procesar_bd_geol.py se usaba:
#   K=col23, Ca=col25, Ti=col27, V=col29, Cr=col31, Mn=col33, Fe=col35
#   Y_ppm=col56, Y_pond=col57

print("\n--- Headers del XLSX en posiciones clave (del script procesar_bd_geol.py) ---")
key_positions = {
    17: 'Al2O3', 19: 'SiO2', 21: 'S', 
    23: 'K', 25: 'Ca', 27: 'Ti', 29: 'V', 31: 'Cr', 33: 'Mn', 35: 'Fe',
    55: 'Y%', 56: 'Y_ppm', 57: 'Y_pond'
}

for idx, expected in key_positions.items():
    actual = headers_raw[idx] if idx < len(headers_raw) else "FUERA DE RANGO"
    
    # Obtener valores para esta columna
    vals = col_data.get(idx, {}).get('non_null', [])
    nums = [float(v) for v in vals if isinstance(v, (int, float))]
    
    if nums:
        info = f"n={len(nums):>3d}, min={min(nums):.4f}, max={max(nums):.4f}, media={np.mean(nums):.4f}, mediana={np.median(nums):.4f}"
    else:
        info = f"n=0 (no numérico)"
    
    print(f"  col[{idx:2d}] esperado='{expected:8s}' → header='{str(actual):30s}' | {info}")


# ═══════════════════════════════════════════════════════════════════
# 4. VERIFICAR UNIDADES CRUZANDO RANGOS
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  VERIFICACIÓN DE UNIDADES POR RANGO DE VALORES")
print("=" * 80)

def get_nums_from_xlsx_col(col_idx):
    """Extraer valores numéricos de una columna del XLSX"""
    vals = col_data.get(col_idx, {}).get('non_null', [])
    nums = []
    for v in vals:
        if isinstance(v, (int, float)):
            nums.append(float(v))
        elif isinstance(v, str) and v.strip().startswith('<'):
            try:
                nums.append(float(v.strip().lstrip('<')) / 2)  # Half detection limit
            except:
                pass
    return np.array(nums) if nums else np.array([])

# Tabla de comparación detallada
comparisons = [
    {'name': 'K', 'pob1_col': 'K__', 'pob1_unit': '%', 'pob2_idx': 23},
    {'name': 'Ca', 'pob1_col': 'Ca_ppm', 'pob1_unit': 'ppm', 'pob2_idx': 25},
    {'name': 'Ti', 'pob1_col': 'Ti__', 'pob1_unit': '%', 'pob2_idx': 27},
    {'name': 'V', 'pob1_col': 'V_ppm', 'pob1_unit': 'ppm', 'pob2_idx': 29},
    {'name': 'Cr', 'pob1_col': 'Cr_ppm', 'pob1_unit': 'ppm', 'pob2_idx': 31},
    {'name': 'Mn', 'pob1_col': 'Mn_', 'pob1_unit': '%', 'pob2_idx': 33},
    {'name': 'Fe', 'pob1_col': 'Fe__', 'pob1_unit': '%', 'pob2_idx': 35},
    {'name': 'Y', 'pob1_col': 'Y_ppm', 'pob1_unit': 'ppm', 'pob2_idx': 56},
]

print(f"\n{'Elem':>5s} | {'Pob1 col':>10s} {'Pob1 unit':>10s} {'Pob1 med':>10s} {'Pob1 rango':>20s} | {'Pob2 header':>25s} {'Pob2 med':>10s} {'Pob2 rango':>20s} | {'Diagnóstico'}")
print("-" * 150)

for comp in comparisons:
    # Pob1
    v1 = pd.to_numeric(df1[comp['pob1_col']], errors='coerce').dropna().values
    
    # Pob2
    v2 = get_nums_from_xlsx_col(comp['pob2_idx'])
    pob2_header = str(headers_raw[comp['pob2_idx']]) if comp['pob2_idx'] < len(headers_raw) else "N/A"
    
    if len(v1) > 0 and len(v2) > 0:
        med1 = np.median(v1)
        med2 = np.median(v2)
        rango1 = f"[{np.min(v1):.2f} - {np.max(v1):.2f}]"
        rango2 = f"[{np.min(v2):.2f} - {np.max(v2):.2f}]"
        
        # Diagnóstico automático
        ratio = med2 / med1 if med1 != 0 else float('inf')
        
        if 0.1 < ratio < 10:
            diag = f"✅ Misma escala (ratio={ratio:.2f})"
        elif 100 < ratio < 100000:
            diag = f"⚠️ Pob2 podría estar en ppm, Pob1 en % (ratio={ratio:.0f})"
        elif 0.00001 < ratio < 0.01:
            diag = f"⚠️ Pob2 podría estar en %, Pob1 en ppm (ratio={ratio:.6f})"
        elif ratio > 100000:
            diag = f"🔴 ESCALA MUY DIFERENTE (ratio={ratio:.0f})"
        else:
            diag = f"❓ Ratio inusual ({ratio:.4f})"
        
        print(f"{comp['name']:>5s} | {comp['pob1_col']:>10s} {comp['pob1_unit']:>10s} {med1:>10.3f} {rango1:>20s} | {pob2_header:>25s} {med2:>10.3f} {rango2:>20s} | {diag}")
    else:
        print(f"{comp['name']:>5s} | {comp['pob1_col']:>10s} — datos insuficientes")


# ═══════════════════════════════════════════════════════════════════
# 5. EXAMINAR HEADERS REALES DEL XLSX PARA DETECTAR UNIDADES
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  HEADERS CON UNIDADES DEL XLSX (filas de contexto)")
print("=" * 80)

# A veces el Excel tiene una fila de unidades. Revisar fila 1 completa y fila 2
print("\nFila 1 (headers):")
for i, h in enumerate(headers_raw):
    if h is not None:
        print(f"  [{i:2d}] {h}")

# Revisar si hay sub-headers (a veces las unidades van en otra fila)
print("\nFila 2 (primera fila de datos):")
for i, cell in enumerate(ws[2]):
    if cell.value is not None:
        print(f"  [{i:2d}] {cell.value}")

# ═══════════════════════════════════════════════════════════════════
# 6. BUSCAR EL PATRÓN DE PARES % + ppm EN EL XLSX
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  DETECTANDO PATRONES DE COLUMNAS PAREADAS (% y ppm)")
print("=" * 80)

# En Pob1, K__ y K van en pares (% y ppm respectivamente)
# ¿Sucede lo mismo en Pob2?

print("\nPob1 - Columnas pareadas:")
pairs_pob1 = [('K__', 'K'), ('Ti__', 'Ti'), ('Mn_', 'Mn'), ('Fe__', 'Fe')]
for pct, ppm in pairs_pob1:
    v_pct = df1[pct].dropna()
    v_ppm = df1[ppm].dropna()
    ratio = (v_ppm / (v_pct * 10000)).dropna()
    print(f"  {pct:6s} (%) ↔ {ppm:6s} (ppm): ratio medio = {ratio.mean():.4f} (esperado ≈ 1.0)")

print(f"\nPob2 - Buscando pares en XLSX:")
# Para cada par de columnas consecutivas, ver si una es ~10000x la otra
for i in range(len(headers_raw) - 1):
    h1 = str(headers_raw[i]) if headers_raw[i] else ""
    h2 = str(headers_raw[i+1]) if headers_raw[i+1] else ""
    
    v1 = get_nums_from_xlsx_col(i)
    v2 = get_nums_from_xlsx_col(i + 1)
    
    if len(v1) >= 5 and len(v2) >= 5:
        # Verificar si son un par %/ppm
        with np.errstate(divide='ignore', invalid='ignore'):
            ratio_vals = v2[:min(len(v1), len(v2))] / (v1[:min(len(v1), len(v2))] * 10000)
            ratio_vals = ratio_vals[np.isfinite(ratio_vals)]
        
        if len(ratio_vals) > 0 and 0.8 < np.median(ratio_vals) < 1.2:
            print(f"  ✅ col[{i}] '{h1}' ↔ col[{i+1}] '{h2}': PAREJA %/ppm confirmada (ratio med = {np.median(ratio_vals):.3f})")

# ═══════════════════════════════════════════════════════════════════
# 7. VERIFICAR VALORES CENSURADOS ("<X")
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  VALORES CENSURADOS (bajo límite de detección)")
print("=" * 80)

print("\nPob2 - Valores con '<' por columna:")
for idx in range(len(headers_raw)):
    vals = col_data.get(idx, {}).get('non_null', [])
    censored = [v for v in vals if isinstance(v, str) and v.strip().startswith('<')]
    if censored:
        unique_limits = set(censored)
        print(f"  col[{idx:2d}] '{str(headers_raw[idx]):25s}': {len(censored)} censurados → {sorted(unique_limits)}")

print("\nPob1 - Valores = 0.5 (posiblemente censurados) por columna:")
for col in ['Y_ppm', 'Ce_ppm', 'Th_ppm', 'La_ppm', 'Pr_ppm', 'Nd_ppm']:
    if col in df1.columns:
        n05 = (df1[col] == 0.5).sum()
        if n05 > 0:
            print(f"  {col:12s}: {n05} valores = 0.5 ({n05/len(df1)*100:.1f}%)")


# ═══════════════════════════════════════════════════════════════════
# 8. DIAGNÓSTICO FINAL - TABLA DE MAPEO CORRECTO
# ═══════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 80)
print("  DIAGNÓSTICO FINAL DE MAPEO")
print("=" * 80)

# Voy a comparar las medianas para determinar qué columnas del XLSX
# deberían mapearse a qué columnas del CSV, considerando conversiones

print("\nBuscando la mejor correspondencia para cada variable de Pob1...")

for comp in comparisons:
    v1 = pd.to_numeric(df1[comp['pob1_col']], errors='coerce').dropna().values
    if len(v1) == 0:
        continue
    
    med1 = np.median(v1)
    unit1 = comp['pob1_unit']
    
    print(f"\n  {comp['name']} ({comp['pob1_col']}, unidad={unit1}, mediana={med1:.4f}):")
    
    # Buscar en el XLSX columnas que contengan el nombre del elemento
    candidates = []
    elem_lower = comp['name'].lower()
    
    for idx in range(len(headers_raw)):
        h = str(headers_raw[idx]).lower() if headers_raw[idx] else ""
        if elem_lower in h or (elem_lower == 'k' and h.strip() in ['k', 'k(%)', 'k (%)', 'k_pct', 'k (%)']) or \
           (elem_lower == 'y' and ('ytrio' in h or h.strip() in ['y', 'y(ppm)', 'y ppm', 'y_ppm'])):
            v2 = get_nums_from_xlsx_col(idx)
            if len(v2) >= 3:
                med2 = np.median(v2)
                ratio = med2 / med1 if med1 != 0 else float('inf')
                candidates.append((idx, headers_raw[idx], med2, ratio, len(v2)))
    
    # También probar el índice que usamos en el script anterior
    v2_used = get_nums_from_xlsx_col(comp['pob2_idx'])
    if len(v2_used) >= 3:
        med2_used = np.median(v2_used)
        ratio_used = med2_used / med1 if med1 != 0 else float('inf')
        print(f"    → col[{comp['pob2_idx']}] (USADA en script): header='{headers_raw[comp['pob2_idx']]}', "
              f"mediana={med2_used:.4f}, ratio vs Pob1={ratio_used:.4f}")
        
        if 0.1 < ratio_used < 10:
            print(f"      ✅ ESCALA COHERENTE")
        elif abs(ratio_used - 10000) / 10000 < 0.5 or abs(ratio_used - 0.0001) / 0.0001 < 0.5:
            print(f"      ⚠️ DIFERENCIA DE FACTOR ~10000 → probablemente una en % y otra en ppm")
        else:
            print(f"      🔴 ESCALA INCOHERENTE")
    
    for idx, header, med2, ratio, n in candidates:
        flag = "✅" if 0.1 < ratio < 10 else "⚠️"
        print(f"    → col[{idx}] '{header}': mediana={med2:.4f}, n={n}, ratio={ratio:.4f} {flag}")


print("\n\n" + "=" * 80)
print("  RESUMEN DE PROBLEMAS DETECTADOS")  
print("=" * 80)
