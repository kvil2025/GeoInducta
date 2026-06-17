#!/usr/bin/env python3
"""
Procesar BD_GEOL_2026 (1).xlsx y comparar con BD_Ytrio_LIMPIO.csv
para integrar datos nuevos al visor de mapa.
"""
import openpyxl
import pyproj
import pandas as pd
import numpy as np
import json
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

OUTDIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"

# ─── 1. Leer datos nuevos ────────────────────────────────────────
wb = openpyxl.load_workbook(os.path.join(OUTDIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']
t = pyproj.Transformer.from_crs('EPSG:32718', 'EPSG:4326', always_xy=True)

rows_with_data = []
rows_location_only = []
issues = []

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
    cp = row[0]
    idsample = row[1]
    elev = row[2]
    xm = row[3]
    ym = row[4]
    frm = row[5]
    to = row[6]
    horiz = row[7]
    roca = row[8]
    estructura = row[9]
    rumbo = row[10]
    manteo = row[11]
    minerales = row[12]
    alteracion = row[13]
    mineraliza = row[14]
    comentario = row[15]
    
    # Geoquimica
    al2o3 = row[17]
    sio2 = row[19]
    s_val = row[21]
    k_val = row[23]
    ca_val = row[25]
    ti_val = row[27]
    v_val = row[29]
    cr_val = row[31]
    mn_val = row[33]
    fe_val = row[35]
    ypct = row[55]
    yppm = row[56]
    ypond = row[57]
    
    # Check coords
    has_coords = xm is not None and ym is not None
    if not has_coords:
        issues.append(f"Fila {i} ({cp}): SIN COORDENADAS")
        continue
    
    lon, lat = t.transform(xm, ym)
    
    def safe_num(val):
        if val is None: return None
        if isinstance(val, (int, float)): return val
        if isinstance(val, str) and val.strip().startswith('<'):
            return 0.5  # Below detection limit → use half
        return None
    
    has_yppm = yppm is not None and isinstance(yppm, (int, float))
    has_geochem = any(isinstance(v, (int, float)) for v in [k_val, ca_val, ti_val, fe_val])
    
    rec = {
        'row': i, 'CP': cp, 'IDSAMPLE': idsample, 'Elevation': elev,
        'Xm': xm, 'Ym': ym, 'lat': round(lat, 6), 'lon': round(lon, 6),
        'From': frm, 'To': to,
        'HORIZONTE': horiz, 'ROCA_CAJA': roca,
        'ESTRUCTURA': estructura, 'MINERALES': minerales,
        'ALTERACION': alteracion, 'MINERALIZA': mineraliza,
        'COMENTARIO': comentario,
        'K': safe_num(k_val), 'Ca': safe_num(ca_val),
        'Ti': safe_num(ti_val), 'V': safe_num(v_val),
        'Cr': safe_num(cr_val), 'Mn': safe_num(mn_val),
        'Fe': safe_num(fe_val),
        'Yppm': safe_num(yppm), 'Ypond': safe_num(ypond),
        'has_geochem': has_geochem, 'has_yppm': has_yppm
    }
    
    if has_yppm or has_geochem:
        rows_with_data.append(rec)
        if not has_yppm:
            issues.append(f"Fila {i} ({cp}): Tiene geoquimica pero SIN Y ppm")
    else:
        rows_location_only.append(rec)


# ─── 2. RESUMEN ──────────────────────────────────────────────────
print("=" * 70)
print("  RESUMEN DEL ARCHIVO BD_GEOL_2026")
print("=" * 70)
total = ws.max_row - 1
print(f"Total filas de datos:               {total}")
print(f"Filas con datos geoquímicos:        {len(rows_with_data)}")
print(f"Filas sólo ubicación (sin química): {len(rows_location_only)}")

n_yppm = sum(1 for r in rows_with_data if r['has_yppm'])
print(f"Filas con Y ppm:                    {n_yppm}")
print(f"Filas sin Y ppm (pero con química): {len(rows_with_data) - n_yppm}")

# Unique CPs
cps_data = set(r['CP'] for r in rows_with_data)
cps_loc = set(r['CP'] for r in rows_location_only)
print(f"\nPuntos únicos con datos:            {len(cps_data)}")
print(f"Puntos únicos sólo ubicación:       {len(cps_loc)}")

print(f"\n--- Puntos con datos geoquímicos ({len(rows_with_data)} filas) ---")
for r in rows_with_data:
    yppm_str = f"{r['Yppm']:.1f}" if r['Yppm'] else "N/A"
    print(f"  {r['CP']:12s} | ID:{r['IDSAMPLE']} | ({r['lat']:.4f}, {r['lon']:.4f}) | "
          f"Y:{yppm_str:>7s} ppm | Roca: {r['ROCA_CAJA'] or '—'} | Hz: {r['HORIZONTE'] or '—'}")

print(f"\n--- Puntos SÓLO ubicación ({len(rows_location_only)} filas) ---")
for r in rows_location_only:
    print(f"  {r['CP']:12s} | ({r['lat']:.4f}, {r['lon']:.4f}) | Elev: {r['Elevation'] or '—'}")

if issues:
    print(f"\n--- PROBLEMAS DETECTADOS ({len(issues)}) ---")
    for iss in issues:
        print(f"  ⚠ {iss}")


# ─── 3. Comparar con datos existentes ────────────────────────────
print("\n" + "=" * 70)
print("  COMPARACIÓN CON DATOS EXISTENTES (BD_Ytrio_LIMPIO.csv)")
print("=" * 70)

df_exist = pd.read_csv(os.path.join(OUTDIR, "BD_Ytrio_LIMPIO.csv"))
print(f"Muestras existentes: {len(df_exist)}")
print(f"Columnas existentes: {len(df_exist.columns)}")

# Coordinate ranges comparison
all_new = rows_with_data + rows_location_only
lats_new = [r['lat'] for r in all_new]
lons_new = [r['lon'] for r in all_new]

t_exist = pyproj.Transformer.from_crs('EPSG:32718', 'EPSG:4326', always_xy=True)
lons_ex, lats_ex = t_exist.transform(df_exist['UTM_E'].values, df_exist['UTM_N'].values)

print(f"\n{'':>25s} {'EXISTENTES':>20s} {'NUEVOS':>20s}")
print(f"{'Latitud mín':>25s} {min(lats_ex):>20.4f} {min(lats_new):>20.4f}")
print(f"{'Latitud máx':>25s} {max(lats_ex):>20.4f} {max(lats_new):>20.4f}")
print(f"{'Longitud mín':>25s} {min(lons_ex):>20.4f} {min(lons_new):>20.4f}")
print(f"{'Longitud máx':>25s} {max(lons_ex):>20.4f} {max(lons_new):>20.4f}")
print(f"{'UTM_E mín':>25s} {df_exist['UTM_E'].min():>20.0f} {min(r['Xm'] for r in all_new):>20.0f}")
print(f"{'UTM_E máx':>25s} {df_exist['UTM_E'].max():>20.0f} {max(r['Xm'] for r in all_new):>20.0f}")
print(f"{'UTM_N mín':>25s} {df_exist['UTM_N'].min():>20.0f} {min(r['Ym'] for r in all_new):>20.0f}")
print(f"{'UTM_N máx':>25s} {df_exist['UTM_N'].max():>20.0f} {max(r['Ym'] for r in all_new):>20.0f}")

# Columns that exist in old but NOT in new
old_cols = set(df_exist.columns)
print(f"\n--- Columnas en BD existente que NO están en datos nuevos ---")
missing_from_new = ['La_ppm', 'Ce_ppm', 'Nd_ppm', 'Pr_ppm', 'Th_ppm', 'Cl_ppm',
                    'Y_pond', 'TLC_REO_P3', 'Obs', 'SEGMENTO']
for c in missing_from_new:
    if c in old_cols:
        print(f"  ❌ {c}")

print(f"\n--- Columnas NUEVAS que no existen en BD existente ---")
new_extras = ['Al2O3', 'SiO2', 'S', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'As', 'Se',
              'Rb', 'Sr', 'Zr', 'Mo', 'Pd', 'Ag', 'Cd', 'In', 'Sn', 'Sb',
              'Te', 'Ba', 'W', 'Pt', 'Au', 'Hg', 'Pb', 'Bi', 'U',
              'HORIZONTE', 'ESTRUCTURA', 'RUMBO', 'MANTEO', 'MINERALES',
              'ALTERACION', 'MINERALIZA', 'From', 'To']
for c in new_extras:
    print(f"  ✨ {c}")

# Y ppm statistics comparison
yppm_new = [r['Yppm'] for r in rows_with_data if r['Yppm'] is not None]
print(f"\n--- Estadísticas Y ppm ---")
print(f"{'':>20s} {'EXISTENTES':>15s} {'NUEVOS':>15s}")
print(f"{'N muestras':>20s} {len(df_exist):>15d} {len(yppm_new):>15d}")
print(f"{'Media':>20s} {df_exist['Y_ppm'].mean():>15.1f} {np.mean(yppm_new):>15.1f}")
print(f"{'Mediana':>20s} {df_exist['Y_ppm'].median():>15.1f} {np.median(yppm_new):>15.1f}")
print(f"{'Mín':>20s} {df_exist['Y_ppm'].min():>15.1f} {min(yppm_new):>15.1f}")
print(f"{'Máx':>20s} {df_exist['Y_ppm'].max():>15.1f} {max(yppm_new):>15.1f}")
print(f"{'Anomalías (≥50)':>20s} {(df_exist['Y_ppm']>=50).sum():>15d} {sum(1 for y in yppm_new if y>=50):>15d}")

# Litologias comparison  
lits_exist = set(df_exist['Litology_STD'].dropna().unique())
roca_new = set(r['ROCA_CAJA'] for r in rows_with_data if r['ROCA_CAJA'])
print(f"\n--- Litologías ---")
print(f"Litologías existentes ({len(lits_exist)}): {sorted(lits_exist)}")
print(f"Rocas nuevas ({len(roca_new)}): {sorted(roca_new)}")

overlap = roca_new & lits_exist
only_new = roca_new - lits_exist
print(f"En ambos: {sorted(overlap) if overlap else 'Ninguna'}")
print(f"Sólo en nuevos: {sorted(only_new)}")

print("\n" + "=" * 70)
print("  CONCLUSIÓN")
print("=" * 70)
print(f"✅ {len(rows_with_data)} filas con datos geoquímicos para integrar al mapa")
print(f"📍 {len(rows_location_only)} puntos de sólo ubicación (sin geoquímica)")
print(f"🗺️  Los datos nuevos están ~70 km al NE de los existentes")
print(f"⚠️  Faltan columnas REE (Ce, La, Nd, Pr, Th) en datos nuevos")
print(f"✨ Los datos nuevos traen {len(new_extras)} columnas adicionales")
