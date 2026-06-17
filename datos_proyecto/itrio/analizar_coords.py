"""Cruce detallado entre pXRF (Datos de muestreo 09.06.xlsx) y BD_GEOL_2026_06_09.xls"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import numpy as np

# 1. Cargar BD_GEOL con coordenadas
df_geo = pd.read_excel(r'G:\Mi unidad\BD_GEOL_2026_06_09.xls', engine='xlrd', sheet_name='BD_29May26')
print(f'BD_GEOL: {len(df_geo)} filas, {df_geo["IDSAMPLE"].nunique()} IDs únicos')

# 2. Cargar pXRF
import openpyxl
wb = openpyxl.load_workbook(r'c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\itrio\Datos de muestreo 09.06.xlsx', data_only=True)
ws = wb['2026 06 09']

# Extraer headers
headers = {}
for col in range(1, ws.max_column+1):
    v = ws.cell(row=1, column=col).value
    if v:
        headers[col] = str(v).strip()

# Extraer datos pXRF
pxrf_rows = []
for r in range(2, ws.max_row+1):
    row_data = {}
    for ci, h in headers.items():
        v = ws.cell(row=r, column=ci).value
        row_data[h] = v
    pxrf_rows.append(row_data)

df_pxrf = pd.DataFrame(pxrf_rows)
print(f'pXRF: {len(df_pxrf)} filas, {df_pxrf["Sample ID"].nunique()} IDs únicos')

# 3. Normalizar IDs para cruce
geo_ids = set()
for v in df_geo['IDSAMPLE'].dropna():
    geo_ids.add(str(int(v)))

pxrf_ids = set()
for v in df_pxrf['Sample ID'].dropna():
    s = str(v).strip()
    # Quitar sufijos como _1, _2
    base = s.split('_')[0]
    pxrf_ids.add(base)

matches = geo_ids & pxrf_ids
print(f'\n=== CRUCE ===')
print(f'IDs en BD_GEOL: {len(geo_ids)}')
print(f'IDs base en pXRF: {len(pxrf_ids)}')
print(f'COINCIDENCIAS: {len(matches)}')

no_match_pxrf = pxrf_ids - geo_ids
no_match_geo = geo_ids - pxrf_ids
print(f'pXRF sin coords: {len(no_match_pxrf)} -> {sorted(no_match_pxrf)[:20]}')
print(f'GEOL sin pXRF: {len(no_match_geo)} IDs')

# 4. Para los matches, mostrar detalle
print(f'\n=== Detalle de matches (primeros 15) ===')
for sid in sorted(list(matches))[:15]:
    sid_int = int(sid)
    geo_row = df_geo[df_geo['IDSAMPLE'] == sid_int].iloc[0]
    pxrf_vals = df_pxrf[df_pxrf['Sample ID'].astype(str).str.split('_').str[0] == sid]
    
    y_vals = []
    for v in pxrf_vals['Y']:
        if v and v != 'ND':
            try: y_vals.append(float(v))
            except: pass
    
    y_avg = np.mean(y_vals) if y_vals else None
    y_geo = geo_row.get('Y ppm', None)
    
    print(f'  {sid}: CP={geo_row["CP"]}, UTM=({geo_row["Xm"]:.0f}, {geo_row["Ym"]:.0f}), '
          f'Horiz={geo_row.get("HORIZONTE","")}, Roca={geo_row.get("ROCA CAJA","")}, '
          f'Y_geo={y_geo}, Y_pxrf_avg={y_avg:.1f if y_avg else "ND"}, n_pxrf={len(pxrf_vals)}')

# 5. Estadísticas de los datos con match
print(f'\n=== Estadísticas de datos con coordenadas ===')
matched_y = []
for sid in matches:
    pxrf_vals = df_pxrf[df_pxrf['Sample ID'].astype(str).str.split('_').str[0] == sid]
    for v in pxrf_vals['Y']:
        if v and v != 'ND':
            try: matched_y.append(float(v))
            except: pass

print(f'Total mediciones con coords: {len(matched_y)}')
if matched_y:
    matched_y.sort()
    print(f'  Min: {min(matched_y):.1f}')
    print(f'  Max: {max(matched_y):.1f}')
    print(f'  Media: {np.mean(matched_y):.1f}')
    print(f'  Y >= 50: {sum(1 for y in matched_y if y >= 50)}')
    print(f'  Y >= 100: {sum(1 for y in matched_y if y >= 100)}')

# 6. Elementos adicionales que existen en pXRF pero no en BD_Ytrio
print(f'\n=== Elementos pXRF adicionales para el visor ===')
key_elements = ['Y', 'Ce', 'La', 'Nd', 'Th', 'Fe', 'Ti', 'Pr', 'Ba', 'Sr', 'Zr', 'Rb', 'Cu', 'Zn', 'Pb']
for elem in key_elements:
    if elem in df_pxrf.columns:
        vals = []
        for v in df_pxrf[elem]:
            if v and v != 'ND':
                try: vals.append(float(v))
                except: pass
        print(f'  {elem:4s}: {len(vals)} valores numéricos de {len(df_pxrf)}, media={np.mean(vals):.1f}' if vals else f'  {elem}: 0 valores')
