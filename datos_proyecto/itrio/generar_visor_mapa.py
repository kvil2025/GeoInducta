#!/usr/bin/env python3
"""
Genera el visor de mapa unificado con datos existentes (BD_Ytrio_LIMPIO.csv)
+ datos nuevos (BD_GEOL_2026 (1).xlsx) + datos pXRF (Datos de muestreo 09.06.xlsx).
Transforma coordenadas UTM 18S → WGS84 y genera HTML con Leaflet.
"""
import pandas as pd
import numpy as np
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
import pyproj
import openpyxl

DATADIR = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001"
OUTDIR  = r"c:\Users\geolo\OneDrive\Documentos\INDUCTA\ANTI\001\itrio"

# ══════════════════════════════════════════════════════════════════
# 1. CARGAR DATOS EXISTENTES
# ══════════════════════════════════════════════════════════════════
df = pd.read_csv(os.path.join(DATADIR, "BD_Ytrio_LIMPIO.csv"))
df['FLAG_OUTLIER'] = df['FLAG_OUTLIER'].fillna('')
df['Litology_STD'] = df['Litology_STD'].fillna('SIN_ASIGNAR')
df['FLAG_DUPLICADO'] = df['FLAG_DUPLICADO'].fillna('')
df['FUENTE'] = 'BD_Ytrio'

transformer = pyproj.Transformer.from_crs('EPSG:32718', 'EPSG:4326', always_xy=True)

cols_old = ['Sample','UTM_E','UTM_N','COTA_M','Y_ppm','Y_pond',
            'Ce_ppm','La_ppm','Th_ppm','Nd_ppm','Pr_ppm','Fe__','Ti__',
            'Litology_STD','FLAG_OUTLIER','FLAG_DUPLICADO','FUENTE']
data_old = df[cols_old].dropna(subset=['UTM_E','UTM_N','Y_ppm']).copy()
data_old['CP'] = ''  # BD_Ytrio no tiene CP
data_old['HORIZONTE'] = ''
data_old['ROCA_CAJA'] = ''
data_old = data_old[(data_old['UTM_E'] > 100000) & (data_old['UTM_N'] > 1000000)].copy()
data_old = data_old.fillna('')

lons_old, lats_old = transformer.transform(data_old['UTM_E'].values, data_old['UTM_N'].values)
data_old['lat'] = np.round(lats_old, 6)
data_old['lon'] = np.round(lons_old, 6)
data_old = data_old[(data_old['lat'] > -60) & (data_old['lat'] < -20) &
                    (data_old['lon'] > -80) & (data_old['lon'] < -60)].copy()

print(f"Datos existentes válidos: {len(data_old)}")

# ══════════════════════════════════════════════════════════════════
# 2. CARGAR DATOS NUEVOS (BD_GEOL_2026)
# ══════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(os.path.join(DATADIR, "BD_GEOL_2026 (1).xlsx"), data_only=True)
ws = wb['Hoja1']

new_rows = []
location_only_rows = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    cp = row[0]
    idsample = row[1]
    elev = row[2]
    xm = row[3]
    ym = row[4]
    horiz = row[7]
    roca = row[8]
    
    # Geoquimica
    k_val = row[23]
    ca_val = row[25]
    ti_val = row[27]
    mn_val = row[33]
    fe_val = row[35]
    yppm = row[56]
    ypond = row[57]
    
    if xm is None or ym is None:
        continue
    
    lon, lat = transformer.transform(xm, ym)
    
    # Normalizar litología
    lit_map = {
        'Granito': 'Granito',
        'Granito de Biotita': 'Granito de Biotita',
        'Granito de Biotita?': 'Granito de Biotita',
        'Granodiorita': 'Granodiorita',
        'Diorita': 'Diorita',
        'Diotrita?': 'Diorita',
        'Aplita?': 'Aplita',
        'Kaolinita': 'Kaolinita',
    }
    lit_std = lit_map.get(roca, roca) if roca else 'SIN_ASIGNAR'
    
    def safe_num(val):
        if val is None: return ''
        if isinstance(val, (int, float)): return val
        if isinstance(val, str) and val.strip().startswith('<'):
            return 0.5
        return ''
    
    has_yppm = isinstance(yppm, (int, float))
    has_geochem = any(isinstance(v, (int, float)) for v in [k_val, ca_val, ti_val, fe_val])
    
    # Sample name: CP + horizonte para diferenciar sub-muestras
    sample_name = f"{cp}" if not horiz else f"{cp}-{horiz}"
    
    rec = {
        'Sample': sample_name,
        'UTM_E': xm,
        'UTM_N': ym,
        'COTA_M': elev if elev else '',
        'Y_ppm': safe_num(yppm),
        'Y_pond': safe_num(ypond),
        'Ce_ppm': '',  # No disponible
        'La_ppm': '',  # No disponible  
        'Th_ppm': '',  # No disponible
        'Nd_ppm': '',  # No disponible
        'Pr_ppm': '',  # No disponible
        'Fe__': safe_num(fe_val),
        'Ti__': safe_num(ti_val),
        'Litology_STD': lit_std,
        'FLAG_OUTLIER': '',
        'FLAG_DUPLICADO': '',
        'FUENTE': 'BD_GEOL_2026',
        'lat': round(lat, 6),
        'lon': round(lon, 6),
        'HORIZONTE': horiz or '',
        'ROCA_CAJA': roca or '',
    }
    
    if has_yppm:
        new_rows.append(rec)
    elif has_geochem:
        new_rows.append(rec)
    else:
        # Punto sólo ubicación
        rec['Y_ppm'] = ''
        location_only_rows.append(rec)

df_new = pd.DataFrame(new_rows)
df_loc = pd.DataFrame(location_only_rows)

print(f"Datos nuevos con geoquímica: {len(df_new)}")
print(f"Puntos sólo ubicación: {len(df_loc)}")

# ══════════════════════════════════════════════════════════════════
# 2b. CARGAR DATOS pXRF (Datos de muestreo 10.06.xlsx) — ACTUALIZADO
# ══════════════════════════════════════════════════════════════════
PXRF_FILE  = os.path.join(OUTDIR, "Datos de muestreo 11.06.xlsx")
COORDS_FILE = r"G:\Mi unidad\BD_GEOL_2026_06_10.xls"

# --- Leer mediciones pXRF (openpyxl) ---
wb_pxrf = openpyxl.load_workbook(PXRF_FILE, data_only=True)
ws_pxrf = wb_pxrf['2026 06 11']

# Columnas del archivo 10.06 (por nombre de header): Y=6, Ce=120, La=118, Nd=124, Th=184, Fe=58, Ti=50
PXRF_COLS = {
    'Y_ppm':  5,   # col 6, 0-based = 5
    'Ce_ppm': 119, # col 120, 0-based = 119
    'La_ppm': 117, # col 118, 0-based = 117
    'Nd_ppm': 123, # col 124, 0-based = 123
    'Th_ppm': 183, # col 184, 0-based = 183
    'Fe__':   57,  # col 58, 0-based = 57
    'Ti__':   49,  # col 50, 0-based = 49
}

# Acumular mediciones por Sample ID
from collections import defaultdict
pxrf_meas = defaultdict(lambda: defaultdict(list))  # {sample_id: {col: [values]}}

for row in ws_pxrf.iter_rows(min_row=2, max_row=ws_pxrf.max_row, values_only=True):
    sid_raw = row[0]
    if sid_raw is None:
        continue
    # Normalizar Sample ID (puede ser int o str con _1 sufijo)
    sid = str(sid_raw).strip()
    for col_name, col_idx in PXRF_COLS.items():
        val = row[col_idx]
        if val is None or (isinstance(val, str) and val.strip().upper() == 'ND'):
            continue
        try:
            pxrf_meas[sid][col_name].append(float(val))
        except (ValueError, TypeError):
            continue

print(f"\npXRF: {sum(len(v) for v in pxrf_meas.values())} mediciones de {len(pxrf_meas)} muestras únicas")

# --- Leer coordenadas (xlrd via pandas) ---
df_coords = pd.read_excel(COORDS_FILE, sheet_name='BD_29May26', engine='xlrd')
# Crear lookup por IDSAMPLE (como entero string)
# Mapa de normalización de litologías
lit_map_pxrf = {
    'Granito': 'Granito', 'Granito ': 'Granito',
    'Granito de Biotita': 'Granito de Biotita',
    'Granito de Biotita?': 'Granito de Biotita',
    'Granodiorita': 'Granodiorita',
    'Diorita': 'Diorita', 'Diotrita?': 'Diorita',
    'Diorita-Metapelita': 'Diorita', 'Diorita - Metapelita': 'Diorita',
    'Aplita': 'Aplita', 'Aplita?': 'Aplita',
    'Kaolinita': 'Kaolinita',
    'Sedimentaria': 'Sedimentaria', 'Sedimentario': 'Sedimentaria',
    'Metapelita': 'Metapelita',
}

coord_lookup = {}
for _, crow in df_coords.iterrows():
    idsample = crow.get('IDSAMPLE')
    if pd.isna(idsample):
        continue
    key = str(int(idsample))
    xm = crow.get('Xm')
    ym = crow.get('Ym')
    if pd.isna(xm) or pd.isna(ym) or xm == -999 or ym == -999:
        continue
    
    roca_raw = str(crow.get('ROCA CAJA', '')).strip() if not pd.isna(crow.get('ROCA CAJA')) else ''
    lit_std = lit_map_pxrf.get(roca_raw, roca_raw) if roca_raw else 'SIN_ASIGNAR'
    
    ypond_val = crow.get('Ypond', '')
    ypond_val = round(float(ypond_val), 1) if not pd.isna(ypond_val) and ypond_val > 0 else ''
    
    coord_lookup[key] = {
        'Xm': float(xm),
        'Ym': float(ym),
        'CP': str(crow.get('CP', '')) if not pd.isna(crow.get('CP')) else '',
        'HORIZONTE': str(crow.get('HORIZONTE', '')) if not pd.isna(crow.get('HORIZONTE')) else '',
        'ROCA_CAJA': roca_raw,
        'Litology_STD': lit_std,
        'COTA_M': float(crow.get('Elevation', '')) if not pd.isna(crow.get('Elevation')) and crow.get('Elevation') != -999 else '',
        'Ypond': ypond_val,
    }

print(f"Coordenadas disponibles: {len(coord_lookup)} puntos")

# --- Promediar réplicas y construir registros ---
pxrf_rows = []
pxrf_no_coords = 0

for sid, measurements in pxrf_meas.items():
    # Buscar coordenadas: probar ID base (sin _1 etc)
    sid_base = sid.split('_')[0]
    coords = coord_lookup.get(sid_base) or coord_lookup.get(sid)
    if not coords:
        pxrf_no_coords += 1
        continue
    
    xm, ym = coords['Xm'], coords['Ym']
    lon, lat = transformer.transform(xm, ym)
    
    # Verificar rango geográfico
    if not (-60 < lat < -20 and -80 < lon < -60):
        continue
    
    # Promediar cada variable
    rec = {
        'Sample': f"pXRF-{sid_base}",
        'UTM_E': xm,
        'UTM_N': ym,
        'COTA_M': coords.get('COTA_M', ''),
        'Y_ppm': '',
        'Y_pond': coords.get('Ypond', ''),
        'Ce_ppm': '',
        'La_ppm': '',
        'Th_ppm': '',
        'Nd_ppm': '',
        'Pr_ppm': '',
        'Fe__': '',
        'Ti__': '',
        'Litology_STD': coords.get('Litology_STD', 'SIN_ASIGNAR'),
        'FLAG_OUTLIER': '',
        'FLAG_DUPLICADO': '',
        'FUENTE': 'pXRF_2026',
        'lat': round(lat, 6),
        'lon': round(lon, 6),
        'HORIZONTE': coords.get('HORIZONTE', ''),
        'ROCA_CAJA': coords.get('ROCA_CAJA', ''),
        'CP': coords.get('CP', ''),
    }
    
    # Usar MÁXIMO de cada variable (no promedio)
    for col_name, vals_list in measurements.items():
        if vals_list:
            rec[col_name] = round(max(vals_list), 2)
    
    # Si tiene al menos un valor, agregar
    has_any = any(isinstance(rec.get(c), (int, float)) for c in PXRF_COLS)
    if has_any:
        pxrf_rows.append(rec)

df_pxrf = pd.DataFrame(pxrf_rows)
print(f"pXRF con coordenadas y datos: {len(df_pxrf)}")
print(f"pXRF sin coordenadas (descartadas): {pxrf_no_coords}")

# ══════════════════════════════════════════════════════════════════
# 3. COMBINAR
# ══════════════════════════════════════════════════════════════════
# Asegurar que todos tengan las mismas columnas
all_dfs = [data_old, df_new]
if len(df_pxrf) > 0:
    all_dfs.append(df_pxrf)

for col in ['HORIZONTE', 'ROCA_CAJA']:
    if col not in data_old.columns:
        data_old[col] = ''

# Recopilar todas las columnas
common_cols = list(data_old.columns)
for adf in all_dfs:
    for c in adf.columns:
        if c not in common_cols:
            common_cols.append(c)

# Asegurar que cada df tenga todas las columnas
for adf in all_dfs:
    for c in common_cols:
        if c not in adf.columns:
            adf[c] = ''

data_all = pd.concat([adf[common_cols] for adf in all_dfs], ignore_index=True)
data_all = data_all.fillna('')

# ── Calcular REE_Global (índice ponderado de tierras raras) ──
# Pesos basados en criticidad/valor económico: Nd(0.35) + Pr(0.25) + Y(0.20) + La(0.10) + Ce(0.10)
def calc_ree_global(row):
    weights = {'Nd_ppm': 0.35, 'Pr_ppm': 0.25, 'Y_ppm': 0.20, 'La_ppm': 0.10, 'Ce_ppm': 0.10}
    total = 0
    has_any = False
    for col, w in weights.items():
        val = row[col]
        if isinstance(val, (int, float)) and val > 0:
            total += val * w
            has_any = True
    return round(total, 2) if has_any else ''

data_all['REE_Global'] = data_all.apply(calc_ree_global, axis=1)
ree_valid = [v for v in data_all['REE_Global'] if isinstance(v, (int, float))]
if ree_valid:
    print(f"\nREE_Global: min={min(ree_valid):.1f}, max={max(ree_valid):.1f}, mean={sum(ree_valid)/len(ree_valid):.1f} ({len(ree_valid)} muestras con valor)")

print(f"\nTotal combinado: {len(data_all)} muestras")
print(f"  - Existentes: {len(data_old)}")
print(f"  - Nuevas (geoquímica): {len(df_new)}")
print(f"  - pXRF 2026: {len(df_pxrf)}")

# ══════════════════════════════════════════════════════════════════
# 4. PREPARAR DATOS PARA MAPA
# ══════════════════════════════════════════════════════════════════
center_lat = float(np.median(data_all['lat']))
center_lon = float(np.median(data_all['lon']))
print(f"Centro: lat={center_lat:.4f}, lon={center_lon:.4f}")
print(f"Rango lat: {data_all['lat'].min():.3f} a {data_all['lat'].max():.3f}")
print(f"Rango lon: {data_all['lon'].min():.3f} a {data_all['lon'].max():.3f}")

records = data_all.to_dict(orient='records')
json_data = json.dumps(records, ensure_ascii=False)

# Puntos sólo ubicación
loc_records = df_loc.to_dict(orient='records') if len(df_loc) > 0 else []
json_loc = json.dumps(loc_records, ensure_ascii=False)

# Litologías
lits = sorted(data_all['Litology_STD'].unique().tolist())
n = len(lits)
lit_colors_py = {l: f"hsl({round(i*360/n)},70%,55%)" for i,l in enumerate(lits)}
lit_colors_json = json.dumps(lit_colors_py)

n_exist = len(data_old)
n_new = len(df_new)
n_pxrf = len(df_pxrf)
n_loc = len(df_loc)
total = len(data_all)

# Cargar fotos vinculadas
fotos_json_path = os.path.join(OUTDIR, "fotos_por_muestra.json")
if os.path.exists(fotos_json_path):
    with open(fotos_json_path, 'r', encoding='utf-8') as fp:
        fotos_por_muestra = json.load(fp)
    n_fotos = sum(len(v) for v in fotos_por_muestra.values())
    print(f"\nFotos vinculadas: {n_fotos} fotos de {len(fotos_por_muestra)} muestras")
else:
    fotos_por_muestra = {}
    n_fotos = 0
fotos_json_str = json.dumps(fotos_por_muestra, ensure_ascii=False)

# ══════════════════════════════════════════════════════════════════
# 5. GENERAR HTML
# ══════════════════════════════════════════════════════════════════
html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Visor Geoquímico — REE Global | Chile (Integrado + pXRF)</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Segoe UI',sans-serif; background:#0f1117; color:#e0e0e0; display:flex; flex-direction:column; height:100vh; overflow:hidden; }}
  .foto-gallery {{ display:flex; gap:6px; flex-wrap:wrap; margin-top:8px; }}
  .foto-gallery img {{ width:100%; max-height:180px; object-fit:cover; border-radius:6px; border:1px solid #2a4a6a; cursor:pointer; transition:transform 0.2s; }}
  .foto-gallery img:hover {{ transform:scale(1.03); border-color:#4fc3f7; }}
  .foto-lightbox {{ position:fixed; top:0; left:0; width:100vw; height:100vh; background:rgba(0,0,0,0.92); z-index:9999; display:flex; align-items:center; justify-content:center; cursor:pointer; }}
  .foto-lightbox img {{ max-width:90vw; max-height:90vh; border-radius:8px; box-shadow:0 0 40px rgba(0,0,0,0.8); }}

  #header {{ background:linear-gradient(135deg,#0d1b2a,#1a2a3a); padding:10px 20px; display:flex; align-items:center; justify-content:space-between; border-bottom:2px solid #1e3a5a; flex-shrink:0; z-index:1000; }}
  #header h1 {{ font-size:17px; font-weight:700; color:#4fc3f7; letter-spacing:1px; }}
  #header .sub {{ font-size:11px; color:#78909c; margin-top:2px; }}
  .pill {{ background:#1e2a3a; border:1px solid #2a4a6a; border-radius:20px; padding:4px 14px; font-size:12px; }}
  .pill b {{ color:#4fc3f7; }}
  .pill.new {{ border-color:#66bb6a; }}
  .pill.new b {{ color:#66bb6a; }}
  .pill.pxrf {{ border-color:#ff9800; }}
  .pill.pxrf b {{ color:#ff9800; }}

  #body {{ display:flex; flex:1; overflow:hidden; }}

  /* SIDEBAR */
  #sidebar {{ width:290px; background:#111820; border-right:1px solid #1e3040; display:flex; flex-direction:column; flex-shrink:0; z-index:500; overflow-y:auto; overflow-x:hidden; }}
  #sidebar::-webkit-scrollbar {{ width:6px; }}
  #sidebar::-webkit-scrollbar-track {{ background:#111820; }}
  #sidebar::-webkit-scrollbar-thumb {{ background:#2a3a4a; border-radius:3px; }}
  .section {{ padding:8px 12px; border-bottom:1px solid #1e3040; }}
  .section h3 {{ font-size:10px; color:#546e7a; text-transform:uppercase; letter-spacing:1px; margin-bottom:5px; }}
  label.lbl {{ font-size:10px; color:#78909c; display:block; margin-bottom:2px; }}
  select, input[type=range] {{ width:100%; }}
  select {{ background:#1a2535; border:1px solid #2a3a4a; color:#e0e0e0; padding:4px 6px; border-radius:5px; font-size:11px; }}
  input[type=range] {{ accent-color:#4fc3f7; margin:2px 0; }}
  .rval {{ display:flex; justify-content:space-between; font-size:9px; color:#546e7a; }}
  .btn {{ width:100%; padding:5px; border:none; border-radius:5px; cursor:pointer; font-size:11px; font-weight:600; margin-top:4px; }}
  .btn-p {{ background:linear-gradient(135deg,#0077b6,#00a8e8); color:#fff; }}
  .btn-p:hover {{ background:linear-gradient(135deg,#0096c7,#00c6ff); }}
  .btn-s {{ background:#1a2535; border:1px solid #2a3a4a; color:#90a4ae; }}
  .btn-s:hover {{ background:#223040; color:#e0e0e0; }}

  /* LIT FILTER */
  #lit-list {{ max-height:120px; overflow-y:auto; }}
  #lit-list::-webkit-scrollbar {{ width:5px; }}
  #lit-list::-webkit-scrollbar-thumb {{ background:#2a3a4a; border-radius:3px; }}
  .lit-row {{ display:flex; align-items:center; gap:5px; padding:1px 0; font-size:10px; cursor:pointer; }}
  .lit-row:hover {{ background:#1a2535; border-radius:3px; }}
  .dot {{ width:8px; height:8px; border-radius:50%; flex-shrink:0; }}
  .lit-row input {{ accent-color:#4fc3f7; cursor:pointer; width:13px; height:13px; }}

  /* PANEL MUESTRA */
  #info-panel {{ padding:8px 12px; border-bottom:1px solid #1e3040; }}
  #info-empty {{ color:#546e7a; font-size:11px; text-align:center; padding:12px 8px; line-height:1.5; }}
  .i-row {{ display:flex; justify-content:space-between; padding:3px 0; border-bottom:1px solid #1a2535; font-size:11px; }}
  .i-lbl {{ color:#78909c; }}
  .i-val {{ font-weight:600; color:#cfd8dc; }}
  .i-val.hi {{ color:#ff5722; }}
  .i-val.md {{ color:#ffa726; }}
  .i-val.lo {{ color:#66bb6a; }}
  .badge {{ display:inline-block; padding:2px 6px; border-radius:8px; font-size:9px; margin-top:3px; }}
  .badge-w {{ background:#2a1a0a; border:1px solid #5d4037; color:#ffa726; }}
  .badge-d {{ background:#1a1a2a; border:1px solid #3949ab; color:#7986cb; }}
  .badge-new {{ background:#0a2a0a; border:1px solid #2e7d32; color:#66bb6a; }}
  .badge-pxrf {{ background:#2a1a0a; border:1px solid #e65100; color:#ff9800; }}
  .badge-loc {{ background:#1a1a2a; border:1px solid #546e7a; color:#90a4ae; }}

  /* MAPA */
  #map-wrap {{ flex:1; position:relative; display:flex; flex-direction:column; }}
  #map {{ flex:1; z-index:1; min-height:400px; height:100%; width:100%; }}

  /* STAT BAR */
  #statbar {{ position:absolute; top:8px; right:10px; display:flex; gap:8px; z-index:800; }}
  .stat {{ background:rgba(13,27,42,0.92); border:1px solid #2a4a6a; border-radius:16px; padding:4px 12px; font-size:12px; backdrop-filter:blur(4px); }}
  .stat b {{ color:#4fc3f7; }}

  .leaflet-container {{ background:#0a0e14 !important; }}

  .layer-btn {{ display:flex; gap:4px; margin-top:5px; }}
  .lbtn {{ flex:1; padding:5px; background:#1a2535; border:1px solid #2a3a4a; color:#78909c; border-radius:5px; cursor:pointer; font-size:11px; text-align:center; }}
  .lbtn.active {{ background:#0077b6; border-color:#4fc3f7; color:#fff; }}

  #mini-chart {{ width:100%; height:45px; }}

  .ovl-row {{ display:flex; align-items:center; gap:8px; padding:3px 0; font-size:10px; }}
  .ovl-row label {{ cursor:pointer; color:#b0bec5; }}
  .toggle {{ position:relative; display:inline-block; width:34px; height:18px; flex-shrink:0; }}
  .toggle input {{ position:absolute; width:34px; height:18px; opacity:0; margin:0; cursor:pointer; z-index:2; }}
  .toggle .slider {{ position:absolute; inset:0; background:#2a3a4a; border-radius:9px; transition:.3s; pointer-events:none; }}
  .toggle .slider::before {{ content:''; position:absolute; height:14px; width:14px; left:2px; bottom:2px; background:#546e7a; border-radius:50%; transition:.3s; }}
  .toggle input:checked + .slider {{ background:#0077b6; }}
  .toggle input:checked + .slider::before {{ transform:translateX(16px); background:#4fc3f7; }}
  .ovl-desc {{ font-size:9px; color:#546e7a; margin-top:2px; }}

  .custom-tooltip {{ background:rgba(10,20,35,0.97)!important; border:1px solid #2a4a6a!important; border-radius:8px!important; color:#e0e0e0!important; font-size:12px!important; padding:8px 12px!important; box-shadow:0 4px 20px rgba(0,0,0,0.7)!important; }}

  /* Source filter tabs */
  .src-tabs {{ display:flex; gap:3px; margin-top:5px; }}
  .src-tab {{ flex:1; padding:4px 6px; background:#1a2535; border:1px solid #2a3a4a; border-radius:5px; cursor:pointer; font-size:10px; text-align:center; color:#78909c; }}
  .src-tab.active {{ color:#fff; }}
  .src-tab.all.active {{ background:#0077b6; border-color:#4fc3f7; }}
  .src-tab.old.active {{ background:#1565c0; border-color:#42a5f5; }}
  .src-tab.new.active {{ background:#2e7d32; border-color:#66bb6a; }}
  .src-tab.pxrf.active {{ background:#e65100; border-color:#ff9800; }}

  /* CP Search */
  .cp-search-wrap {{ position:relative; margin-top:6px; }}
  .cp-search-wrap input {{ width:100%; background:#1a2535; border:1px solid #2a4a6a; border-radius:6px; padding:7px 10px 7px 28px; color:#e0e0e0; font-size:12px; outline:none; transition:border-color .2s; }}
  .cp-search-wrap input:focus {{ border-color:#4fc3f7; }}
  .cp-search-wrap input::placeholder {{ color:#546e7a; }}
  .cp-search-wrap .search-icon {{ position:absolute; left:8px; top:50%; transform:translateY(-50%); font-size:12px; color:#546e7a; pointer-events:none; }}
  .cp-results {{ max-height:180px; overflow-y:auto; margin-top:4px; border-radius:6px; }}
  .cp-result-item {{ padding:5px 10px; font-size:11px; cursor:pointer; border-bottom:1px solid #1a2535; background:#0f1a28; display:flex; justify-content:space-between; align-items:center; }}
  .cp-result-item:hover {{ background:#1a3050; }}
  .cp-result-item .cp-name {{ color:#4fc3f7; font-weight:600; }}
  .cp-result-item .cp-val {{ color:#ff9800; font-size:10px; }}
  .cp-clear {{ position:absolute; right:8px; top:50%; transform:translateY(-50%); cursor:pointer; color:#78909c; font-size:14px; display:none; }}
  .cp-clear:hover {{ color:#ff7043; }}
</style>
</head>
<body>

<div id="header">
  <div>
    <h1>🇨🇱 VISOR GEOQUÍMICO — REE [INTEGRADO]</h1>
    <div class="sub">Campaña de Exploración Ytrio y Tierras Raras | BD_Ytrio + BD_GEOL_2026 + pXRF | {total:,} muestras</div>
  </div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;">
    <div class="pill">Total: <b id="stat-n">{total:,}</b></div>
    <div class="pill">Existentes: <b>{n_exist:,}</b></div>
    <div class="pill new">Nuevas: <b>{n_new:,}</b></div>
    <div class="pill pxrf">pXRF: <b>{n_pxrf:,}</b></div>
    <div class="pill">Y media: <b id="stat-mean">—</b> ppm</div>
    <div class="pill">Y máx: <b id="stat-max">—</b> ppm</div>
    <div class="pill">Anomalías: <b id="stat-anom">—</b></div>
  </div>
</div>

<div id="body">
  <div id="sidebar">

    <div class="section">
      <h3>🗺 Capa Base</h3>
      <div class="layer-btn">
        <div class="lbtn active" id="btn-osm" onclick="setLayer('osm')">Calles</div>
        <div class="lbtn" id="btn-sat" onclick="setLayer('sat')">Satélite</div>
        <div class="lbtn" id="btn-topo" onclick="setLayer('topo')">Topo</div>
        <div class="lbtn" id="btn-dark" onclick="setLayer('dark')">Oscuro</div>
      </div>
    </div>

    <div class="section">
      <h3>🧭 Capas Geológicas</h3>
      <div class="ovl-desc">Superponer información geológica sobre el mapa</div>
      <div class="ovl-row" style="margin-top:5px;">
        <label class="toggle"><input type="checkbox" id="chk-geo" onchange="toggleGeoLayer(this.checked)"><span class="slider"></span></label>
        <label for="chk-geo">🪨 Geología (Macrostrat)</label>
      </div>
      <div class="ovl-row">
        <label class="toggle"><input type="checkbox" id="chk-geo-lines" onchange="toggleGeoLines(this.checked)"><span class="slider"></span></label>
        <label for="chk-geo-lines">📐 Contactos geológicos</label>
      </div>
      <div class="ovl-row">
        <label class="lbl" style="margin-top:4px;">Opacidad capa geológica</label>
      </div>
      <input type="range" id="geo-alpha" min="0.1" max="0.9" step="0.05" value="0.5" oninput="setGeoOpacity(this.value)">
      <div class="rval"><span>10%</span><span id="geo-al-val">50%</span><span>90%</span></div>
    </div>

    <div class="section">
      <h3>🌡️ Mapa de Isoconcentraciones (IDW)</h3>
      <div class="ovl-desc">Interpolación IDW de concentraciones elementales</div>
      <div class="ovl-row" style="margin-top:5px;">
        <label class="toggle"><input type="checkbox" id="chk-idw" onchange="toggleIDW(this.checked)"><span class="slider"></span></label>
        <label for="chk-idw">🌡️ Superficie IDW</label>
      </div>
      <label class="lbl" style="margin-top:6px;">Variable IDW</label>
      <select id="idw-var" onchange="if(document.getElementById('chk-idw').checked) renderIDW()">
        <option value="REE_Global">⭐ REE Global (ponderado)</option>
        <option value="Y_ppm">Y (ppm) — Ytrio</option>
        <option value="Y_pond">Y ponderado</option>
        <option value="Ce_ppm">Ce (ppm) — Cerio</option>
        <option value="La_ppm">La (ppm) — Lantano</option>
        <option value="Th_ppm">Th (ppm) — Torio</option>
        <option value="Nd_ppm">Nd (ppm) — Neodimio</option>
        <option value="Pr_ppm">Pr (ppm) — Praseodimio</option>
        <option value="Fe__">Fe (%)</option>
      </select>
      <label class="lbl" style="margin-top:6px;">Opacidad IDW</label>
      <input type="range" id="idw-alpha" min="0.1" max="0.8" step="0.05" value="0.45" oninput="document.getElementById('idw-al-val').textContent=Math.round(this.value*100)+'%'; if(idwOverlay) idwOverlay.setOpacity(parseFloat(this.value))">
      <div class="rval"><span>10%</span><span id="idw-al-val">45%</span><span>80%</span></div>
      <label class="lbl" style="margin-top:6px;">Resolución</label>
      <select id="idw-res" onchange="if(document.getElementById('chk-idw').checked) renderIDW()">
        <option value="2">Alta (lento)</option>
        <option value="3" selected>Media</option>
        <option value="5">Baja (rápido)</option>
      </select>
      <label class="lbl" style="margin-top:6px;">Potencia (p)</label>
      <select id="idw-power" onchange="if(document.getElementById('chk-idw').checked) renderIDW()">
        <option value="1">1 (suave)</option>
        <option value="2" selected>2 (estándar)</option>
        <option value="3">3 (agudo)</option>
      </select>
      <div id="idw-status" style="font-size:9px;color:#546e7a;margin-top:4px;"></div>
    </div>

    <div class="section">
      <h3>📂 Fuente de datos</h3>
      <div class="src-tabs">
        <div class="src-tab all active" id="src-all" onclick="setSource('all')">Todas ({total})</div>
        <div class="src-tab old" id="src-old" onclick="setSource('old')">Existentes ({n_exist})</div>
        <div class="src-tab new" id="src-new" onclick="setSource('new')">Nuevas ({n_new})</div>
        <div class="src-tab pxrf" id="src-pxrf" onclick="setSource('pxrf')">pXRF ({n_pxrf})</div>
      </div>
      <div class="ovl-row" style="margin-top:6px;">
        <label class="toggle"><input type="checkbox" id="chk-loc" onchange="updateMarkers()" checked><span class="slider"></span></label>
        <label for="chk-loc">📍 Puntos sólo ubicación ({n_loc})</label>
      </div>
    </div>

    <div class="section">
      <h3>🔍 Buscar Control Point</h3>
      <div class="cp-search-wrap">
        <span class="search-icon">🔎</span>
        <input type="text" id="cp-search" placeholder="Ej: QUI-CA238, CA150..." oninput="searchCP(this.value)" autocomplete="off"/>
        <span class="cp-clear" id="cp-clear" onclick="clearCPSearch()">&times;</span>
      </div>
      <div id="cp-results" class="cp-results"></div>
      <div id="cp-status" style="font-size:10px;color:#546e7a;margin-top:3px"></div>
    </div>

    <div class="section">
      <h3>🎨 Visualización</h3>
      <label class="lbl">Variable a mostrar</label>
      <select id="var-sel" onchange="updateMarkers()">
        <option value="REE_Global">⭐ REE Global (ponderado)</option>
        <option value="Y_ppm">Y (ppm) — Ytrio</option>
        <option value="Y_pond">Y ponderado</option>
        <option value="Ce_ppm">Ce (ppm) — Cerio</option>
        <option value="Th_ppm">Th (ppm) — Torio</option>
        <option value="La_ppm">La (ppm) — Lantano</option>
        <option value="Nd_ppm">Nd (ppm) — Neodimio</option>
        <option value="Pr_ppm">Pr (ppm) — Praseodimio</option>
        <option value="Fe__">Fe (%)</option>
      </select>

      <label class="lbl" style="margin-top:8px;">Paleta de color</label>
      <select id="pal-sel" onchange="updateMarkers()">
        <option value="spectral">Spectral (científico)</option>
        <option value="hot">Hot (anomalías)</option>
        <option value="viridis">Viridis</option>
        <option value="plasma">Plasma</option>
      </select>

      <label class="lbl" style="margin-top:8px;">Tamaño de punto</label>
      <input type="range" id="pt-size" min="3" max="18" value="7" oninput="document.getElementById('pt-val').textContent=this.value; updateMarkers()">
      <div class="rval"><span>3</span><span id="pt-val">7</span><span>18</span></div>

      <label class="lbl" style="margin-top:6px;">Opacidad</label>
      <input type="range" id="pt-alpha" min="0.2" max="1" step="0.05" value="0.8" oninput="document.getElementById('al-val').textContent=this.value; updateMarkers()">
      <div class="rval"><span>0.2</span><span id="al-val">0.8</span><span>1.0</span></div>
    </div>

    <div class="section">
      <h3>🔍 Filtro Y (ppm)</h3>
      <label class="lbl">Mínimo: <span id="ymin-v">0</span></label>
      <input type="range" id="ymin-sl" min="0" max="300" value="0" oninput="document.getElementById('ymin-v').textContent=this.value; updateMarkers()">
      <label class="lbl" style="margin-top:4px;">Mostrar solo</label>
      <select id="anom-f" onchange="updateMarkers()">
        <option value="0">Todas las muestras</option>
        <option value="20">Y ≥ 20 ppm</option>
        <option value="50">Y ≥ 50 ppm (anomalía)</option>
        <option value="100">Y ≥ 100 ppm (fuerte)</option>
        <option value="200">Y ≥ 200 ppm (excepcional)</option>
      </select>
    </div>

    <div class="section">
      <h3>🪨 Litologías</h3>
      <label class="lit-row">
        <input type="checkbox" id="chk-all" checked onchange="toggleAll(this)">
        <span style="color:#90a4ae;">Todas</span>
      </label>
      <div id="lit-list"></div>
    </div>

    <div class="section">
      <h3>📊 Distribución Y visible</h3>
      <canvas id="mini-chart"></canvas>
    </div>

    <div id="info-panel">
      <div id="info-empty">👆 Haz clic en un punto para ver los datos completos de la muestra</div>
      <div id="info-detail" style="display:none"></div>
    </div>
  </div>

  <div id="map-wrap">
    <div id="map"></div>
    <div id="statbar">
      <div class="stat">Visibles: <b id="vis-n">—</b></div>
      <div class="stat">Zoom: <b id="zoom-lv">—</b></div>
    </div>
  </div>
</div>

<script>
const RAW = {json_data};
const LOC_POINTS = {json_loc};
const LIT_COLORS = {lit_colors_json};
const CENTER = [{center_lat}, {center_lon}];
const FOTOS_BY_SAMPLE = {fotos_json_str};

function openLightbox(src) {{
  const lb = document.createElement('div');
  lb.className = 'foto-lightbox';
  lb.innerHTML = '<img src="'+src+'"/>';
  lb.onclick = () => lb.remove();
  document.body.appendChild(lb);
}}

let sourceFilter = 'all'; // 'all', 'old', 'new', 'pxrf'

// ── Paletas ────────────────────────────────────────────────────
const PALS = {{
  spectral:['#313695','#4575b4','#74add1','#abd9e9','#ffffbf','#fee090','#fdae61','#f46d43','#d73027','#a50026'],
  hot:     ['#050005','#300000','#700000','#b00000','#e00000','#ff3000','#ff7000','#ffa500','#ffd000','#ffff30'],
  viridis: ['#440154','#482878','#3e4989','#31688e','#26828e','#1f9e89','#35b779','#6ece58','#b5de2b','#fde725'],
  plasma:  ['#0d0887','#46039f','#7201a8','#9c179e','#bd3786','#d8576b','#ed7953','#fb9f3a','#fdcf18','#f0f921'],
}};

function getColor(val, mn, mx, pal) {{
  const t = Math.max(0, Math.min(1, (val-mn)/(mx-mn)));
  const p = PALS[pal], n = p.length-1;
  const i = Math.min(Math.floor(t*n), n-1), f = t*n-i;
  return lerpCol(p[i], p[i+1]||p[n], f);
}}
function lerpCol(a,b,t) {{
  const ah=parseInt(a.slice(1),16), bh=parseInt(b.slice(1),16);
  const r=Math.round(((ah>>16)&255)+(((bh>>16)&255)-((ah>>16)&255))*t);
  const g=Math.round(((ah>>8)&255)+(((bh>>8)&255)-((ah>>8)&255))*t);
  const bl=Math.round((ah&255)+((bh&255)-(ah&255))*t);
  return `#${{r.toString(16).padStart(2,'0')}}${{g.toString(16).padStart(2,'0')}}${{bl.toString(16).padStart(2,'0')}}`;
}}

// ── Leaflet ────────────────────────────────────────────────────
const LAYERS = {{
  osm:  L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{maxZoom:19, attribution:'© OpenStreetMap'}}),
  sat:  L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}', {{maxZoom:19, attribution:'© Esri'}}),
  topo: L.tileLayer('https://{{s}}.tile.opentopomap.org/{{z}}/{{x}}/{{y}}.png', {{maxZoom:17, attribution:'© OpenTopoMap'}}),
  dark: L.tileLayer('https://{{s}}.basemaps.cartocdn.com/dark_all/{{z}}/{{x}}/{{y}}{{r}}.png', {{maxZoom:19, attribution:'© CartoDB'}}),
}};

const map = L.map('map', {{ center:CENTER, zoom:9, zoomControl:true }});
let currentLayer = LAYERS.osm.addTo(map);

function setLayer(key) {{
  map.removeLayer(currentLayer);
  currentLayer = LAYERS[key].addTo(map);
  document.querySelectorAll('.lbtn').forEach(b=>b.classList.remove('active'));
  document.getElementById('btn-'+key).classList.add('active');
  if(geoLayer && map.hasLayer(geoLayer)) geoLayer.bringToFront();
  if(geoLinesLayer && map.hasLayer(geoLinesLayer)) geoLinesLayer.bringToFront();
}}

map.on('zoomend', () => document.getElementById('zoom-lv').textContent = map.getZoom());
document.getElementById('zoom-lv').textContent = map.getZoom();

// ── Source filter ──────────────────────────────────────────────
function setSource(src) {{
  sourceFilter = src;
  document.querySelectorAll('.src-tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('src-'+src).classList.add('active');
  updateMarkers();
}}

// ── Capas geológicas ──────────────────────────────────────────
let geoLayer = null;
let geoLinesLayer = null;
map.createPane('geology');
map.getPane('geology').style.zIndex = 350;

function toggleGeoLayer(on) {{
  if(on) {{
    if(!geoLayer) {{
      geoLayer = L.tileLayer('https://tiles.macrostrat.org/carto/{{z}}/{{x}}/{{y}}.png', {{
        maxZoom:19, opacity: parseFloat(document.getElementById('geo-alpha').value),
        attribution:'© Macrostrat', pane:'geology', errorTileUrl:''
      }});
    }}
    geoLayer.addTo(map);
  }} else if(geoLayer) map.removeLayer(geoLayer);
}}

function toggleGeoLines(on) {{
  if(on) {{
    if(!geoLinesLayer) {{
      geoLinesLayer = L.tileLayer.wms('https://mrdata.usgs.gov/services/sgmc2', {{
        layers:'SGMCContactsAndFaults', format:'image/png', transparent:true,
        opacity:parseFloat(document.getElementById('geo-alpha').value),
        attribution:'© USGS', pane:'geology'
      }});
    }}
    geoLinesLayer.addTo(map);
  }} else if(geoLinesLayer) map.removeLayer(geoLinesLayer);
}}

function setGeoOpacity(val) {{
  document.getElementById('geo-al-val').textContent = Math.round(val*100)+'%';
  if(geoLayer && map.hasLayer(geoLayer)) geoLayer.setOpacity(parseFloat(val));
  if(geoLinesLayer && map.hasLayer(geoLinesLayer)) geoLinesLayer.setOpacity(parseFloat(val));
}}

// ── Litologias checkboxes ──────────────────────────────────────
const litList = [...new Set(RAW.map(r=>r.Litology_STD))].sort();
let visLits = new Set(litList);
const litBox = document.getElementById('lit-list');
litList.forEach(l => {{
  const d = document.createElement('label');
  d.className = 'lit-row';
  d.innerHTML = `<input type="checkbox" class="lchk" value="${{l}}" checked>
    <div class="dot" style="background:${{LIT_COLORS[l]||'#78909c'}}"></div>
    <span title="${{l}}">${{l.length>26?l.slice(0,24)+'…':l}}</span>`;
  d.querySelector('input').addEventListener('change',e=>{{
    if(e.target.checked) visLits.add(l); else visLits.delete(l);
    document.getElementById('chk-all').indeterminate=true;
    updateMarkers();
  }});
  litBox.appendChild(d);
}});

function toggleAll(chk) {{
  document.querySelectorAll('.lchk').forEach(c=>c.checked=chk.checked);
  if(chk.checked) litList.forEach(l=>visLits.add(l)); else visLits.clear();
  updateMarkers();
}}

// ── Markers ────────────────────────────────────────────────────
let layer = L.layerGroup().addTo(map);
let locLayer = L.layerGroup().addTo(map);

let cpFilter = '';

function getFiltered() {{
  const v = document.getElementById('var-sel').value;
  const ymin = parseFloat(document.getElementById('ymin-sl').value)||0;
  const af = parseFloat(document.getElementById('anom-f').value)||0;
  return RAW.filter(r => {{
    if(sourceFilter === 'old' && r.FUENTE !== 'BD_Ytrio') return false;
    if(sourceFilter === 'new' && r.FUENTE !== 'BD_GEOL_2026') return false;
    if(sourceFilter === 'pxrf' && r.FUENTE !== 'pXRF_2026') return false;
    const yppm = typeof r.Y_ppm === 'number' ? r.Y_ppm : 0;
    return yppm>=ymin && yppm>=af && visLits.has(r.Litology_STD) && r[v]!=='';
  }});
}}

function searchCP(q) {{
  const clear = document.getElementById('cp-clear');
  const results = document.getElementById('cp-results');
  const status = document.getElementById('cp-status');
  q = q.trim();
  clear.style.display = q ? 'block' : 'none';
  
  if(!q || q.length < 1) {{
    results.innerHTML = '';
    status.textContent = '';
    return;
  }}
  
  const qu = q.toUpperCase();
  
  // Encontrar CPs únicos que coincidan — agrupar muestras por CP
  const cpMap = new Map();
  RAW.forEach(r => {{
    const cp = r.CP || '';
    if(!cp) return;
    if(!cp.toUpperCase().includes(qu)) return;
    if(!cpMap.has(cp)) {{
      cpMap.set(cp, {{
        cp: cp,
        samples: [],
        maxY: -Infinity,
        maxCe: -Infinity,
        lit: r.Litology_STD || '',
        fuente: r.FUENTE || '',
        lat: r.lat,
        lon: r.lon,
        repr: r
      }});
    }}
    const entry = cpMap.get(cp);
    entry.samples.push(r);
    if(typeof r.Y_ppm === 'number' && r.Y_ppm > entry.maxY) {{
      entry.maxY = r.Y_ppm;
      entry.repr = r;
    }}
    if(typeof r.Ce_ppm === 'number' && r.Ce_ppm > entry.maxCe) entry.maxCe = r.Ce_ppm;
  }});
  
  const matches = Array.from(cpMap.values()).sort((a,b) => a.cp.localeCompare(b.cp));
  
  if(matches.length === 0) {{
    results.innerHTML = '<div style=\"padding:6px;font-size:10px;color:#78909c\">Sin resultados para \"' + q + '\"</div>';
    status.textContent = '';
  }} else {{
    let html = '';
    matches.slice(0, 30).forEach(m => {{
      const y = m.maxY > -Infinity ? m.maxY.toFixed(0) : '\u2014';
      const ce = m.maxCe > -Infinity ? m.maxCe.toFixed(0) : '\u2014';
      const nSamples = m.samples.length;
      const badge = m.fuente === 'pXRF_2026' ? '<span style=\"color:#ff9800\">&#9679;</span>' : (m.fuente === 'BD_GEOL_2026' ? '<span style=\"color:#66bb6a\">&#9679;</span>' : '<span style=\"color:#42a5f5\">&#9679;</span>');
      html += '<div class=\"cp-result-item\" onclick=\"flyToCP(\\'' + m.cp.replace(/'/g,"\\\\'") + '\\')\">'+
        '<div style=\"flex:1\"><span class=\"cp-name\">' + badge + ' ' + m.cp + '</span>'+
        '<div style=\"font-size:9px;color:#78909c;margin-top:1px\">' + m.lit + ' \u00b7 ' + nSamples + ' muestra' + (nSamples>1?'s':'') + '</div></div>'+
        '<div style=\"text-align:right\"><span class=\"cp-val\">Y: ' + y + '</span>'+
        '<div style=\"font-size:9px;color:#b0bec5\">Ce: ' + ce + '</div></div>'+
        '</div>';
    }});
    results.innerHTML = html;
    status.textContent = matches.length + (matches.length > 30 ? '+ ' : ' ') + 'CP encontrados';
  }}
}}

function flyToCP(cp) {{
  const pts = RAW.filter(r => r.CP === cp);
  if(pts.length > 0) {{
    // Encontrar el de mayor Y
    let best = pts[0];
    pts.forEach(r => {{ if(typeof r.Y_ppm === 'number' && r.Y_ppm > (best.Y_ppm||0)) best = r; }});
    map.flyTo([best.lat, best.lon], 16, {{duration:1}});
    setTimeout(() => showDetail(best), 500);
    // Limpiar input y resultados
    document.getElementById('cp-search').value = cp;
    document.getElementById('cp-results').innerHTML = '';
    document.getElementById('cp-status').textContent = '\u2705 ' + cp + ' \u2014 ' + pts.length + ' muestra(s)';
  }}
}}

function clearCPSearch() {{
  document.getElementById('cp-search').value = '';
  document.getElementById('cp-clear').style.display = 'none';
  document.getElementById('cp-results').innerHTML = '';
  document.getElementById('cp-status').textContent = '';
}}

function calcP95(arr) {{
  const s=[...arr].sort((a,b)=>a-b);
  return s[Math.floor(s.length*0.95)]||s[s.length-1];
}}

// Mapa de nombres legibles para variables
const VAR_LABELS = {{
  'REE_Global': 'REE Global', 'Y_ppm': 'Y (ppm)', 'Y_pond': 'Y pond.',
  'Ce_ppm': 'Ce (ppm)', 'La_ppm': 'La (ppm)', 'Th_ppm': 'Th (ppm)',
  'Nd_ppm': 'Nd (ppm)', 'Pr_ppm': 'Pr (ppm)', 'Fe__': 'Fe (%)', 'Ti__': 'Ti (%)'
}};
const VAR_UNITS = {{
  'REE_Global': '', 'Y_ppm': 'ppm', 'Y_pond': '', 'Ce_ppm': 'ppm',
  'La_ppm': 'ppm', 'Th_ppm': 'ppm', 'Nd_ppm': 'ppm', 'Pr_ppm': 'ppm',
  'Fe__': '%', 'Ti__': '%'
}};

function buildTooltip(r, varKey) {{
  const lbl = VAR_LABELS[varKey] || varKey;
  const unit = VAR_UNITS[varKey] || '';
  const val = typeof r[varKey]==='number' ? r[varKey].toFixed(1) + (unit ? ' '+unit : '') : '—';
  const isNew = r.FUENTE === 'BD_GEOL_2026';
  const isPxrf = r.FUENTE === 'pXRF_2026';
  const srcBadge = isPxrf ? '<span style="color:#ff9800;font-size:9px">📊 pXRF</span>' : (isNew ? '<span style="color:#66bb6a;font-size:9px">🆕 GEOL_2026</span>' : '');
  return `
    <div style="font-weight:700;color:#4fc3f7;font-size:13px;margin-bottom:4px">📍 ${{r.Sample}}</div>
    ${{r.CP ? '<div style="color:#b0bec5;font-size:11px;margin-bottom:3px">🏠 CP: <b>'+r.CP+'</b></div>' : ''}}
    ${{srcBadge}}
    <div style="margin:4px 0;padding:4px 6px;background:rgba(79,195,247,0.12);border-radius:4px;border-left:3px solid #4fc3f7">
      <b style="color:#4fc3f7">${{lbl}}:</b> <span style="color:#ff7043;font-weight:700;font-size:13px">${{val}}</span>
    </div>
    <div><b>Litología:</b> ${{r.Litology_STD}}</div>
    ${{r.HORIZONTE ? '<div><b>Horizonte:</b> '+r.HORIZONTE+'</div>' : ''}}
    <div style="font-size:10px;color:#78909c;margin-top:3px">Clic para más detalles</div>
  `;
}}

function updateMarkers() {{
  layer.clearLayers();
  locLayer.clearLayers();
  
  const data = getFiltered();
  const v = document.getElementById('var-sel').value;
  const pal = document.getElementById('pal-sel').value;
  const sz = parseInt(document.getElementById('pt-size').value);
  const alpha = parseFloat(document.getElementById('pt-alpha').value);
  
  const vals = data.map(r=>r[v]).filter(x=>typeof x==='number'&&x>0);
  const mn = 0;
  const mx = Math.min(Math.max(...vals), calcP95(vals));

  data.forEach(r => {{
    const val = r[v];
    if(typeof val !== 'number') return;
    const col = getColor(val, mn, mx, pal);
    const isNew = r.FUENTE === 'BD_GEOL_2026';
    const isPxrf = r.FUENTE === 'pXRF_2026';
    const yppm = typeof r.Y_ppm === 'number' ? r.Y_ppm : 0;
    const m = L.circleMarker([r.lat, r.lon], {{
      radius: (isNew || isPxrf) ? sz+1 : sz,
      fillColor: col,
      color: isPxrf ? '#ff9800' : (isNew ? '#66bb6a' : (yppm>=100 ? '#ffffff' : col)),
      weight: isPxrf ? 2 : (isNew ? 2 : (yppm>=100 ? 1.5 : 0.3)),
      fillOpacity: alpha,
      opacity: 1
    }});

    const srcBadge = isPxrf ? '<span style="color:#ff9800;font-size:9px">📊 pXRF 2026</span>' : (isNew ? '<span style="color:#66bb6a;font-size:9px">🆕 BD_GEOL_2026</span>' : '');
    m.bindTooltip(buildTooltip(r, v), {{className:'custom-tooltip', sticky:true}});

    m.on('click', ()=>showDetail(r));
    layer.addLayer(m);
  }});

  // Location-only points
  const showLoc = document.getElementById('chk-loc').checked;
  if(showLoc && (sourceFilter === 'all' || sourceFilter === 'new' || sourceFilter === 'pxrf')) {{
    LOC_POINTS.forEach(r => {{
      const m = L.circleMarker([r.lat, r.lon], {{
        radius: 5, fillColor:'#546e7a', color:'#78909c',
        weight:1.5, fillOpacity:0.6, opacity:1
      }});
      m.bindTooltip(`
        <div style="font-weight:700;color:#90a4ae;font-size:12px">📍 ${{r.Sample || r.CP}}</div>
        <div style="color:#546e7a;font-size:10px">Sólo ubicación (sin geoquímica)</div>
        <div><b>Elev:</b> ${{r.COTA_M || '—'}} m</div>
      `, {{className:'custom-tooltip', sticky:true}});
      m.on('click', ()=>{{
        document.getElementById('info-empty').style.display='none';
        const det = document.getElementById('info-detail');
        det.style.display='block';
        det.innerHTML = `
          <div style="font-size:13px;font-weight:700;color:#90a4ae;margin-bottom:8px">📍 ${{r.Sample || r.CP}}</div>
          <span class="badge badge-loc">📍 Sólo ubicación</span>
          <span class="badge badge-new">🆕 BD_GEOL_2026</span>
          <div class="i-row" style="margin-top:8px"><span class="i-lbl">Lat / Lon</span><span class="i-val" style="font-size:10px">${{r.lat.toFixed(5)}} / ${{r.lon.toFixed(5)}}</span></div>
          <div class="i-row"><span class="i-lbl">Elevación</span><span class="i-val">${{r.COTA_M || '—'}} m</span></div>
          <button class="btn btn-p" style="margin-top:10px" onclick="map.flyTo([${{r.lat}},${{r.lon}}],15,{{duration:1}})">📍 Centrar en mapa</button>
        `;
      }});
      locLayer.addLayer(m);
    }});
  }}

  // Stats
  document.getElementById('vis-n').textContent = (data.length + (showLoc ? LOC_POINTS.length : 0)).toLocaleString();
  document.getElementById('stat-n').textContent = data.length.toLocaleString();
  const sv = data.map(r=>r[v]).filter(x=>typeof x==='number');
  if(sv.length) {{
    document.getElementById('stat-mean').textContent = (sv.reduce((a,b)=>a+b,0)/sv.length).toFixed(1);
    document.getElementById('stat-max').textContent = Math.max(...sv).toFixed(1);
    const anomThresh = (v==='REE_Global') ? 150 : (v==='Fe__' ? 5 : 50);
    document.getElementById('stat-anom').textContent = sv.filter(x=>x>=anomThresh).length.toLocaleString();
  }}

  drawMiniChart(data);
  updateLegend(mn, mx, pal);
}}

function updateLegend(mn, mx, pal) {{
  const p = PALS[pal];
  const bar = document.getElementById('legend-bar');
  if(bar) bar.style.background = `linear-gradient(to right,${{p.join(',')}})`;
  const v = document.getElementById('var-sel').value;
  const titleEl = document.getElementById('legend-title');
  if(titleEl) titleEl.textContent = 'Escala ' + (VAR_LABELS[v] || v);
  const rangeEl = document.getElementById('legend-range');
  if(rangeEl) rangeEl.innerHTML = `<span>${{mn.toFixed(0)}}</span><span>${{(mx/2).toFixed(0)}}</span><span>${{mx.toFixed(0)}}+</span>`;
}}

// ── Mini histograma ────────────────────────────────────────────
function drawMiniChart(data) {{
  const canvas = document.getElementById('mini-chart');
  if(!canvas) return;
  canvas.width = canvas.offsetWidth || 240;
  canvas.height = 55;
  const ctx = canvas.getContext('2d');
  ctx.clearRect(0,0,canvas.width,canvas.height);

  const curVar = document.getElementById('var-sel').value;
  const vals = data.map(r=>r[curVar]).filter(v=>typeof v==='number'&&v>0);
  if(!vals.length) return;
  const p95 = [...vals].sort((a,b)=>a-b)[Math.floor(vals.length*0.95)] || Math.max(...vals);
  const cap = p95 * 1.2;
  const capped = vals.map(v=>Math.min(v, cap));
  const bins = 25, max_v = Math.max(...capped);
  const bw = max_v/bins;
  const hist = new Array(bins).fill(0);
  capped.forEach(v => {{ const i=Math.min(Math.floor(v/bw),bins-1); hist[i]++; }});
  const maxH = Math.max(...hist);
  const w = canvas.width/bins, h = canvas.height-14;

  hist.forEach((c,i) => {{
    const x = i*w, bh = (c/maxH)*h;
    const col = getColor(i*bw, 0, max_v, document.getElementById('pal-sel').value);
    ctx.fillStyle = col;
    ctx.fillRect(x+1, h-bh+2, w-2, bh);
  }});
  const unit = VAR_UNITS[curVar] || '';
  ctx.fillStyle='#546e7a'; ctx.font='9px monospace';
  ctx.fillText('0', 2, canvas.height-1);
  ctx.fillText(Math.round(max_v)+(unit||''), canvas.width-40, canvas.height-1);
}}

// ── Detalle muestra ────────────────────────────────────────────
function showDetail(r) {{
  document.getElementById('info-empty').style.display='none';
  const det = document.getElementById('info-detail');
  det.style.display='block';
  const isNew = r.FUENTE === 'BD_GEOL_2026';
  const isPxrf = r.FUENTE === 'pXRF_2026';
  const yc = (typeof r.Y_ppm==='number') ? (r.Y_ppm>100?'hi':r.Y_ppm>50?'md':'lo') : '';
  const fmt = (v,d=2) => typeof v==='number'?v.toFixed(d):'—';
  det.innerHTML = `
    <div style="font-size:13px;font-weight:700;color:#4fc3f7;margin-bottom:4px">📍 ${{r.Sample}}</div>
    ${{r.CP ? '<div style="font-size:12px;color:#b0bec5;margin-bottom:6px">🏠 Control Point: <b style="color:#4fc3f7">'+r.CP+'</b></div>' : ''}}
    ${{isPxrf ? '<span class="badge badge-pxrf">📊 pXRF 2026</span>' : (isNew ? '<span class="badge badge-new">🆕 BD_GEOL_2026</span>' : '<span class="badge badge-d">📋 BD_Ytrio</span>')}}&nbsp;
    <div class="i-row" style="margin-top:6px"><span class="i-lbl">Litología</span><span class="i-val" style="font-size:10px;color:${{LIT_COLORS[r.Litology_STD]||'#78909c'}}">${{r.Litology_STD}}</span></div>
    ${{r.ROCA_CAJA ? '<div class="i-row"><span class="i-lbl">Roca caja</span><span class="i-val" style="font-size:10px">'+r.ROCA_CAJA+'</span></div>' : ''}}
    ${{r.HORIZONTE ? '<div class="i-row"><span class="i-lbl">Horizonte</span><span class="i-val" style="font-size:10px">'+r.HORIZONTE+'</span></div>' : ''}}
    <div class="i-row"><span class="i-lbl">Lat / Lon</span><span class="i-val" style="font-size:10px">${{r.lat.toFixed(5)}} / ${{r.lon.toFixed(5)}}</span></div>
    <div class="i-row"><span class="i-lbl">UTM E / N</span><span class="i-val" style="font-size:10px">${{typeof r.UTM_E==='number'?r.UTM_E.toLocaleString():'—'}} / ${{typeof r.UTM_N==='number'?r.UTM_N.toLocaleString():'—'}}</span></div>
    <div class="i-row"><span class="i-lbl">Cota (m)</span><span class="i-val">${{r.COTA_M||'—'}}</span></div>
    <div style="margin:6px 0 3px;font-size:10px;color:#546e7a;text-transform:uppercase;letter-spacing:1px">── Tierras Raras ──</div>
    <div class="i-row" style="background:rgba(79,195,247,0.1);border-radius:4px;padding:2px 4px;border-left:3px solid #4fc3f7"><span class="i-lbl" style="color:#4fc3f7">⭐ REE Global</span><span class="i-val" style="color:#ff7043;font-weight:700">${{fmt(r.REE_Global,1)}}</span></div>
    <div class="i-row"><span class="i-lbl">Y (ppm)</span><span class="i-val ${{yc}}">${{fmt(r.Y_ppm)}}</span></div>
    <div class="i-row"><span class="i-lbl">Y pond.</span><span class="i-val">${{fmt(r.Y_pond,3)}}</span></div>
    <div class="i-row"><span class="i-lbl">Ce (ppm)</span><span class="i-val">${{fmt(r.Ce_ppm)}}</span></div>
    <div class="i-row"><span class="i-lbl">La (ppm)</span><span class="i-val">${{fmt(r.La_ppm)}}</span></div>
    <div class="i-row"><span class="i-lbl">Th (ppm)</span><span class="i-val">${{fmt(r.Th_ppm)}}</span></div>
    <div class="i-row"><span class="i-lbl">Nd (ppm)</span><span class="i-val">${{fmt(r.Nd_ppm)}}</span></div>
    <div class="i-row"><span class="i-lbl">Pr (ppm)</span><span class="i-val">${{fmt(r.Pr_ppm)}}</span></div>
    <div style="margin:6px 0 3px;font-size:10px;color:#546e7a;text-transform:uppercase;letter-spacing:1px">── Elementos Mayores ──</div>
    <div class="i-row"><span class="i-lbl">Fe (%)</span><span class="i-val">${{fmt(r.Fe__,4)}}</span></div>
    <div class="i-row"><span class="i-lbl">Ti (%)</span><span class="i-val">${{fmt(r.Ti__,4)}}</span></div>
    ${{r.FLAG_OUTLIER?`<span class="badge badge-w">⚠ ${{r.FLAG_OUTLIER}}</span>`:''}}
    ${{r.FLAG_DUPLICADO?`<span class="badge badge-d">🔁 Duplicado</span>`:''}}
    <button class="btn btn-p" style="margin-top:10px" onclick="map.flyTo([${{r.lat}},${{r.lon}}],15,{{duration:1}})">📍 Centrar en mapa</button>
  `;
  
  // Agregar fotos si existen
  const sampleBase = r.Sample ? r.Sample.replace('pXRF-','') : '';
  const fotos = FOTOS_BY_SAMPLE[sampleBase] || [];
  if(fotos.length > 0) {{
    let html = '<div style="margin:8px 0 3px;font-size:10px;color:#546e7a;text-transform:uppercase;letter-spacing:1px">── Fotos de campo (' + fotos.length + ') ──</div>';
    html += '<div class="foto-gallery">';
    fotos.forEach(f => {{
      html += '<img src="fotos/' + f + '" alt="' + f + '" onclick="openLightbox(this.src)" loading="lazy"/>';
    }});
    html += '</div>';
    det.innerHTML += html;
  }}
  map.flyTo([r.lat, r.lon], Math.max(map.getZoom(), 13), {{duration:1}});
}}

// ── Leyenda flotante ───────────────────────────────────────────
const legendCtrl = L.control({{position:'bottomright'}});
legendCtrl.onAdd = () => {{
  const d = L.DomUtil.create('div','');
  d.style.cssText='background:rgba(13,27,42,0.92);border:1px solid #2a4a6a;border-radius:8px;padding:10px 14px;min-width:200px;backdrop-filter:blur(4px)';
  d.innerHTML=`
    <div id="legend-title" style="font-size:11px;color:#78909c;text-transform:uppercase;margin-bottom:5px">Escala REE Global</div>
    <div id="legend-bar" style="height:12px;border-radius:4px;background:linear-gradient(to right,#313695,#74add1,#ffffbf,#fdae61,#a50026)"></div>
    <div id="legend-range" style="display:flex;justify-content:space-between;font-size:10px;color:#78909c;margin-top:3px"><span>Min</span><span>Max</span></div>
    <div style="margin-top:8px;font-size:10px;color:#546e7a">
      <div style="color:#66bb6a">🟢 Datos nuevos (borde verde)</div>
      <div style="color:#ff9800">🟠 Datos pXRF (borde naranja)</div>
      <div style="color:#78909c">⚫ Sólo ubicación (gris)</div>
    </div>
  `;
  return d;
}};
legendCtrl.addTo(map);

// ── Zoom to fit all data ──────────────────────────────────────
function zoomToAll() {{
  const pts = RAW.map(r=>[r.lat,r.lon]).concat(LOC_POINTS.map(r=>[r.lat,r.lon]));
  if(pts.length) map.fitBounds(pts, {{padding:[30,30]}});
}}

// ── IDW (Inverse Distance Weighting) — OPTIMIZADO ────────────
let idwOverlay = null;
let idwBusy = false;
map.createPane('idw');
map.getPane('idw').style.zIndex = 340;

function toggleIDW(on) {{
  if(on) {{
    renderIDW();
  }} else if(idwOverlay) {{
    map.removeLayer(idwOverlay);
    idwOverlay = null;
    document.getElementById('idw-status').textContent = '';
  }}
}}

function renderIDW() {{
  if(idwBusy) return;
  const statusEl = document.getElementById('idw-status');
  statusEl.textContent = '⏳ Calculando...';
  
  if(idwOverlay) {{ map.removeLayer(idwOverlay); idwOverlay = null; }}
  
  const varName = document.getElementById('idw-var').value;
  const step = parseInt(document.getElementById('idw-res').value);
  const power = parseFloat(document.getElementById('idw-power').value);
  const opacity = parseFloat(document.getElementById('idw-alpha').value);
  
  // Filtrar datos con valor numérico y dentro del viewport + buffer
  const bounds = map.getBounds();
  const latBuf = (bounds.getNorth() - bounds.getSouth()) * 0.3;
  const lonBuf = (bounds.getEast() - bounds.getWest()) * 0.3;
  
  let pts = RAW.filter(r => {{
    if(typeof r[varName] !== 'number' || r[varName] <= 0) return false;
    return r.lat >= bounds.getSouth()-latBuf && r.lat <= bounds.getNorth()+latBuf
        && r.lon >= bounds.getWest()-lonBuf && r.lon <= bounds.getEast()+lonBuf;
  }});
  
  if(pts.length < 3) {{ statusEl.textContent = 'Zoom: pocos datos visibles'; return; }}
  
  // Subsamplear si son muchos (max 400)
  const MAX_PTS = 400;
  if(pts.length > MAX_PTS) {{
    pts.sort((a,b) => b[varName] - a[varName]);
    // Mantener top anomalías + muestra aleatoria del resto
    const top = pts.slice(0, Math.floor(MAX_PTS * 0.3));
    const rest = pts.slice(Math.floor(MAX_PTS * 0.3));
    const sampleN = MAX_PTS - top.length;
    for(let i = rest.length - 1; i > 0; i--) {{
      const j = Math.floor(Math.random() * (i+1));
      [rest[i], rest[j]] = [rest[j], rest[i]];
    }}
    pts = top.concat(rest.slice(0, sampleN));
  }}
  
  const vals = pts.map(r => r[varName]);
  const vMax = Math.min(Math.max(...vals), percentile95(vals));
  
  // Canvas
  const mapSize = map.getSize();
  const cw = Math.ceil(mapSize.x / step);
  const ch = Math.ceil(mapSize.y / step);
  const canvas = document.createElement('canvas');
  canvas.width = cw;
  canvas.height = ch;
  const ctx = canvas.getContext('2d');
  const imgData = ctx.createImageData(cw, ch);
  
  // Pre-calcular posiciones en píxeles
  const ptPix = pts.map(p => {{
    const px = map.latLngToContainerPoint([p.lat, p.lon]);
    return {{ x: px.x/step, y: px.y/step, v: p[varName] }};
  }});
  
  // Grid espacial para búsqueda rápida de vecinos
  const CELL = 20;
  const gridW = Math.ceil(cw / CELL) + 1;
  const gridH = Math.ceil(ch / CELL) + 1;
  const grid = new Array(gridW * gridH);
  for(let i = 0; i < grid.length; i++) grid[i] = [];
  
  ptPix.forEach((p, i) => {{
    const gx = Math.max(0, Math.min(gridW-1, Math.floor(p.x / CELL)));
    const gy = Math.max(0, Math.min(gridH-1, Math.floor(p.y / CELL)));
    grid[gy * gridW + gx].push(i);
  }});
  
  const K = Math.min(8, ptPix.length);
  idwBusy = true;
  
  // Procesar en chunks de filas
  let rowIdx = 0;
  const ROWS_PER_CHUNK = 10;
  
  function processChunk() {{
    const endRow = Math.min(rowIdx + ROWS_PER_CHUNK, ch);
    
    for(let py = rowIdx; py < endRow; py++) {{
      for(let px = 0; px < cw; px++) {{
        // Buscar vecinos en celdas cercanas (expandir si necesario)
        const gx0 = Math.floor(px / CELL);
        const gy0 = Math.floor(py / CELL);
        
        let candidates = [];
        for(let radius = 0; radius <= 3 && candidates.length < K; radius++) {{
          candidates = [];
          for(let dy = -radius; dy <= radius; dy++) {{
            for(let dx = -radius; dx <= radius; dx++) {{
              const gxi = gx0 + dx;
              const gyi = gy0 + dy;
              if(gxi >= 0 && gxi < gridW && gyi >= 0 && gyi < gridH) {{
                const cell = grid[gyi * gridW + gxi];
                for(const idx of cell) {{
                  const ddx = px - ptPix[idx].x;
                  const ddy = py - ptPix[idx].y;
                  candidates.push({{ d: Math.sqrt(ddx*ddx + ddy*ddy), v: ptPix[idx].v }});
                }}
              }}
            }}
          }}
        }}
        
        // Si aún no hay suficientes, fallback a todos (raro)
        if(candidates.length < 1) {{
          const pidx = (py * cw + px) * 4;
          imgData.data[pidx+3] = 0;
          continue;
        }}
        
        candidates.sort((a,b) => a.d - b.d);
        const nearest = candidates.slice(0, K);
        
        let wSum = 0, vSum = 0;
        for(const n of nearest) {{
          if(n.d < 0.3) {{ vSum = n.v; wSum = 1; break; }}
          const w = 1 / Math.pow(n.d, power);
          wSum += w;
          vSum += w * n.v;
        }}
        
        const val = wSum > 0 ? vSum / wSum : 0;
        const t = Math.max(0, Math.min(1, val / vMax));
        
        const rgb = spectralRGB(t);
        const pidx = (py * cw + px) * 4;
        imgData.data[pidx] = rgb[0];
        imgData.data[pidx+1] = rgb[1];
        imgData.data[pidx+2] = rgb[2];
        imgData.data[pidx+3] = t < 0.05 ? 0 : Math.round(180 * (0.2 + 0.8*t));
      }}
    }}
    
    rowIdx = endRow;
    statusEl.textContent = `⏳ ${{Math.round(rowIdx/ch*100)}}%`;
    
    if(rowIdx < ch) {{
      setTimeout(processChunk, 0);
    }} else {{
      ctx.putImageData(imgData, 0, 0);
      idwOverlay = L.imageOverlay(canvas.toDataURL(), bounds, {{
        opacity: opacity, pane: 'idw', interactive: false
      }}).addTo(map);
      statusEl.textContent = `✅ ${{pts.length}} pts, ${{cw}}×${{ch}}px, p=${{power}}`;
      idwBusy = false;
    }}
  }}
  
  setTimeout(processChunk, 30);
}}

function percentile95(arr) {{
  const s = [...arr].sort((a,b) => a-b);
  return s[Math.min(Math.floor(s.length * 0.95), s.length-1)];
}}

function spectralRGB(t) {{
  const S = [
    [0.00, 49, 54, 149],
    [0.12, 69, 117, 180],
    [0.25, 116, 173, 209],
    [0.38, 171, 217, 233],
    [0.50, 255, 255, 191],
    [0.62, 254, 224, 144],
    [0.75, 253, 174, 97],
    [0.85, 244, 109, 67],
    [0.92, 215, 48, 39],
    [1.00, 165, 0, 38],
  ];
  let i = 0;
  for(; i < S.length-1; i++) if(t <= S[i+1][0]) break;
  const s0 = S[i], s1 = S[Math.min(i+1, S.length-1)];
  const f = s1[0]===s0[0] ? 0 : (t-s0[0])/(s1[0]-s0[0]);
  return [
    Math.round(s0[1]+(s1[1]-s0[1])*f),
    Math.round(s0[2]+(s1[2]-s0[2])*f),
    Math.round(s0[3]+(s1[3]-s0[3])*f),
  ];
}}

// Recalcular al mover mapa (con debounce)
let idwTimer = null;
map.on('moveend', () => {{
  if(document.getElementById('chk-idw').checked) {{
    clearTimeout(idwTimer);
    idwTimer = setTimeout(renderIDW, 600);
  }}
}});

// ── Init ───────────────────────────────────────────────────────
setTimeout(() => {{ map.invalidateSize(); updateMarkers(); zoomToAll(); }}, 300);
</script>
</body>
</html>"""

outpath = os.path.join(OUTDIR, "visor_mapa_chile.html")
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"\n✅ Visor generado: {outpath}")
print(f"📊 Tamaño: {os.path.getsize(outpath)/1024:.0f} KB")
print(f"🗺️  Total puntos en mapa: {total} geoquímicos ({n_exist} exist + {n_new} nuevas + {n_pxrf} pXRF) + {n_loc} ubicación")
